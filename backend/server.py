from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Cookie, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import math
import base64
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional
import uuid
from uuid import uuid4
from datetime import datetime, timezone, timedelta
import pytz
import bcrypt
import jwt
import io
import shutil
import mimetypes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Import accounting router
from accounting_routes import accounting_router

# Zona horaria de Puerto Rico (GMT-4)
PUERTO_RICO_TZ = pytz.timezone('America/Puerto_Rico')

# Función para calcular distancia entre dos coordenadas (Haversine)
def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Calcula la distancia en metros entre dos puntos geográficos
    usando la fórmula de Haversine
    """
    R = 6371000  # Radio de la Tierra en metros
    
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c  # Distancia en metros
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import requests
from enum import Enum
from email_service import send_email, get_task_assigned_email, get_task_completed_email, get_comment_email, get_welcome_email, get_task_reminder_email, get_task_status_changed_email

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']

# Configure MongoDB client with robust connection settings for Atlas/Kubernetes
# Key settings for cloud deployments:
# - directConnection=False allows replica set discovery
# - tlsAllowInvalidCertificates helps with some Atlas configurations
# - Shorter initial timeouts prevent blocking deployment health checks
client = AsyncIOMotorClient(
    mongo_url,
    serverSelectionTimeoutMS=30000,  # 30 seconds for server selection
    connectTimeoutMS=20000,          # 20 seconds for initial connection
    socketTimeoutMS=20000,           # 20 seconds for socket operations
    retryWrites=True,
    retryReads=True,
    w='majority',
    maxPoolSize=10,
    minPoolSize=0,                   # Allow pool to shrink to 0 when idle
    maxIdleTimeMS=45000,             # Close idle connections after 45s
    waitQueueTimeoutMS=10000,        # Max wait time for connection from pool
    appName='promanage-erp'          # Identify app in Atlas logs
)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-this')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 7

# ==================== AUTO-CREATE DEFAULT ADMIN ====================
async def create_default_admin():
    """
    Create default admin user if no users exist in database.
    This function is designed to be resilient and non-blocking for production deployments.
    """
    import asyncio
    max_retries = 3  # Reduced retries to prevent long blocking
    retry_delay = 10  # 10 seconds between retries
    
    # Wait for application to fully start and network to stabilize
    await asyncio.sleep(15)
    
    for attempt in range(max_retries):
        try:
            # Test connection first with a simple ping
            await client.admin.command('ping')
            print(f"✅ MongoDB connection successful on attempt {attempt + 1}")
            
            user_count = await db.users.count_documents({})
            if user_count == 0:
                import bcrypt
                default_password = "Admin2024!"
                hashed_password = bcrypt.hashpw(default_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                
                default_admin = {
                    "id": "user_admin_default",
                    "user_id": "user_admin_default",
                    "name": "Jorge Carrero Rodriguez",
                    "email": "j.carrero@ohsmspr.com",
                    "password": hashed_password,
                    "role": "super_admin",
                    "picture": None,
                    "is_temp_password": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.users.insert_one(default_admin)
                print("✅ Default admin created: j.carrero@ohsmspr.com / Admin2024!")
            else:
                print(f"ℹ️ Database has {user_count} users, skipping default admin creation")
            return  # Success, exit function
        except Exception as e:
            print(f"⚠️ Attempt {attempt + 1}/{max_retries} - Error in admin creation: {e}")
            if attempt < max_retries - 1:
                await asyncio.sleep(retry_delay)
    
    print("⚠️ Could not create default admin after retries - app will continue normally")
    print("ℹ️ Admin can be created manually via the registration endpoint if needed")

@app.on_event("startup")
async def startup_event():
    """
    FastAPI startup event - runs admin creation in background.
    IMPORTANT: This must not block application startup for health checks.
    """
    import asyncio
    # Fire and forget - don't await, let it run in background
    asyncio.create_task(create_default_admin_background())
    print("🚀 Application started - admin creation running in background")

async def create_default_admin_background():
    """Background task wrapper for admin creation with full error isolation"""
    try:
        await create_default_admin()
    except Exception as e:
        # Log but never crash - this is a non-critical background task
        print(f"⚠️ Background admin creation error (non-fatal): {e}")

# ==================== HEALTH CHECK ENDPOINTS ====================
# These are critical for Kubernetes liveness/readiness probes

@app.get("/health")
@app.get("/api/health")
async def health_check():
    """
    Basic health check - returns immediately without DB dependency.
    Used by Kubernetes liveness probe to verify app is running.
    """
    return {"status": "healthy", "service": "promanage-erp"}

@app.get("/api/health/ready")
async def readiness_check():
    """
    Readiness check - verifies database connectivity.
    Used by Kubernetes readiness probe to verify app can serve traffic.
    """
    try:
        # Quick ping to verify MongoDB connection
        await client.admin.command('ping')
        return {"status": "ready", "database": "connected"}
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=503, detail=f"Database not ready: {str(e)}")

# ================================================================

class ProjectStatus(str, Enum):
    PLANNING = "planning"
    IN_PROGRESS = "in_progress"
    ON_HOLD = "on_hold"
    COMPLETED = "completed"
    CANCELLED = "cancelled"

class TaskStatus(str, Enum):
    TODO = "todo"
    IN_PROGRESS = "in_progress"
    REVIEW = "review"
    DONE = "done"

class Priority(str, Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    URGENT = "urgent"

class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    PROJECT_MANAGER = "project_manager"
    RRHH = "rrhh"
    EMPLEADO = "empleado"
    CLIENT = "client"
    ACCOUNTANT = "accountant"

# Role permissions helper
ROLE_PERMISSIONS = {
    "super_admin": ["*"],  # All permissions
    "project_manager": [
        "view_assigned_projects", "view_project_budget", "view_project_delta",
        "approve_hours", "view_timesheets", "upload_project_docs", "mark_client_visible",
        "clock_in_out", "view_own_hours", "view_own_history", "upload_docs"
    ],
    "rrhh": [
        "create_employees", "edit_employees", "view_all_hours", "view_payroll",
        "view_employee_docs", "apply_labor_rules", "block_users",
        "clock_in_out", "view_own_hours", "view_own_history", "upload_docs"
    ],
    "empleado": [
        "clock_in_out", "view_own_hours", "view_own_history", 
        "view_assigned_projects", "upload_docs"
    ],
    "client": ["view_own_profile", "upload_client_docs"],
    "accountant": [
        "view_accounting", "manage_chart_of_accounts", "create_journal_entries",
        "view_general_ledger", "view_financial_statements", "manage_ar_ap",
        "bank_reconciliation", "view_tax_reports", "generate_reports",
        "clock_in_out", "view_own_hours", "view_own_history"
    ]
}

def has_permission(role: str, permission: str) -> bool:
    if role == "super_admin":
        return True
    return permission in ROLE_PERMISSIONS.get(role, [])

class PaymentStatus(str, Enum):
    PENDING = "pending"
    PARTIAL = "partial"
    PAID = "paid"

class UserRegister(BaseModel):
    name: str
    email: Optional[str] = None
    password: Optional[str] = None
    role: str = "empleado"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    name: Optional[str] = "Sin nombre"
    email: Optional[str] = ""
    role: str  # Changed from UserRole to str for flexibility with legacy data
    picture: Optional[str] = None
    created_at: Optional[str] = None

class EmployeeProfile(BaseModel):
    employee_id: Optional[str] = None
    user_id: Optional[str] = None
    # Personal Info
    phone: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    zipcode: Optional[str] = ""
    country: Optional[str] = ""
    date_of_birth: Optional[str] = ""
    gender: Optional[str] = ""
    marital_status: Optional[str] = ""
    nationality: Optional[str] = ""
    id_number: Optional[str] = ""  # Cédula/SSN
    # Employment Info
    department: Optional[str] = ""
    position: Optional[str] = ""
    hire_date: Optional[str] = ""
    employment_type: Optional[str] = ""  # full-time, part-time, contractor
    worker_classification: Optional[str] = ""  # employee or contractor (servicios profesionales)
    salary: Optional[float] = 0
    hourly_rate: Optional[float] = 0
    pay_frequency: Optional[str] = ""  # weekly, biweekly, semimonthly, monthly
    bank_name: Optional[str] = ""
    bank_account: Optional[str] = ""
    routing_number: Optional[str] = ""
    account_type: Optional[str] = ""  # checking, savings
    # Emergency Contact
    emergency_contact_name: Optional[str] = ""
    emergency_contact_phone: Optional[str] = ""
    emergency_contact_relationship: Optional[str] = ""
    # Notes
    notes: Optional[str] = ""
    
    @field_validator('salary', 'hourly_rate', mode='before')
    @classmethod
    def convert_empty_string_to_zero(cls, v):
        if v == "" or v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

class ProjectCreate(BaseModel):
    name: str
    description: str
    start_date: str
    end_date: str
    status: ProjectStatus = ProjectStatus.PLANNING
    priority: Priority = Priority.MEDIUM
    budget_total: float = 0
    project_value: float = 0
    payment_status: PaymentStatus = PaymentStatus.PENDING
    po_summary: Optional[str] = None
    resource: Optional[str] = None
    initials: Optional[str] = None
    project_number: Optional[str] = None
    client: Optional[str] = None
    sponsor: Optional[str] = None
    po_number: Optional[str] = None
    po_quantity: Optional[float] = None
    proposal_number: Optional[str] = None
    team_members: List[str] = []
    # Geofencing fields
    location_latitude: Optional[float] = None
    location_longitude: Optional[float] = None
    geofence_radius: float = 100  # metros
    geofence_enabled: bool = False

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[ProjectStatus] = None
    priority: Optional[Priority] = None
    budget_total: Optional[float] = None
    project_value: Optional[float] = None
    payment_status: Optional[PaymentStatus] = None
    po_summary: Optional[str] = None
    resource: Optional[str] = None
    initials: Optional[str] = None
    project_number: Optional[str] = None
    client: Optional[str] = None
    sponsor: Optional[str] = None
    po_number: Optional[str] = None
    po_quantity: Optional[float] = None
    proposal_number: Optional[str] = None
    team_members: Optional[List[str]] = None
    # Geofencing fields
    location_latitude: Optional[float] = None
    location_longitude: Optional[float] = None
    geofence_radius: Optional[float] = None
    geofence_enabled: Optional[bool] = None

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    project_id: str
    name: str
    description: str
    start_date: str
    end_date: str
    status: ProjectStatus
    priority: Priority
    budget_total: float
    budget_spent: float = 0
    project_value: float = 0
    profit: float = 0
    payment_status: PaymentStatus = PaymentStatus.PENDING
    po_summary: Optional[str] = None
    resource: Optional[str] = None
    initials: Optional[str] = None
    project_number: Optional[str] = None
    client: Optional[str] = None
    sponsor: Optional[str] = None
    po_number: Optional[str] = None
    po_quantity: Optional[float] = None
    proposal_number: Optional[str] = None
    cover_image: Optional[str] = None
    created_by: str
    team_members: List[str]
    created_at: str
    updated_at: str

class TaskCreate(BaseModel):
    project_id: str
    title: str
    description: str
    assigned_to: Optional[str] = None
    status: TaskStatus = TaskStatus.TODO
    priority: Priority = Priority.MEDIUM
    due_date: Optional[str] = None
    progress: int = 0

class Task(BaseModel):
    model_config = ConfigDict(extra="ignore")
    task_id: str
    project_id: str
    title: str
    description: str
    assigned_to: Optional[str] = None
    status: TaskStatus
    priority: Priority
    due_date: Optional[str] = None
    progress: int
    created_at: str
    updated_at: str

class BudgetCategoryCreate(BaseModel):
    project_id: str
    name: str
    allocated_amount: float

class BudgetCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    category_id: str
    project_id: str
    name: str
    allocated_amount: float
    spent_amount: float = 0
    created_at: str

class ExpenseCreate(BaseModel):
    project_id: str
    category_id: str
    description: str
    amount: float
    date: str

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    expense_id: str
    project_id: str
    category_id: str
    description: str
    amount: float
    date: str
    created_by: str
    created_at: str

class CommentCreate(BaseModel):
    project_id: str
    content: str

class Comment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    comment_id: str
    project_id: str
    user_id: str
    user_name: str
    content: str
    timestamp: str

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    notification_id: str
    user_id: str
    type: str
    message: str
    read: bool = False
    timestamp: str

# Project Logs Models
class ProjectLogCreate(BaseModel):
    project_id: str
    log_type: str  # work, update, problem, milestone, note
    title: str
    description: str
    hours_worked: Optional[float] = None
    attachments: Optional[List[str]] = []

class ProjectLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    project_id: str
    project_name: Optional[str] = None
    user_id: str
    user_name: str
    log_type: str
    title: str
    description: str
    hours_worked: Optional[float] = None
    attachments: List[str] = []
    created_at: str
    updated_at: Optional[str] = None

class DocumentFolder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    folder_id: str
    project_id: str
    name: str
    parent_folder_id: Optional[str] = None
    created_by: str
    created_by_name: str
    created_at: str

class DocumentFolderCreate(BaseModel):
    project_id: str
    name: str
    parent_folder_id: Optional[str] = None

class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    document_id: str
    project_id: str
    folder_id: Optional[str] = None
    filename: str
    original_filename: str
    file_size: int
    file_type: str
    uploaded_by: str
    uploaded_by_name: str
    uploaded_at: str

class LaborCreate(BaseModel):
    project_id: str
    labor_category: str
    hours_per_week: float
    hourly_rate: float
    estimated_total_hours: float
    consumed_hours: float = 0
    overtime_hours: float = 0
    overtime_rate: float = 0
    expenses: float = 0
    comments: Optional[str] = None

class Labor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    labor_id: str
    project_id: str
    labor_category: str
    hours_per_week: float
    hourly_rate: float
    estimated_total_hours: float
    consumed_hours: float
    consumed_cost: float
    overtime_hours: float
    overtime_rate: float
    expenses: float
    total_cost: float
    comments: Optional[str]
    created_at: str
    updated_at: str

class TimesheetCreate(BaseModel):
    project_id: str
    user_id: str
    user_name: str
    date: str
    hours_worked: float
    description: str
    task_id: Optional[str] = None

class Timesheet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    timesheet_id: str
    project_id: str
    user_id: str
    user_name: str
    date: str
    hours_worked: float
    description: str
    task_id: Optional[str]
    created_at: str
    updated_at: str

class ClockEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    clock_id: str
    user_id: str
    user_name: str
    project_id: str
    project_name: str
    clock_in: str
    clock_out: Optional[str] = None
    hours_worked: Optional[float] = None
    status: str  # active, completed
    date: str
    notes: Optional[str] = None
    clock_in_latitude: Optional[float] = None
    clock_in_longitude: Optional[float] = None
    clock_in_address: Optional[str] = None
    clock_out_latitude: Optional[float] = None
    clock_out_longitude: Optional[float] = None
    clock_out_address: Optional[str] = None

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    user_id: str
    user_name: Optional[str] = "N/A"
    action: str  # create, update, delete
    entity_type: str  # project, task, expense, user, etc.
    entity_id: str
    entity_name: Optional[str] = "N/A"
    details: Optional[dict] = None
    timestamp: str

class InvoiceItem(BaseModel):
    description: str
    hours: float
    rate: float
    amount: float

class InvoiceCreate(BaseModel):
    project_id: str
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    sponsor_name: Optional[str] = None
    tax_rate: float = 0.0
    tax_type_id: Optional[str] = None
    tax_type_name: Optional[str] = None
    notes: Optional[str] = None
    custom_number: Optional[str] = None

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    invoice_id: str
    invoice_number: str
    project_id: str
    project_name: str
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    sponsor_name: Optional[str] = None
    po_number: Optional[str] = None
    items: List[InvoiceItem]
    subtotal: float
    tax_rate: float
    tax_amount: float
    selected_taxes: Optional[List[dict]] = []
    total: float
    amount_paid: float = 0.0
    balance_due: float = 0.0
    status: str  # draft, sent, paid, overdue, partial
    notes: Optional[str] = None
    terms: Optional[str] = None
    price_breakdown: Optional[dict] = None  # {material_equipment: float, labor: float, total: float}
    created_by: str
    created_at: str
    due_date: Optional[str] = None
    sent_date: Optional[str] = None
    paid_date: Optional[str] = None

class PaymentCreate(BaseModel):
    amount: float
    payment_method: str  # cash, card, transfer, check
    reference: Optional[str] = None
    notes: Optional[str] = None

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    payment_id: str
    invoice_id: str
    amount: float
    payment_method: str
    reference: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    created_at: str

# Estimate Models
class EstimateItem(BaseModel):
    description: str
    quantity: float = 1.0
    unit_price: float
    amount: float

class EstimateCreate(BaseModel):
    project_id: Optional[str] = None
    client_profile_id: Optional[str] = None
    client_company: Optional[str] = None
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    title: str
    description: Optional[str] = None
    items: List[EstimateItem]
    tax_rate: float = 0.0
    selected_taxes: Optional[List[dict]] = []
    discount_percent: float = 0.0
    notes: Optional[str] = None
    terms: Optional[str] = None
    valid_until: Optional[str] = None
    custom_number: Optional[str] = None
    price_breakdown: Optional[dict] = None

class Estimate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    estimate_id: str
    estimate_number: str
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    client_profile_id: Optional[str] = None
    client_company: Optional[str] = None
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    title: str
    description: Optional[str] = None
    items: List[EstimateItem]
    subtotal: float
    discount_percent: float = 0.0
    discount_amount: float = 0.0
    tax_rate: float = 0.0
    tax_amount: float = 0.0
    selected_taxes: Optional[List[dict]] = []
    total: float
    status: str  # draft, sent, approved, rejected, converted
    notes: Optional[str] = None
    terms: Optional[str] = None
    valid_until: Optional[str] = None
    created_by: str
    created_by_name: str
    created_at: str
    sent_date: Optional[str] = None
    approved_date: Optional[str] = None
    converted_invoice_id: Optional[str] = None
    # Price Breakdown (separate from items)
    price_breakdown: Optional[dict] = None  # {material_equipment: float, labor: float, total: float}

# Purchase Order Models
class PurchaseOrderItem(BaseModel):
    description: str
    quantity: float = 1.0
    unit_price: float
    amount: float

class PurchaseOrderCreate(BaseModel):
    project_id: Optional[str] = None
    supplier_name: str
    supplier_email: Optional[str] = None
    supplier_phone: Optional[str] = None
    supplier_address: Optional[str] = None
    title: str
    description: Optional[str] = None
    items: List[PurchaseOrderItem]
    tax_rate: float = 0.0
    discount_percent: float = 0.0
    notes: Optional[str] = None
    terms: Optional[str] = None
    expected_delivery_date: Optional[str] = None
    custom_number: Optional[str] = None

class PurchaseOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    po_id: str
    po_number: str
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    supplier_name: str
    supplier_email: Optional[str] = None
    supplier_phone: Optional[str] = None
    supplier_address: Optional[str] = None
    title: str
    description: Optional[str] = None
    items: List[PurchaseOrderItem]
    subtotal: float
    discount_percent: float = 0.0
    discount_amount: float = 0.0
    tax_rate: float = 0.0
    tax_amount: float = 0.0
    total: float
    status: str  # draft, approved, sent, partially_received, completed, cancelled
    notes: Optional[str] = None
    terms: Optional[str] = None
    expected_delivery_date: Optional[str] = None
    created_by: str
    created_by_name: str
    created_at: str
    approved_date: Optional[str] = None
    sent_date: Optional[str] = None
    received_date: Optional[str] = None
    linked_expense_id: Optional[str] = None

async def log_audit(user_id: str, user_name: str, action: str, entity_type: str, entity_id: str, entity_name: str, details: dict = None):
    """Helper function to create audit log entries"""
    log_entry = {
        "log_id": f"log_{uuid4().hex[:16]}",
        "user_id": user_id,
        "user_name": user_name,
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity_name": entity_name,
        "details": details or {},
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log_entry)

class IntegrationConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    integration_id: str
    integration_type: str  # slack, google_calendar, github, gitlab
    enabled: bool
    config: dict  # Contains webhook_url, api_key, etc.
    created_at: str
    updated_at: str

async def send_slack_notification(webhook_url: str, message: str, title: str = None, color: str = "good"):
    """Send notification to Slack webhook"""
    if not webhook_url:
        return
    
    import requests
    
    payload = {
        "attachments": [{
            "color": color,
            "title": title or "Notificación de Proyectos",
            "text": message,
            "footer": "Sistema de Gestión de Proyectos",
            "ts": int(datetime.now(timezone.utc).timestamp())
        }]
    }
    
    try:
        requests.post(webhook_url, json=payload, timeout=5)
    except Exception as e:
        print(f"Error sending Slack notification: {e}")

async def notify_slack_event(event_type: str, details: dict):
    """Send notification to Slack if integration is enabled"""
    try:
        integration = await db.integrations.find_one(
            {"integration_type": "slack", "enabled": True},
            {"_id": 0}
        )
        
        if not integration:
            return
        
        webhook_url = integration.get('config', {}).get('webhook_url')
        if not webhook_url:
            return
        
        # Format message based on event type
        messages = {
            "project_created": f"🎉 Nuevo proyecto creado: **{details.get('name')}**",
            "task_completed": f"✅ Tarea completada: **{details.get('title')}** en proyecto **{details.get('project')}**",
            "invoice_paid": f"💰 Factura pagada: **{details.get('invoice_number')}** - ${details.get('amount')}",
            "project_overdue": f"⚠️ Proyecto retrasado: **{details.get('name')}** - Vencía: {details.get('due_date')}",
            "payment_received": f"💵 Pago recibido: **${details.get('amount')}** para factura **{details.get('invoice')}**"
        }
        
        message = messages.get(event_type, f"Evento: {event_type}")
        color = "good" if event_type in ["task_completed", "invoice_paid", "payment_received"] else "#764FA5"
        
        await send_slack_notification(webhook_url, message, "Notificación del Sistema", color)
    except Exception as e:
        print(f"Error in notify_slack_event: {e}")

async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)):
    token = session_token
    if not token:
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    
    session_doc = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Sesión expirada")
    
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Return a simple object with required attributes instead of Pydantic model
    class SimpleUser:
        def __init__(self, doc):
            self.user_id = doc.get('user_id', doc.get('id', ''))
            self.name = doc.get('name', 'Sin nombre')
            self.email = doc.get('email', '')
            self.role = doc.get('role', 'empleado')
            self.picture = doc.get('picture')
            self.created_at = doc.get('created_at', '')
            self.id = self.user_id  # Alias
    
    return SimpleUser(user_doc)

@api_router.post("/auth/setup")
async def initial_setup(user_data: UserRegister):
    """
    Setup inicial - Solo funciona si NO hay super_admins en la base de datos.
    Esto permite crear el primer admin en una instalación nueva.
    """
    # Check if any super_admin exists
    existing_admin = await db.users.find_one({"role": "super_admin"}, {"_id": 0})
    if existing_admin:
        raise HTTPException(status_code=403, detail="El sistema ya tiene un administrador configurado. Use login normal.")
    
    if not user_data.email or not user_data.password:
        raise HTTPException(status_code=400, detail="Email y contraseña son requeridos")
    
    # Check if email exists
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    user_doc = {
        "id": user_id,
        "user_id": user_id,
        "name": user_data.name,
        "email": user_data.email,
        "password": hashed_password,
        "role": "super_admin",  # Force super_admin for setup
        "picture": None,
        "is_temp_password": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    return {"message": "Super Admin creado exitosamente. Ya puedes iniciar sesión."}

class EmergencyResetRequest(BaseModel):
    email: str
    new_password: str
    secret_key: str

@api_router.post("/auth/emergency-reset")
async def emergency_password_reset(data: EmergencyResetRequest):
    """
    Reset de contraseña de emergencia para cuando no puedes acceder.
    Requiere una clave secreta temporal.
    """
    # Clave secreta temporal - CAMBIAR DESPUÉS DE USAR
    EMERGENCY_KEY = "OHSMS2024RESET"
    
    if data.secret_key != EMERGENCY_KEY:
        raise HTTPException(status_code=403, detail="Clave de emergencia inválida")
    
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    
    # Find user by email
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Hash new password
    hashed_password = bcrypt.hashpw(data.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Update password
    await db.users.update_one(
        {"email": data.email},
        {"$set": {"password": hashed_password, "is_temp_password": False}}
    )
    
    return {"message": f"Contraseña actualizada para {data.email}. Ya puedes iniciar sesión."}

@api_router.post("/auth/register")
async def register(user_data: UserRegister, request: Request, session_token: Optional[str] = Cookie(None)):
    """Registro de usuarios - Solo Super Admin y RRHH pueden crear cuentas"""
    # Verificar autenticación - solo admins y RRHH pueden crear usuarios
    try:
        current_user = await get_current_user(request, session_token)
        if current_user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
            raise HTTPException(status_code=403, detail="Solo Super Admin y RRHH pueden crear usuarios")
        
        # RRHH no puede crear super_admin
        if current_user.role == UserRole.RRHH.value and user_data.role == UserRole.SUPER_ADMIN.value:
            raise HTTPException(status_code=403, detail="No puedes crear usuarios Super Admin")
    except HTTPException as e:
        if e.status_code == 401:
            raise HTTPException(status_code=403, detail="El registro público está deshabilitado. Contacta al administrador.")
        raise e
    
    # Si se proporciona email, verificar que no exista
    if user_data.email:
        existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Email ya registrado")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    # Determine if this is a temporary password (created by admin)
    is_temp_password = False
    original_password = user_data.password  # Store for email
    
    # Crear documento de usuario
    user_doc = {
        "user_id": user_id,
        "name": user_data.name,
        "email": user_data.email or "",
        "role": user_data.role,
        "picture": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Solo hashear password si se proporciona
    if user_data.password:
        hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        user_doc["password"] = hashed_password
        user_doc["is_temp_password"] = True  # Mark as temporary
        is_temp_password = True
    else:
        user_doc["password"] = ""
        user_doc["is_temp_password"] = False
    
    await db.users.insert_one(user_doc)
    
    # Send welcome email with credentials if email is provided
    if user_data.email and original_password and is_temp_password:
        try:
            html, text = get_welcome_email(user_data.name, user_data.email, original_password)
            await send_email(user_data.email, "🎉 Bienvenido a ProManage - Tus credenciales de acceso", html, text)
        except Exception as e:
            print(f"Error sending welcome email: {e}")
    
    user_doc_without_password = {k: v for k, v in user_doc.items() if k not in ['password', 'is_temp_password']}
    return user_doc_without_password

@api_router.post("/auth/login")
async def login(credentials: UserLogin, response: Response):
    print(f"Login attempt for: {credentials.email}")
    print(f"Password received: {credentials.password}")
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    print(f"User found: {user_doc is not None}")
    if not user_doc:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    # Check if user is blocked
    if user_doc.get("is_blocked", False):
        raise HTTPException(status_code=403, detail="Tu cuenta ha sido bloqueada. Contacta al administrador.")
    
    print(f"Stored hash: {user_doc['password'][:60]}...")
    password_match = bcrypt.checkpw(credentials.password.encode('utf-8'), user_doc['password'].encode('utf-8'))
    print(f"Password match: {password_match}")
    if not password_match:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    session_token = f"session_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    
    session_doc = {
        "user_id": user_doc["user_id"],
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=ACCESS_TOKEN_EXPIRE_DAYS * 24 * 60 * 60
    )
    
    user_without_password = {k: v for k, v in user_doc.items() if k != 'password'}
    
    # Check if password is temporary
    is_temp_password = user_doc.get('is_temp_password', False)
    
    return {
        "user": user_without_password, 
        "token": session_token,
        "requires_password_change": is_temp_password
    }

@api_router.post("/auth/session")
async def create_session_from_emergent(request: Request, response: Response):
    session_id = request.headers.get('X-Session-ID')
    if not session_id:
        raise HTTPException(status_code=400, detail="No session ID provided")
    
    try:
        emergent_response = requests.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id},
            timeout=10
        )
        
        if emergent_response.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        data = emergent_response.json()
        email = data.get('email')
        name = data.get('name')
        picture = data.get('picture')
        emergent_session_token = data.get('session_token')
        
        user_doc = await db.users.find_one({"email": email}, {"_id": 0})
        
        if not user_doc:
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            user_doc = {
                "user_id": user_id,
                "name": name,
                "email": email,
                "role": UserRole.EMPLEADO.value,
                "picture": picture,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(user_doc)
        else:
            if user_doc.get('name') != name or user_doc.get('picture') != picture:
                await db.users.update_one(
                    {"user_id": user_doc['user_id']},
                    {"$set": {"name": name, "picture": picture}}
                )
                user_doc['name'] = name
                user_doc['picture'] = picture
        
        session_token = emergent_session_token
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        session_doc = {
            "user_id": user_doc["user_id"],
            "session_token": session_token,
            "expires_at": expires_at,
            "created_at": datetime.now(timezone.utc)
        }
        await db.user_sessions.insert_one(session_doc)
        
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            max_age=7 * 24 * 60 * 60
        )
        
        user_without_password = {k: v for k, v in user_doc.items() if k != 'password'}
        return {"user": user_without_password, "token": session_token}
    
    except Exception as e:
        logging.error(f"Error creating session: {str(e)}")
        raise HTTPException(status_code=500, detail="Error al crear sesión")

@api_router.get("/auth/me")
async def get_me(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    # Return as dict to avoid Pydantic validation issues
    return {
        "user_id": user.user_id,
        "name": user.name,
        "email": user.email,
        "role": user.role,
        "picture": user.picture,
        "created_at": user.created_at
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response, session_token: Optional[str] = Cookie(None)):
    token = session_token
    if not token:
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if token:
        # Delete session from database
        await db.user_sessions.delete_one({"session_token": token})
    
    # Delete cookie - must set max_age=0 to force deletion
    response.set_cookie(
        key="session_token",
        value="",
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=0  # This forces cookie deletion
    )
    
    return {"message": "Sesión cerrada exitosamente"}

@api_router.post("/auth/change-password")
async def change_password(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Change user password and clear temporary password flag"""
    user = await get_current_user(request, session_token)
    
    current_password = data.get("current_password")
    new_password = data.get("new_password")
    
    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Se requieren contraseña actual y nueva")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")
    
    # Get user with password
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Verify current password
    if not bcrypt.checkpw(current_password.encode('utf-8'), user_doc['password'].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Contraseña actual incorrecta")
    
    # Hash new password
    hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Update password and clear temp flag
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"password": hashed_password, "is_temp_password": False}}
    )
    
    return {"message": "Contraseña actualizada exitosamente"}

@api_router.post("/projects", response_model=Project)
async def create_project(project_data: ProjectCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    print("=== CREATE PROJECT REQUEST ===")
    print(f"Project data received: {project_data}")
    
    try:
        user = await get_current_user(request, session_token)
        print(f"User authenticated: {user.email} (role: {user.role})")
    except Exception as e:
        print(f"Authentication error: {e}")
        raise
    
    # Solo admins y project managers pueden crear proyectos
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        print(f"Permission denied for role: {user.role}")
        raise HTTPException(status_code=403, detail="No tienes permiso para crear proyectos")
    
    project_id = f"proj_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    
    project_doc = {
        "project_id": project_id,
        "name": project_data.name,
        "description": project_data.description,
        "start_date": project_data.start_date,
        "end_date": project_data.end_date,
        "status": project_data.status,
        "priority": project_data.priority,
        "budget_total": project_data.budget_total,
        "budget_spent": 0,
        "project_value": project_data.project_value,
        "profit": project_data.project_value - 0,
        "payment_status": project_data.payment_status,
        "po_summary": project_data.po_summary,
        "resource": project_data.resource,
        "initials": project_data.initials,
        "project_number": project_data.project_number,
        "client": project_data.client,
        "sponsor": project_data.sponsor,
        "po_number": project_data.po_number,
        "po_quantity": project_data.po_quantity,
        "proposal_number": project_data.proposal_number,
        "cover_image": None,
        "created_by": user.user_id,
        "team_members": project_data.team_members,
        "created_at": now,
        "updated_at": now
    }
    
    print(f"Inserting project: {project_id}")
    try:
        await db.projects.insert_one(project_doc)
        print(f"Project inserted successfully: {project_id}")
    except Exception as e:
        print(f"Database insert error: {e}")
        raise
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "create",
        "project",
        project_id,
        project_data.name,
        {"budget": project_data.budget_total, "status": project_data.status}
    )
    
    # Send Slack notification
    await notify_slack_event("project_created", {
        "name": project_data.name,
        "created_by": user.name
    })
    
    return Project(**project_doc)

def extract_project_number(project_number: str) -> tuple:
    """Extrae año y número para ordenamiento. P-2025-15 -> (2025, 15)"""
    import re
    if not project_number:
        return (0, 0)
    match = re.search(r'(\d{4})[^\d]*(\d+)', project_number)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return (0, 0)

@api_router.get("/projects", response_model=List[Project])
async def get_projects(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Super Admin y Project Manager pueden ver todos los proyectos
    if user.role in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    else:
        projects = await db.projects.find(
            {"$or": [{"created_by": user.user_id}, {"team_members": user.user_id}]},
            {"_id": 0}
        ).to_list(1000)
    
    # Calcular profit para cada proyecto y asegurar payment_status
    for p in projects:
        p['project_value'] = p.get('project_value', 0)
        p['budget_spent'] = p.get('budget_spent', 0)
        p['profit'] = p['project_value'] - p['budget_spent']
        p['payment_status'] = p.get('payment_status', 'pending')
    
    # Ordenar por número de proyecto (año, número)
    projects.sort(key=lambda p: extract_project_number(p.get('project_number', '')))
    
    return [Project(**p) for p in projects]

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # PM y Admin tienen acceso completo, otros solo si son creadores o del equipo
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    
    # Asegurar que existan los campos y calcular profit
    project_doc['project_value'] = project_doc.get('project_value', 0)
    project_doc['budget_spent'] = project_doc.get('budget_spent', 0)
    project_doc['profit'] = project_doc['project_value'] - project_doc['budget_spent']
    project_doc['payment_status'] = project_doc.get('payment_status', 'pending')
    
    return Project(**project_doc)

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_data: ProjectUpdate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # PM puede editar proyectos, Admin puede editar todos
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para editar este proyecto")
    
    update_data = project_data.model_dump(exclude_unset=True)
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # Recalcular ganancia si se actualiza project_value
    if 'project_value' in update_data or 'budget_spent' in update_data:
        new_value = update_data.get('project_value', project_doc.get('project_value', 0))
        current_spent = project_doc.get('budget_spent', 0)
        update_data['profit'] = new_value - current_spent
    
    await db.projects.update_one({"project_id": project_id}, {"$set": update_data})
    
    updated_project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    # Recalcular profit con datos actuales
    updated_project['profit'] = updated_project.get('project_value', 0) - updated_project.get('budget_spent', 0)
    return Project(**updated_project)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if user.role != UserRole.SUPER_ADMIN.value and project_doc['created_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar este proyecto")
    
    documents = await db.documents.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    for doc in documents:
        file_path = UPLOAD_DIR / doc['filename']
        if file_path.exists():
            file_path.unlink()
    
    await db.projects.delete_one({"project_id": project_id})
    await db.tasks.delete_many({"project_id": project_id})
    await db.budget_categories.delete_many({"project_id": project_id})
    await db.expenses.delete_many({"project_id": project_id})
    await db.comments.delete_many({"project_id": project_id})
    await db.documents.delete_many({"project_id": project_id})
    
    return {"message": "Proyecto eliminado exitosamente"}

# ==================== PROJECT TEAM MEMBERS ====================

@api_router.get("/projects/{project_id}/team")
async def get_project_team(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Get team members for a project with their user details"""
    user = await get_current_user(request, session_token)
    
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    team_member_ids = project.get("team_members", [])
    
    # Get user details for each team member
    team_members = []
    for member_id in team_member_ids:
        member = await db.users.find_one({"user_id": member_id}, {"_id": 0, "password_hash": 0})
        if member:
            team_members.append({
                "user_id": member.get("user_id"),
                "name": member.get("name"),
                "email": member.get("email"),
                "role": member.get("role"),
                "position": member.get("position", "")
            })
    
    return team_members

@api_router.post("/projects/{project_id}/team")
async def add_team_member(project_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Add a team member to a project"""
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo PM o administradores pueden gestionar el equipo")
    
    project = await db.projects.find_one({"project_id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    member_id = data.get("user_id")
    if not member_id:
        raise HTTPException(status_code=400, detail="user_id es requerido")
    
    # Verify user exists
    member = await db.users.find_one({"user_id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Add to team_members if not already there
    team_members = project.get("team_members", [])
    if member_id not in team_members:
        team_members.append(member_id)
        await db.projects.update_one(
            {"project_id": project_id},
            {"$set": {"team_members": team_members, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {"message": f"{member.get('name')} añadido al equipo"}

@api_router.delete("/projects/{project_id}/team/{member_id}")
async def remove_team_member(project_id: str, member_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Remove a team member from a project"""
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo PM o administradores pueden gestionar el equipo")
    
    project = await db.projects.find_one({"project_id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    team_members = project.get("team_members", [])
    if member_id in team_members:
        team_members.remove(member_id)
        await db.projects.update_one(
            {"project_id": project_id},
            {"$set": {"team_members": team_members, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {"message": "Miembro eliminado del equipo"}

# ==================== CHANGE ORDERS ====================
@api_router.post("/change-orders")
async def create_change_order(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project = await db.projects.find_one({"project_id": data.get("project_id")})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    change_order = {
        "id": str(uuid4()),
        "project_id": data.get("project_id"),
        "description": data.get("description", ""),
        "budget_change": data.get("budget_change", 0),
        "value_change": data.get("value_change", 0),
        "reason": data.get("reason", ""),
        "status": "pending",
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.change_orders.insert_one(change_order)
    await log_audit(user.user_id, user.name, "create", "change_order", change_order["id"], data.get("description", "Change Order"), {"project": project.get("name"), "budget_change": data.get("budget_change"), "value_change": data.get("value_change")})
    return {"message": "Change Order creada", "id": change_order["id"]}

@api_router.get("/change-orders")
async def get_change_orders(project_id: str = None, request: Request = None, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    query = {"project_id": project_id} if project_id else {}
    orders = await db.change_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return orders

@api_router.put("/change-orders/{order_id}")
async def update_change_order(order_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo PM o administradores pueden aprobar")
    
    order = await db.change_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Change Order no encontrada")
    
    new_status = data.get("status")
    
    if new_status == "approved":
        # Actualizar presupuesto y valor del proyecto
        await db.projects.update_one(
            {"project_id": order["project_id"]},
            {"$inc": {
                "budget_total": order.get("budget_change", 0),
                "project_value": order.get("value_change", 0)
            }}
        )
    
    await db.change_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": new_status,
            "reviewed_by": user.user_id,
            "reviewed_by_name": user.name,
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "review_notes": data.get("notes", "")
        }}
    )
    
    await log_audit(user.user_id, user.name, "approve" if new_status == "approved" else "update", "change_order", order_id, order.get("description", "Change Order"), {"status": new_status})
    return {"message": f"Change Order {new_status}"}

@api_router.delete("/change-orders/{order_id}")
async def delete_change_order(order_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo PM o administradores pueden eliminar")
    
    order = await db.change_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Change Order no encontrada")
    
    # Si fue aprobada, revertir los cambios en el proyecto
    if order.get("status") == "approved":
        await db.projects.update_one(
            {"project_id": order["project_id"]},
            {"$inc": {
                "budget_total": -order.get("budget_change", 0),
                "project_value": -order.get("value_change", 0)
            }}
        )
    
    await db.change_orders.delete_one({"id": order_id})
    await log_audit(user.user_id, user.name, "delete", "change_order", order_id, order.get("description", "Change Order"), {"project_id": order["project_id"]})
    return {"message": "Change Order eliminada"}

@api_router.post("/tasks", response_model=Task)
async def create_task(task_data: TaskCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": task_data.project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    
    task_doc = {
        "task_id": task_id,
        "project_id": task_data.project_id,
        "title": task_data.title,
        "description": task_data.description,
        "assigned_to": task_data.assigned_to,
        "status": task_data.status,
        "priority": task_data.priority,
        "due_date": task_data.due_date,
        "progress": task_data.progress,
        "created_at": now,
        "updated_at": now
    }
    
    await db.tasks.insert_one(task_doc)
    
    # Create notification if task is assigned
    if task_data.assigned_to:
        notification_doc = {
            "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
            "user_id": task_data.assigned_to,
            "type": "task_assigned",
            "message": f"{user.name} te asignó la tarea: {task_data.title}",
            "read": False,
            "timestamp": now,
            "related_id": task_id
        }
        await db.notifications.insert_one(notification_doc)
        
        # Send email notification
        assigned_user = await db.users.find_one({"user_id": task_data.assigned_to}, {"_id": 0})
        if assigned_user:
            html, text = get_task_assigned_email(
                assigned_user['name'],
                task_data.title,
                project_doc['name'],
                user.name
            )
            await send_email(
                assigned_user['email'],
                f"Nueva tarea asignada: {task_data.title}",
                html,
                text
            )
    
    return Task(**task_doc)

@api_router.get("/tasks", response_model=List[Task])
async def get_tasks(project_id: Optional[str] = None, request: Request = None, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(1000)
    return [Task(**t) for t in tasks]

@api_router.put("/tasks/{task_id}", response_model=Task)
async def update_task(task_id: str, task_data: TaskCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    task_doc = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task_doc:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    update_data = task_data.model_dump()
    now = datetime.now(timezone.utc).isoformat()
    update_data['updated_at'] = now
    
    await db.tasks.update_one({"task_id": task_id}, {"$set": update_data})
    
    # Create notification if status changed to done
    if task_data.status == TaskStatus.DONE and task_doc.get('status') != TaskStatus.DONE:
        project_doc = await db.projects.find_one({"project_id": task_doc['project_id']}, {"_id": 0})
        if project_doc and project_doc.get('created_by'):
            notification_doc = {
                "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
                "user_id": project_doc['created_by'],
                "type": "task_completed",
                "message": f"{user.name} completó la tarea: {task_data.title}",
                "read": False,
                "timestamp": now,
                "related_id": task_id
            }
            await db.notifications.insert_one(notification_doc)
            
            # Send email notification
            owner_user = await db.users.find_one({"user_id": project_doc['created_by']}, {"_id": 0})
            if owner_user:
                html, text = get_task_completed_email(
                    owner_user['name'],
                    task_data.title,
                    project_doc['name'],
                    user.name
                )
                await send_email(
                    owner_user['email'],
                    f"Tarea completada: {task_data.title}",
                    html,
                    text
                )
    
    # Create notification if assigned user changed
    if task_data.assigned_to and task_data.assigned_to != task_doc.get('assigned_to'):
        notification_doc = {
            "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
            "user_id": task_data.assigned_to,
            "type": "task_assigned",
            "message": f"{user.name} te asignó la tarea: {task_data.title}",
            "read": False,
            "timestamp": now,
            "related_id": task_id
        }
        await db.notifications.insert_one(notification_doc)
    
    updated_task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    return Task(**updated_task)

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    await db.tasks.delete_one({"task_id": task_id})
    return {"message": "Tarea eliminada exitosamente"}

@api_router.post("/budget/categories", response_model=BudgetCategory)
async def create_budget_category(category_data: BudgetCategoryCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    
    category_doc = {
        "category_id": category_id,
        "project_id": category_data.project_id,
        "name": category_data.name,
        "allocated_amount": category_data.allocated_amount,
        "spent_amount": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.budget_categories.insert_one(category_doc)
    return BudgetCategory(**category_doc)

@api_router.get("/budget/categories", response_model=List[BudgetCategory])
async def get_budget_categories(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    categories = await db.budget_categories.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    return [BudgetCategory(**c) for c in categories]

@api_router.put("/budget/categories/{category_id}", response_model=BudgetCategory)
async def update_budget_category(category_id: str, category_data: BudgetCategoryCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    category = await db.budget_categories.find_one({"category_id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    
    update_data = {
        "name": category_data.name,
        "allocated_amount": category_data.allocated_amount
    }
    
    await db.budget_categories.update_one({"category_id": category_id}, {"$set": update_data})
    
    updated_category = await db.budget_categories.find_one({"category_id": category_id}, {"_id": 0})
    return BudgetCategory(**updated_category)

@api_router.delete("/budget/categories/{category_id}")
async def delete_budget_category(category_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    category = await db.budget_categories.find_one({"category_id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    
    # Check if category has expenses
    expenses = await db.expenses.find({"category_id": category_id}, {"_id": 0}).to_list(1)
    if expenses:
        raise HTTPException(status_code=400, detail="No se puede eliminar una categoría con gastos registrados")
    
    await db.budget_categories.delete_one({"category_id": category_id})
    return {"message": "Categoría eliminada exitosamente"}

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense_data: ExpenseCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    expense_id = f"exp_{uuid.uuid4().hex[:12]}"
    
    expense_doc = {
        "expense_id": expense_id,
        "project_id": expense_data.project_id,
        "category_id": expense_data.category_id,
        "description": expense_data.description,
        "amount": expense_data.amount,
        "date": expense_data.date,
        "created_by": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.expenses.insert_one(expense_doc)
    
    await db.budget_categories.update_one(
        {"category_id": expense_data.category_id},
        {"$inc": {"spent_amount": expense_data.amount}}
    )
    
    # Actualizar budget_spent y recalcular profit
    project = await db.projects.find_one({"project_id": expense_data.project_id}, {"_id": 0})
    new_spent = project.get('budget_spent', 0) + expense_data.amount
    project_value = project.get('project_value', 0)
    new_profit = project_value - new_spent
    
    await db.projects.update_one(
        {"project_id": expense_data.project_id},
        {"$set": {"budget_spent": new_spent, "profit": new_profit}}
    )
    
    return Expense(**expense_doc)

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    expenses = await db.expenses.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    return [Expense(**e) for e in expenses]

@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(expense_id: str, expense_data: ExpenseCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    # Revertir el gasto anterior
    old_amount = expense['amount']
    old_category_id = expense['category_id']
    
    await db.budget_categories.update_one(
        {"category_id": old_category_id},
        {"$inc": {"spent_amount": -old_amount}}
    )
    
    project = await db.projects.find_one({"project_id": expense['project_id']}, {"_id": 0})
    old_spent = project.get('budget_spent', 0) - old_amount
    old_profit = project.get('project_value', 0) - old_spent
    
    await db.projects.update_one(
        {"project_id": expense['project_id']},
        {"$set": {"budget_spent": old_spent, "profit": old_profit}}
    )
    
    # Aplicar el nuevo gasto
    update_data = {
        "category_id": expense_data.category_id,
        "description": expense_data.description,
        "amount": expense_data.amount,
        "date": expense_data.date
    }
    
    await db.expenses.update_one({"expense_id": expense_id}, {"$set": update_data})
    
    await db.budget_categories.update_one(
        {"category_id": expense_data.category_id},
        {"$inc": {"spent_amount": expense_data.amount}}
    )
    
    new_spent = old_spent + expense_data.amount
    new_profit = project.get('project_value', 0) - new_spent
    
    await db.projects.update_one(
        {"project_id": expense['project_id']},
        {"$set": {"budget_spent": new_spent, "profit": new_profit}}
    )
    
    updated_expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    return Expense(**updated_expense)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    # Revertir el gasto de la categoría
    await db.budget_categories.update_one(
        {"category_id": expense['category_id']},
        {"$inc": {"spent_amount": -expense['amount']}}
    )
    
    # Revertir el gasto del proyecto
    project = await db.projects.find_one({"project_id": expense['project_id']}, {"_id": 0})
    new_spent = project.get('budget_spent', 0) - expense['amount']
    new_profit = project.get('project_value', 0) - new_spent
    
    await db.projects.update_one(
        {"project_id": expense['project_id']},
        {"$set": {"budget_spent": new_spent, "profit": new_profit}}
    )
    
    await db.expenses.delete_one({"expense_id": expense_id})
    return {"message": "Gasto eliminado exitosamente"}

@api_router.post("/labor", response_model=Labor)
async def create_labor(labor_data: LaborCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    labor_id = f"labor_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    
    # Calcular total_cost (costo estimado)
    regular_cost = labor_data.estimated_total_hours * labor_data.hourly_rate
    overtime_cost = labor_data.overtime_hours * labor_data.overtime_rate
    total_cost = regular_cost + overtime_cost + labor_data.expenses
    
    # Calcular consumed_cost (costo real basado en horas consumidas)
    consumed_cost = labor_data.consumed_hours * labor_data.hourly_rate
    
    labor_doc = {
        "labor_id": labor_id,
        "project_id": labor_data.project_id,
        "labor_category": labor_data.labor_category,
        "hours_per_week": labor_data.hours_per_week,
        "hourly_rate": labor_data.hourly_rate,
        "estimated_total_hours": labor_data.estimated_total_hours,
        "consumed_hours": labor_data.consumed_hours,
        "consumed_cost": consumed_cost,
        "overtime_hours": labor_data.overtime_hours,
        "overtime_rate": labor_data.overtime_rate,
        "expenses": labor_data.expenses,
        "total_cost": total_cost,
        "comments": labor_data.comments,
        "created_at": now,
        "updated_at": now
    }
    
    await db.labor.insert_one(labor_doc)
    return Labor(**labor_doc)

@api_router.get("/labor", response_model=List[Labor])
async def get_labor(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    labor_records = await db.labor.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Asegurar que los registros antiguos tengan los nuevos campos
    for record in labor_records:
        if 'consumed_hours' not in record:
            record['consumed_hours'] = 0
        if 'consumed_cost' not in record:
            record['consumed_cost'] = 0
    
    return [Labor(**l) for l in labor_records]

@api_router.put("/labor/{labor_id}", response_model=Labor)
async def update_labor(labor_id: str, labor_data: LaborCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    labor = await db.labor.find_one({"labor_id": labor_id}, {"_id": 0})
    if not labor:
        raise HTTPException(status_code=404, detail="Registro de labor no encontrado")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Calcular total_cost (costo estimado)
    regular_cost = labor_data.estimated_total_hours * labor_data.hourly_rate
    overtime_cost = labor_data.overtime_hours * labor_data.overtime_rate
    total_cost = regular_cost + overtime_cost + labor_data.expenses
    
    # Calcular consumed_cost (costo real basado en horas consumidas)
    consumed_cost = labor_data.consumed_hours * labor_data.hourly_rate
    
    update_data = {
        "labor_category": labor_data.labor_category,
        "hours_per_week": labor_data.hours_per_week,
        "hourly_rate": labor_data.hourly_rate,
        "estimated_total_hours": labor_data.estimated_total_hours,
        "consumed_hours": labor_data.consumed_hours,
        "consumed_cost": consumed_cost,
        "overtime_hours": labor_data.overtime_hours,
        "overtime_rate": labor_data.overtime_rate,
        "expenses": labor_data.expenses,
        "total_cost": total_cost,
        "comments": labor_data.comments,
        "updated_at": now
    }
    
    await db.labor.update_one({"labor_id": labor_id}, {"$set": update_data})
    
    updated_labor = await db.labor.find_one({"labor_id": labor_id}, {"_id": 0})
    return Labor(**updated_labor)

@api_router.delete("/labor/{labor_id}")
async def delete_labor(labor_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    labor = await db.labor.find_one({"labor_id": labor_id}, {"_id": 0})
    if not labor:
        raise HTTPException(status_code=404, detail="Registro de labor no encontrado")
    
    await db.labor.delete_one({"labor_id": labor_id})
    return {"message": "Registro de labor eliminado exitosamente"}

@api_router.post("/timesheet", response_model=Timesheet)
async def create_timesheet(timesheet_data: TimesheetCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    timesheet_id = f"timesheet_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    
    timesheet_doc = {
        "timesheet_id": timesheet_id,
        "project_id": timesheet_data.project_id,
        "user_id": timesheet_data.user_id,
        "user_name": timesheet_data.user_name,
        "date": timesheet_data.date,
        "hours_worked": timesheet_data.hours_worked,
        "description": timesheet_data.description,
        "task_id": timesheet_data.task_id,
        "created_at": now,
        "updated_at": now
    }
    
    await db.timesheet.insert_one(timesheet_doc)
    return Timesheet(**timesheet_doc)

@api_router.get("/timesheet", response_model=List[Timesheet])
async def get_timesheets(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    timesheets = await db.timesheet.find({"project_id": project_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    return [Timesheet(**t) for t in timesheets]

@api_router.put("/timesheet/{timesheet_id}", response_model=Timesheet)
async def update_timesheet(timesheet_id: str, timesheet_data: TimesheetCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    timesheet = await db.timesheet.find_one({"timesheet_id": timesheet_id}, {"_id": 0})
    if not timesheet:
        raise HTTPException(status_code=404, detail="Registro de timesheet no encontrado")
    
    now = datetime.now(timezone.utc).isoformat()
    
    update_data = {
        "user_id": timesheet_data.user_id,
        "user_name": timesheet_data.user_name,
        "date": timesheet_data.date,
        "hours_worked": timesheet_data.hours_worked,
        "description": timesheet_data.description,
        "task_id": timesheet_data.task_id,
        "updated_at": now
    }
    
    await db.timesheet.update_one({"timesheet_id": timesheet_id}, {"$set": update_data})
    
    updated_timesheet = await db.timesheet.find_one({"timesheet_id": timesheet_id}, {"_id": 0})
    return Timesheet(**updated_timesheet)

@api_router.delete("/timesheet/{timesheet_id}")
async def delete_timesheet(timesheet_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    timesheet = await db.timesheet.find_one({"timesheet_id": timesheet_id}, {"_id": 0})
    if not timesheet:
        raise HTTPException(status_code=404, detail="Registro de timesheet no encontrado")
    
    await db.timesheet.delete_one({"timesheet_id": timesheet_id})
    return {"message": "Registro de timesheet eliminado exitosamente"}

@api_router.post("/clock/in", response_model=ClockEntry)
async def clock_in(
    project_id: str = Query(...),
    latitude: float = Query(...),
    longitude: float = Query(...),
    address: Optional[str] = Query(None),
    notes: Optional[str] = Query(None),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Validate location data (mandatory)
    if latitude is None or longitude is None:
        raise HTTPException(
            status_code=400, 
            detail="La ubicación GPS es obligatoria para ponchar. Por favor, permite el acceso a tu ubicación."
        )
    
    # Check if user is already clocked in
    active_clock = await db.clock_entries.find_one({
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    if active_clock:
        raise HTTPException(
            status_code=400, 
            detail=f"Ya tienes un ponche activo en el proyecto {active_clock.get('project_name')}. Debes ponchar salida primero."
        )
    
    # Get project info
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Check if user is assigned to project
    if user.user_id not in project.get('team_members', []) and user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No estás asignado a este proyecto")
    
    # ========== GEOFENCING VALIDATION ==========
    geofence_enabled = project.get('geofence_enabled', False)
    project_lat = project.get('location_latitude')
    project_lon = project.get('location_longitude')
    geofence_radius = project.get('geofence_radius', 100)
    
    # Si el proyecto no tiene geofencing, verificar ubicación general de empresa
    if not geofence_enabled or not project_lat or not project_lon:
        company = await db.company_settings.find_one({}, {"_id": 0})
        if company and company.get('geofence_enabled'):
            geofence_enabled = True
            project_lat = company.get('location_latitude')
            project_lon = company.get('location_longitude')
            geofence_radius = company.get('geofence_radius', 100)
    
    # Validar geofencing si está habilitado
    if geofence_enabled and project_lat and project_lon:
        distance = calculate_distance(latitude, longitude, project_lat, project_lon)
        if distance > geofence_radius:
            raise HTTPException(
                status_code=403,
                detail=f"No puedes ponchar desde esta ubicación. Estás a {int(distance)} metros del área de trabajo permitida (máximo {int(geofence_radius)} metros)."
            )
    
    # Create clock entry
    clock_id = f"clk_{uuid4().hex[:16]}"
    # Usar zona horaria de Puerto Rico en lugar de UTC
    now = datetime.now(PUERTO_RICO_TZ)
    
    clock_doc = {
        "clock_id": clock_id,
        "user_id": user.user_id,
        "user_name": user.name,
        "project_id": project_id,
        "project_name": project.get('name'),
        "clock_in": now.isoformat(),
        "clock_out": None,
        "hours_worked": None,
        "status": "active",
        "date": now.date().isoformat(),
        "notes": notes,
        "clock_in_latitude": latitude,
        "clock_in_longitude": longitude,
        "clock_in_address": address
    }
    
    await db.clock_entries.insert_one(clock_doc)
    
    # Send Slack notification
    await notify_slack_event("clock_in", {
        "user": user.name,
        "project": project.get('name'),
        "time": now.strftime("%H:%M")
    })
    
    return ClockEntry(**clock_doc)

@api_router.post("/clock/out", response_model=ClockEntry)
async def clock_out(
    latitude: float = Query(...),
    longitude: float = Query(...),
    address: Optional[str] = Query(None),
    notes: Optional[str] = Query(None),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    print(f"🔍 Clock OUT - User: {user.user_id} ({user.name})")
    
    # Validate location data (mandatory)
    if latitude is None or longitude is None:
        print(f"❌ Clock OUT - Location missing: lat={latitude}, lon={longitude}")
        raise HTTPException(
            status_code=400, 
            detail="La ubicación GPS es obligatoria para ponchar. Por favor, permite el acceso a tu ubicación."
        )
    
    # Find active clock entry
    active_clock = await db.clock_entries.find_one({
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    print(f"🔍 Clock OUT - Active clock found: {active_clock is not None}")
    if active_clock:
        print(f"🔍 Clock OUT - Clock ID: {active_clock.get('clock_id')}, clock_in: {active_clock.get('clock_in')}, is_manual: {active_clock.get('is_manual')}")
    
    if not active_clock:
        # Debug: list all active clocks for this user
        all_active = await db.clock_entries.find({"user_id": user.user_id}).to_list(10)
        print(f"🔍 Clock OUT - All clocks for user: {len(all_active)}")
        for c in all_active:
            print(f"   - clock_id: {c.get('clock_id')}, status: {c.get('status')}, clock_in: {c.get('clock_in')}")
        raise HTTPException(status_code=400, detail="No tienes un ponche activo")
    
    # Calculate hours worked - handle both timezone-aware and naive datetime strings
    try:
        clock_in_str = active_clock['clock_in']
        # Handle manual entries that might not have timezone info
        if '+' not in clock_in_str and 'Z' not in clock_in_str and len(clock_in_str) <= 19:
            # Naive datetime string like "2026-01-05T08:00:00" - assume Puerto Rico timezone
            naive_dt = datetime.fromisoformat(clock_in_str)
            clock_in_time = PUERTO_RICO_TZ.localize(naive_dt)
        else:
            clock_in_time = datetime.fromisoformat(clock_in_str)
        print(f"✅ Clock OUT - Parsed clock_in_time: {clock_in_time}")
    except Exception as e:
        print(f"❌ Clock OUT - Error parsing clock_in: {active_clock['clock_in']} - {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error procesando hora de entrada: {str(e)}")
    # Usar zona horaria de Puerto Rico en lugar de UTC
    clock_out_time = datetime.now(PUERTO_RICO_TZ)
    hours_worked = (clock_out_time - clock_in_time).total_seconds() / 3600
    
    # Update clock entry
    await db.clock_entries.update_one(
        {"clock_id": active_clock['clock_id']},
        {"$set": {
            "clock_out": clock_out_time.isoformat(),
            "hours_worked": round(hours_worked, 2),
            "status": "completed",
            "notes": notes or active_clock.get('notes'),
            "clock_out_latitude": latitude,
            "clock_out_longitude": longitude,
            "clock_out_address": address
        }}
    )
    
    # Create timesheet entry automatically
    timesheet_id = f"ts_{uuid4().hex[:16]}"
    timesheet_doc = {
        "timesheet_id": timesheet_id,
        "project_id": active_clock['project_id'],
        "project_name": active_clock.get('project_name', ''),
        "user_id": user.user_id,
        "user_name": user.name,
        "date": active_clock['date'],
        "hours_worked": round(hours_worked, 2),
        "description": notes or f"Trabajo registrado mediante ponche ({clock_in_time.strftime('%H:%M')} - {clock_out_time.strftime('%H:%M')})",
        "task_id": None,
        "clock_id": active_clock['clock_id'],
        "created_at": clock_out_time.isoformat(),
        "updated_at": clock_out_time.isoformat()
    }
    
    await db.timesheet.insert_one(timesheet_doc)
    print(f"✅ Timesheet creado automáticamente: {timesheet_id} para proyecto {active_clock['project_id']}")
    
    # Get updated clock entry
    updated_clock = await db.clock_entries.find_one({"clock_id": active_clock['clock_id']}, {"_id": 0})
    
    return ClockEntry(**updated_clock)

@api_router.get("/clock/active", response_model=Optional[ClockEntry])
async def get_active_clock(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    active_clock = await db.clock_entries.find_one({
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    if not active_clock:
        return None
    
    return ClockEntry(**active_clock)

@api_router.get("/clock/history", response_model=List[ClockEntry])
async def get_clock_history(
    date: Optional[str] = None,
    limit: int = 50,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    query = {"user_id": user.user_id}
    if date:
        query["date"] = date
    
    clock_entries = await db.clock_entries.find(
        query, 
        {"_id": 0}
    ).sort("clock_in", -1).limit(limit).to_list(limit)
    
    return [ClockEntry(**entry) for entry in clock_entries]

@api_router.get("/clock/all", response_model=List[ClockEntry])
async def get_all_clock_entries(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    project_id: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    """Get all clock entries (admin/PM only)"""
    user = await get_current_user(request, session_token)
    
    # Admin and PM can see all entries
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="Solo PM, RRHH o administradores pueden ver todos los ponches")
    
    query = {}
    
    if user_id:
        query["user_id"] = user_id
    
    if project_id:
        query["project_id"] = project_id
    
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = start_date
        if end_date:
            query["date"]["$lte"] = end_date
    
    clock_entries = await db.clock_entries.find(
        query, 
        {"_id": 0}
    ).sort("clock_in", -1).to_list(1000)
    
    return [ClockEntry(**entry) for entry in clock_entries]

@api_router.get("/clock/projects", response_model=List[dict])
async def get_assigned_projects(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get projects where user is assigned (for clock-in)"""
    user = await get_current_user(request, session_token)
    
    # Admins can see all active projects
    if user.role == UserRole.SUPER_ADMIN.value:
        projects = await db.projects.find(
            {"status": {"$ne": "completed"}},
            {"_id": 0, "project_id": 1, "name": 1, "status": 1}
        ).to_list(1000)
    else:
        # Regular users see only assigned projects
        projects = await db.projects.find(
            {
                "team_members": user.user_id,
                "status": {"$ne": "completed"}
            },
            {"_id": 0, "project_id": 1, "name": 1, "status": 1}
        ).to_list(1000)
    
    return projects

class ClockEditRequest(BaseModel):
    clock_in: Optional[str] = None
    clock_out: Optional[str] = None
    notes: Optional[str] = None

class ManualClockRequest(BaseModel):
    user_id: str
    project_id: str
    date: str  # YYYY-MM-DD
    clock_in_time: str  # HH:MM
    clock_out_time: Optional[str] = None  # HH:MM
    notes: Optional[str] = None

@api_router.post("/clock/manual")
async def create_manual_clock_entry(
    data: ManualClockRequest,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Create a manual clock entry (admin/RRHH/PM only)"""
    current_user = await get_current_user(request, session_token)
    
    # Only admins, RRHH and PM can create manual entries
    if current_user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo administradores, RRHH o PM pueden crear ponches manuales")
    
    # Validate user exists
    target_user = await db.users.find_one({"user_id": data.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Validate project exists
    project = await db.projects.find_one({"project_id": data.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Build datetime strings WITH timezone for Puerto Rico
    clock_in_naive = datetime.strptime(f"{data.date}T{data.clock_in_time}:00", "%Y-%m-%dT%H:%M:%S")
    clock_in_dt = PUERTO_RICO_TZ.localize(clock_in_naive)
    clock_in_full = clock_in_dt.isoformat()
    
    clock_out_full = None
    clock_out_dt = None
    if data.clock_out_time:
        clock_out_naive = datetime.strptime(f"{data.date}T{data.clock_out_time}:00", "%Y-%m-%dT%H:%M:%S")
        clock_out_dt = PUERTO_RICO_TZ.localize(clock_out_naive)
        clock_out_full = clock_out_dt.isoformat()
    
    # Calculate hours worked
    hours_worked = 0
    status = "active"
    if clock_out_dt:
        hours_worked = round((clock_out_dt - clock_in_dt).total_seconds() / 3600, 2)
        status = "completed"
    
    # Create clock entry
    clock_id = f"clk_{uuid4().hex[:16]}"
    clock_doc = {
        "clock_id": clock_id,
        "user_id": data.user_id,
        "user_name": target_user.get('name', target_user.get('email', 'Sin nombre')),
        "project_id": data.project_id,
        "project_name": project.get('name', 'Sin nombre'),
        "date": data.date,
        "clock_in": clock_in_full,
        "clock_out": clock_out_full,
        "hours_worked": hours_worked,
        "status": status,
        "notes": data.notes or "Entrada manual",
        "is_manual": True,
        "created_by": current_user.user_id,
        "created_by_name": current_user.name,
        "created_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.clock_entries.insert_one(clock_doc)
    
    # Log audit
    await log_audit(
        current_user.user_id,
        current_user.name,
        "create",
        "clock_entry",
        clock_id,
        f"Ponche manual - {target_user.get('name')}",
        {"manual_entry": True, "for_user": target_user.get('name'), "date": data.date}
    )
    
    return ClockEntry(**clock_doc)

@api_router.put("/clock/{clock_id}")
async def update_clock_entry(
    clock_id: str,
    data: ClockEditRequest,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Edit a clock entry (admin/RRHH/PM only)"""
    user = await get_current_user(request, session_token)
    
    # Only admins, RRHH and PM can edit clock entries
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Solo administradores, RRHH o PM pueden editar ponches")
    
    # Find the clock entry
    clock_entry = await db.clock_entries.find_one({"clock_id": clock_id}, {"_id": 0})
    if not clock_entry:
        raise HTTPException(status_code=404, detail="Ponche no encontrado")
    
    # Build update object
    update_data = {"updated_at": datetime.now(PUERTO_RICO_TZ).isoformat()}
    
    if data.clock_in:
        update_data["clock_in"] = data.clock_in
    
    if data.clock_out:
        update_data["clock_out"] = data.clock_out
        # Recalculate hours worked
        clock_in_time = datetime.fromisoformat(data.clock_in if data.clock_in else clock_entry['clock_in'])
        clock_out_time = datetime.fromisoformat(data.clock_out)
        hours_worked = (clock_out_time - clock_in_time).total_seconds() / 3600
        update_data["hours_worked"] = round(hours_worked, 2)
        update_data["status"] = "completed"
    
    if data.notes is not None:
        update_data["notes"] = data.notes
    
    # Update the clock entry
    await db.clock_entries.update_one(
        {"clock_id": clock_id},
        {"$set": update_data}
    )
    
    # Update associated timesheet if exists
    if "hours_worked" in update_data:
        await db.timesheet.update_one(
            {"clock_id": clock_id},
            {"$set": {"hours": update_data["hours_worked"]}}
        )
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "update",
        "clock_entry",
        clock_id,
        {"old": clock_entry, "new": update_data}
    )
    
    # Return updated entry
    updated_entry = await db.clock_entries.find_one({"clock_id": clock_id}, {"_id": 0})
    return ClockEntry(**updated_entry)

@api_router.delete("/clock/{clock_id}")
async def delete_clock_entry(
    clock_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Delete a clock entry (admin and RRHH only)"""
    user = await get_current_user(request, session_token)
    
    # Admin and RRHH can delete clock entries
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="Solo administradores y RRHH pueden eliminar ponches")
    
    # Find the clock entry
    clock_entry = await db.clock_entries.find_one({"clock_id": clock_id}, {"_id": 0})
    if not clock_entry:
        raise HTTPException(status_code=404, detail="Ponche no encontrado")
    
    # Delete the clock entry
    await db.clock_entries.delete_one({"clock_id": clock_id})
    
    # Also delete associated timesheet if exists
    # Try by clock_id first, then by user/date/project match
    deleted_ts = await db.timesheet.delete_one({"clock_id": clock_id})
    if deleted_ts.deleted_count == 0:
        # Try matching by user, date, project and approximate time
        clock_in_time = clock_entry.get('clock_in', '')
        if clock_in_time:
            await db.timesheet.delete_one({
                "user_id": clock_entry.get('user_id'),
                "date": clock_entry.get('date'),
                "project_id": clock_entry.get('project_id'),
                "description": {"$regex": "ponche", "$options": "i"}
            })
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "delete",
        "clock_entry",
        clock_id,
        f"Ponche de {clock_entry.get('user_name', 'Unknown')}",
        {"date": clock_entry.get('date'), "project": clock_entry.get('project_name')}
    )
    
    return {"message": "Ponche eliminado exitosamente"}

@api_router.post("/comments", response_model=Comment)
async def create_comment(comment_data: CommentCreate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    comment_id = f"comm_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    
    comment_doc = {
        "comment_id": comment_id,
        "project_id": comment_data.project_id,
        "user_id": user.user_id,
        "user_name": user.name,
        "content": comment_data.content,
        "timestamp": now
    }
    
    await db.comments.insert_one(comment_doc)
    
    # Notify project owner and team members
    project_doc = await db.projects.find_one({"project_id": comment_data.project_id}, {"_id": 0})
    if project_doc:
        notify_users = [project_doc.get('created_by')] + project_doc.get('team_members', [])
        notify_users = [u for u in notify_users if u != user.user_id]  # Don't notify self
        
        for notify_user in notify_users:
            notification_doc = {
                "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
                "user_id": notify_user,
                "type": "new_comment",
                "message": f"{user.name} comentó en {project_doc['name']}",
                "read": False,
                "timestamp": now,
                "related_id": comment_data.project_id
            }
            await db.notifications.insert_one(notification_doc)
    
    return Comment(**comment_doc)

@api_router.get("/comments", response_model=List[Comment])
async def get_comments(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    comments = await db.comments.find({"project_id": project_id}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    return [Comment(**c) for c in comments]

# ============== PROJECT LOGS ENDPOINTS ==============

@api_router.post("/project-logs", response_model=ProjectLog)
async def create_project_log(
    log_data: ProjectLogCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Get project name
    project = await db.projects.find_one({"project_id": log_data.project_id}, {"_id": 0, "name": 1})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    log_id = f"plog_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    log_doc = {
        "log_id": log_id,
        "project_id": log_data.project_id,
        "project_name": project.get('name'),
        "user_id": user.user_id,
        "user_name": user.name,
        "log_type": log_data.log_type,
        "title": log_data.title,
        "description": log_data.description,
        "hours_worked": log_data.hours_worked,
        "attachments": log_data.attachments or [],
        "created_at": now,
        "updated_at": now
    }
    
    await db.project_logs.insert_one(log_doc)
    
    await log_audit(user.user_id, user.name, "create", "project_log", log_id, log_data.title, {"project_id": log_data.project_id})
    
    return ProjectLog(**log_doc)

@api_router.get("/project-logs", response_model=List[ProjectLog])
async def get_project_logs(
    project_id: Optional[str] = None,
    log_type: Optional[str] = None,
    user_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if log_type:
        query["log_type"] = log_type
    if user_id:
        query["user_id"] = user_id
    if start_date and end_date:
        query["created_at"] = {"$gte": start_date, "$lte": end_date}
    
    logs = await db.project_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [ProjectLog(**log) for log in logs]

@api_router.get("/project-logs/{log_id}", response_model=ProjectLog)
async def get_project_log(
    log_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    log = await db.project_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Log no encontrado")
    
    return ProjectLog(**log)

@api_router.put("/project-logs/{log_id}", response_model=ProjectLog)
async def update_project_log(
    log_id: str,
    log_data: ProjectLogCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    log = await db.project_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Log no encontrado")
    
    # Only creator or admin can edit
    if log.get('user_id') != user.user_id and user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No tienes permiso para editar este log")
    
    update_data = {
        "log_type": log_data.log_type,
        "title": log_data.title,
        "description": log_data.description,
        "hours_worked": log_data.hours_worked,
        "attachments": log_data.attachments or [],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.project_logs.update_one({"log_id": log_id}, {"$set": update_data})
    
    updated = await db.project_logs.find_one({"log_id": log_id}, {"_id": 0})
    return ProjectLog(**updated)

@api_router.delete("/project-logs/{log_id}")
async def delete_project_log(
    log_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    log = await db.project_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Log no encontrado")
    
    # Only creator or admin can delete
    if log.get('user_id') != user.user_id and user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No tienes permiso para eliminar este log")
    
    await db.project_logs.delete_one({"log_id": log_id})
    
    await log_audit(user.user_id, user.name, "delete", "project_log", log_id, log.get('title'), {})
    
    return {"message": "Log eliminado exitosamente"}

@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    notifications = await db.notifications.find({"user_id": user.user_id}, {"_id": 0}).sort("timestamp", -1).to_list(100)
    return [Notification(**n) for n in notifications]

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    await db.notifications.update_one(
        {"notification_id": notification_id, "user_id": user.user_id},
        {"$set": {"read": True}}
    )
    
    return {"message": "Notificación marcada como leída"}

@api_router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.notifications.update_many(
        {"user_id": user.user_id, "read": False},
        {"$set": {"read": True}}
    )
    
    return {"message": f"{result.modified_count} notificaciones marcadas como leídas"}

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Super Admin y Project Manager ven todos los proyectos
    if user.role in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        projects_query = {}
    else:
        projects_query = {"$or": [{"created_by": user.user_id}, {"team_members": user.user_id}]}
    
    total_projects = await db.projects.count_documents(projects_query)
    active_projects = await db.projects.count_documents({**projects_query, "status": ProjectStatus.IN_PROGRESS})
    completed_projects = await db.projects.count_documents({**projects_query, "status": ProjectStatus.COMPLETED})
    
    projects = await db.projects.find(projects_query, {"_id": 0}).to_list(1000)
    total_budget = sum(p.get('budget_total', 0) for p in projects)
    total_spent = sum(p.get('budget_spent', 0) for p in projects)
    total_value = sum(p.get('project_value', 0) for p in projects)
    total_profit = total_value - total_spent
    
    return {
        "total_projects": total_projects,
        "active_projects": active_projects,
        "completed_projects": completed_projects,
        "total_budget": total_budget,
        "total_spent": total_spent,
        "budget_remaining": total_budget - total_spent,
        "total_value": total_value,
        "total_profit": total_profit
    }

@api_router.get("/alerts")
async def get_alerts(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    alerts = []
    
    # 1. Proyectos con presupuesto > 80%
    projects = await db.projects.find({"status": "in_progress"}, {"_id": 0}).to_list(1000)
    for p in projects:
        budget = p.get('budget_total', 0)
        spent = p.get('budget_spent', 0)
        if budget > 0 and (spent / budget) >= 0.8:
            pct = int((spent / budget) * 100)
            alerts.append({
                "type": "budget",
                "severity": "high" if pct >= 100 else "medium",
                "title": f"Presupuesto al {pct}%",
                "message": f"Proyecto '{p.get('name')}' ha consumido {pct}% del presupuesto",
                "link": f"/projects/{p.get('project_id')}"
            })
    
    # 2. Facturas vencidas o por vencer
    today = datetime.now(timezone.utc).date()
    invoices = await db.invoices.find({"status": {"$in": ["pending", "sent"]}}, {"_id": 0}).to_list(1000)
    for inv in invoices:
        due_date_str = inv.get('due_date')
        if due_date_str:
            try:
                due_date = datetime.fromisoformat(due_date_str.replace('Z', '+00:00')).date()
                days_diff = (due_date - today).days
                if days_diff < 0:
                    alerts.append({
                        "type": "invoice",
                        "severity": "high",
                        "title": "Factura vencida",
                        "message": f"Factura #{inv.get('invoice_number')} vencida hace {abs(days_diff)} días",
                        "link": "/invoices"
                    })
                elif days_diff <= 3:
                    alerts.append({
                        "type": "invoice",
                        "severity": "medium",
                        "title": "Factura por vencer",
                        "message": f"Factura #{inv.get('invoice_number')} vence en {days_diff} días",
                        "link": "/invoices"
                    })
            except:
                pass
    
    # 3. Aprobaciones pendientes (solo admin)
    if user.role == UserRole.SUPER_ADMIN.value:
        pending_approvals = await db.approvals.count_documents({"status": "pending"})
        if pending_approvals > 0:
            alerts.append({
                "type": "approval",
                "severity": "medium",
                "title": f"{pending_approvals} aprobaciones pendientes",
                "message": "Hay solicitudes esperando tu aprobación",
                "link": "/approvals"
            })
    
    return alerts

# ==================== APPROVALS SYSTEM ====================
## ==================== SISTEMA DE SOLICITUDES ====================

# Tipos de solicitudes
class RequestCategory(str, Enum):
    EMPLOYEE = "employee"  # Solicitudes de empleado (RRHH aprueba)
    PROJECT = "project"    # Solicitudes de proyecto (PM aprueba)

class RequestType(str, Enum):
    # Employee requests
    VACATION = "vacation"
    PERMISSION = "permission"
    OVERTIME = "overtime"
    SICK_LEAVE = "sick_leave"
    # Project requests
    MATERIAL_PURCHASE = "material_purchase"
    EMERGENCY_EXPENSE = "emergency_expense"
    ADDITIONAL_RESOURCE = "additional_resource"
    EQUIPMENT_RENTAL = "equipment_rental"

REQUEST_CATEGORY_MAP = {
    "vacation": "employee",
    "permission": "employee", 
    "overtime": "employee",
    "sick_leave": "employee",
    "material_purchase": "project",
    "emergency_expense": "project",
    "additional_resource": "project",
    "equipment_rental": "project"
}

@api_router.post("/requests")
async def create_request(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Crear solicitud (empleados pueden crear cualquier tipo)"""
    user = await get_current_user(request, session_token)
    
    request_type = data.get("type")
    category = REQUEST_CATEGORY_MAP.get(request_type, "employee")
    
    new_request = {
        "id": str(uuid4()),
        "category": category,
        "type": request_type,
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        # Employee request fields
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "hours": data.get("hours"),
        # Project request fields
        "project_id": data.get("project_id"),
        "project_name": data.get("project_name"),
        "amount": data.get("amount", 0),
        "urgency": data.get("urgency", "normal"),  # normal, urgent, critical
        # Common fields
        "requested_by": user.user_id,
        "requested_by_name": user.name or "Usuario",
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
        "attachments": data.get("attachments", []),
        "notes": data.get("notes", "")
    }
    
    await db.requests.insert_one(new_request)
    await log_audit(user.user_id, user.name, "create", "request", new_request["id"], data.get("title", "Solicitud"), {"type": request_type, "category": category})
    
    return {"message": "Solicitud creada exitosamente", "id": new_request["id"]}

@api_router.get("/requests")
async def get_requests(
    request: Request, 
    category: Optional[str] = None,
    status: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """Obtener solicitudes según rol del usuario"""
    user = await get_current_user(request, session_token)
    
    query = {}
    
    # Filtrar por categoría si se especifica
    if category:
        query["category"] = category
    
    # Filtrar por estado si se especifica
    if status:
        query["status"] = status
    
    # Super Admin ve todo
    if user.role == UserRole.SUPER_ADMIN.value:
        pass
    # RRHH ve solicitudes de empleados + las propias
    elif user.role == UserRole.RRHH.value:
        query["$or"] = [
            {"category": "employee"},
            {"requested_by": user.user_id}
        ]
    # Project Manager ve solicitudes de proyectos + las propias
    elif user.role == UserRole.PROJECT_MANAGER.value:
        query["$or"] = [
            {"category": "project"},
            {"requested_by": user.user_id}
        ]
    # Empleados solo ven sus propias solicitudes
    else:
        query["requested_by"] = user.user_id
    
    requests_list = await db.requests.find(query, {"_id": 0}).sort("requested_at", -1).to_list(200)
    return requests_list

@api_router.get("/requests/my")
async def get_my_requests(request: Request, session_token: Optional[str] = Cookie(None)):
    """Obtener mis solicitudes"""
    user = await get_current_user(request, session_token)
    requests_list = await db.requests.find({"requested_by": user.user_id}, {"_id": 0}).sort("requested_at", -1).to_list(100)
    return requests_list

@api_router.put("/requests/{request_id}")
async def update_request(request_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Aprobar/Rechazar solicitud"""
    user = await get_current_user(request, session_token)
    
    # Obtener la solicitud
    req = await db.requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    # Verificar permisos según categoría
    category = req.get("category")
    can_approve = False
    
    if user.role == UserRole.SUPER_ADMIN.value:
        can_approve = True
    elif user.role == UserRole.RRHH.value and category == "employee":
        can_approve = True
    elif user.role == UserRole.PROJECT_MANAGER.value and category == "project":
        can_approve = True
    
    if not can_approve:
        raise HTTPException(status_code=403, detail="No tiene permisos para aprobar esta solicitud")
    
    update_data = {
        "status": data.get("status"),  # approved, rejected
        "reviewed_by": user.user_id,
        "reviewed_by_name": user.name or "Usuario",
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "review_notes": data.get("notes", "")
    }
    
    await db.requests.update_one({"id": request_id}, {"$set": update_data})
    await log_audit(user.user_id, user.name, "review", "request", request_id, req.get("title", "Solicitud"), {"status": data.get("status")})
    
    return {"message": f"Solicitud {data.get('status')}"}

@api_router.delete("/requests/{request_id}")
async def delete_request(request_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Cancelar solicitud propia (solo si está pendiente)"""
    user = await get_current_user(request, session_token)
    
    req = await db.requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    # Solo el creador puede cancelar y solo si está pendiente
    if req["requested_by"] != user.user_id and user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No puede cancelar esta solicitud")
    
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Solo se pueden cancelar solicitudes pendientes")
    
    await db.requests.delete_one({"id": request_id})
    return {"message": "Solicitud cancelada"}

## ==================== LEGACY APPROVALS (mantener compatibilidad) ====================

@api_router.post("/approvals")
async def create_approval(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    approval = {
        "id": str(uuid4()),
        "type": data.get("type"),
        "reference_id": data.get("reference_id"),
        "reference_name": data.get("reference_name"),
        "amount": data.get("amount", 0),
        "requested_by": user.user_id,
        "requested_by_name": user.name or "Usuario",
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
        "notes": data.get("notes", "")
    }
    
    await db.approvals.insert_one(approval)
    return {"message": "Solicitud de aprobación creada", "id": approval["id"]}

@api_router.get("/approvals")
async def get_approvals(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value, UserRole.RRHH.value]:
        approvals = await db.approvals.find({}, {"_id": 0}).sort("requested_at", -1).to_list(100)
    else:
        approvals = await db.approvals.find({"requested_by": user.user_id}, {"_id": 0}).sort("requested_at", -1).to_list(100)
    
    return approvals

@api_router.put("/approvals/{approval_id}")
async def update_approval(approval_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="No tiene permisos para aprobar")
    
    update_data = {
        "status": data.get("status"),
        "reviewed_by": user.user_id,
        "reviewed_by_name": user.name or "Usuario",
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "review_notes": data.get("notes", "")
    }
    
    result = await db.approvals.update_one({"id": approval_id}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Aprobación no encontrada")
    
    return {"message": f"Solicitud {data.get('status')}"}

@api_router.delete("/approvals/{approval_id}")
async def delete_approval(approval_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar aprobaciones")
    
    result = await db.approvals.delete_one({"id": approval_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Aprobación no encontrada")
    
    return {"message": "Aprobación eliminada"}

@api_router.get("/projects/{project_id}/stats")
async def get_project_stats(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    tasks = await db.tasks.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    total_tasks = len(tasks)
    completed_tasks = len([t for t in tasks if t.get('status') == TaskStatus.DONE])
    in_progress_tasks = len([t for t in tasks if t.get('status') == TaskStatus.IN_PROGRESS])
    
    categories = await db.budget_categories.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    expenses = await db.expenses.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    project_value = project_doc.get('project_value', 0)
    budget_spent = project_doc.get('budget_spent', 0)
    profit = project_value - budget_spent
    
    return {
        "total_tasks": total_tasks,
        "completed_tasks": completed_tasks,
        "in_progress_tasks": in_progress_tasks,
        "budget_total": project_doc.get('budget_total', 0),
        "budget_spent": budget_spent,
        "budget_remaining": project_doc.get('budget_total', 0) - budget_spent,
        "project_value": project_value,
        "profit": profit,
        "categories": categories,
        "recent_expenses": expenses[-10:] if expenses else []
    }

@api_router.get("/reports/project/{project_id}/export")
async def export_project_report(project_id: str, format: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    tasks = await db.tasks.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    expenses = await db.expenses.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    categories = await db.budget_categories.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen del Proyecto"
        
        title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        header_font = Font(bold=True)
        
        ws['A1'] = f"Reporte del Proyecto: {project_doc['name']}"
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        ws['A3'] = "Información del Proyecto"
        ws['A3'].font = header_font
        ws['A4'] = "Descripción:"
        ws['B4'] = project_doc['description']
        ws['A5'] = "Estado:"
        ws['B5'] = project_doc['status']
        ws['A6'] = "Presupuesto Total:"
        ws['B6'] = f"${project_doc['budget_total']:,.2f}"
        ws['A7'] = "Presupuesto Gastado:"
        ws['B7'] = f"${project_doc.get('budget_spent', 0):,.2f}"
        
        ws['A9'] = "Tareas"
        ws['A9'].font = header_font
        ws['A10'] = "Título"
        ws['B10'] = "Estado"
        ws['C10'] = "Prioridad"
        ws['D10'] = "Progreso"
        for col in ['A10', 'B10', 'C10', 'D10']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        row = 11
        for task in tasks:
            ws[f'A{row}'] = task['title']
            ws[f'B{row}'] = task['status']
            ws[f'C{row}'] = task['priority']
            ws[f'D{row}'] = f"{task['progress']}%"
            row += 1
        
        ws[f'A{row+1}'] = "Gastos"
        ws[f'A{row+1}'].font = header_font
        ws[f'A{row+2}'] = "Descripción"
        ws[f'B{row+2}'] = "Monto"
        ws[f'C{row+2}'] = "Fecha"
        for col in [f'A{row+2}', f'B{row+2}', f'C{row+2}']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        exp_row = row + 3
        for expense in expenses:
            ws[f'A{exp_row}'] = expense['description']
            ws[f'B{exp_row}'] = f"${expense['amount']:,.2f}"
            ws[f'C{exp_row}'] = expense['date']
            exp_row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=proyecto_{project_id}.xlsx"}
        )
    
    elif format == "pdf":
        # Get company info for consistent styling
        company = await db.company.find_one({}, {"_id": 0}) or {}
        
        # Define corporate colors matching frontend
        PRIMARY_COLOR = colors.HexColor('#f97316')  # Orange
        SECONDARY_COLOR = colors.HexColor('#475569')  # Slate
        TEXT_COLOR = colors.HexColor('#1e293b')  # Dark
        LIGHT_BG = colors.HexColor('#f8fafc')  # Light gray
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles matching invoice CSS
        company_style = ParagraphStyle(
            'Company',
            parent=styles['Normal'],
            fontSize=10,
            textColor=PRIMARY_COLOR,
            fontName='Helvetica-Bold'
        )
        company_detail_style = ParagraphStyle(
            'CompanyDetail',
            parent=styles['Normal'],
            fontSize=8,
            textColor=SECONDARY_COLOR
        )
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=TEXT_COLOR,
            spaceAfter=5,
            alignment=2  # Right aligned
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=TEXT_COLOR,
            fontName='Helvetica-Bold',
            spaceAfter=8,
            spaceBefore=15
        )
        normal_right = ParagraphStyle('NormalRight', parent=styles['Normal'], alignment=2)
        
        # Company header
        company_name = company.get('company_name', 'OHSMS PR')
        elements.append(Paragraph(company_name, company_style))
        if company.get('address'):
            elements.append(Paragraph(company.get('address', ''), company_detail_style))
        if company.get('city') or company.get('state'):
            elements.append(Paragraph(f"{company.get('city', '')}, {company.get('state', '')} {company.get('zip_code', '')}", company_detail_style))
        if company.get('phone'):
            elements.append(Paragraph(f"Tel: {company.get('phone', '')}", company_detail_style))
        
        elements.append(Spacer(1, 10))
        
        # Document title
        elements.append(Paragraph("REPORTE DE PROYECTO", title_style))
        elements.append(Paragraph(f"#{project_doc['name']}", normal_right))
        elements.append(Paragraph(f"Fecha: {datetime.now(PUERTO_RICO_TZ).strftime('%d/%m/%Y')}", normal_right))
        elements.append(Spacer(1, 15))
        
        # Info table with new style
        info_data = [
            ['Descripción:', project_doc['description']],
            ['Estado:', project_doc['status']],
            ['Presupuesto Total:', f"${project_doc['budget_total']:,.2f}"],
            ['Presupuesto Gastado:', f"${project_doc.get('budget_spent', 0):,.2f}"]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), LIGHT_BG),
            ('TEXTCOLOR', (0, 0), (-1, -1), TEXT_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LINEBELOW', (0, 0), (-1, -1), 0.3, PRIMARY_COLOR)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table style matching invoice CSS
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, PRIMARY_COLOR),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_COLOR),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fcfcfd')]),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
        ])
        
        elements.append(Paragraph("Tareas", heading_style))
        if tasks:
            task_data = [['Título', 'Estado', 'Prioridad', 'Progreso']]
            for task in tasks:
                task_data.append([
                    task['title'][:30],
                    task['status'],
                    task['priority'],
                    f"{task['progress']}%"
                ])
            
            task_table = Table(task_data, colWidths=[2.5*inch, 1.2*inch, 1*inch, 0.8*inch])
            task_table.setStyle(table_style)
            elements.append(task_table)
        
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("Gastos", heading_style))
        if expenses:
            expense_data = [['Descripción', 'Monto', 'Fecha']]
            for expense in expenses:
                expense_data.append([
                    expense['description'][:40],
                    f"${expense['amount']:,.2f}",
                    expense['date']
                ])
            
            expense_table = Table(expense_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            expense_table.setStyle(table_style)
            elements.append(expense_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=proyecto_{project_id}.pdf"}
        )
    
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado")

@api_router.get("/users")
async def get_users(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Filter out hidden users and sort by name alphabetically
    users = await db.users.find({"hidden": {"$ne": True}}, {"_id": 0, "password": 0}).sort("name", 1).to_list(1000)
    
    # Convert datetime to string and ensure all required fields exist
    for u in users:
        if 'created_at' in u and not isinstance(u['created_at'], str):
            u['created_at'] = u['created_at'].isoformat() if hasattr(u['created_at'], 'isoformat') else str(u['created_at'])
        # Ensure required fields have defaults
        u.setdefault('user_id', u.get('id', ''))
        u.setdefault('name', 'Sin nombre')
        u.setdefault('email', '')
        u.setdefault('role', 'empleado')
        u.setdefault('picture', None)
        u.setdefault('created_at', '')
    
    return users

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    current_user = await get_current_user(request, session_token)
    
    # Only admins can delete users
    if current_user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar usuarios")
    
    # Cannot delete yourself
    if current_user.user_id == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    
    user_to_delete = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Delete user sessions
    await db.user_sessions.delete_many({"user_id": user_id})
    
    # Delete user notifications
    await db.notifications.delete_many({"user_id": user_id})
    
    # Unassign user from all tasks
    await db.tasks.update_many(
        {"assigned_to": user_id},
        {"$set": {"assigned_to": None}}
    )
    
    # Remove user from project teams
    await db.projects.update_many(
        {"team_members": user_id},
        {"$pull": {"team_members": user_id}}
    )
    
    # Delete the user
    await db.users.delete_one({"user_id": user_id})
    
    return {"message": f"Usuario {user_to_delete['name']} eliminado exitosamente"}

@api_router.put("/users/{user_id}/block")
async def block_user(user_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Block or unblock a user. Admin and RRHH can do this."""
    current_user = await get_current_user(request, session_token)
    
    # Admin and RRHH can block users
    if current_user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="No tienes permisos para bloquear usuarios")
    
    # Cannot block yourself
    if current_user.user_id == user_id:
        raise HTTPException(status_code=400, detail="No puedes bloquearte a ti mismo")
    
    user_to_block = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_to_block:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # RRHH cannot block super_admin
    if current_user.role == UserRole.RRHH.value and user_to_block.get("role") == UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No puedes bloquear a un Super Admin")
    
    # Toggle block status
    current_blocked = user_to_block.get("is_blocked", False)
    new_blocked = not current_blocked
    
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"is_blocked": new_blocked, "blocked_at": datetime.now(timezone.utc).isoformat() if new_blocked else None, "blocked_by": current_user.user_id if new_blocked else None}}
    )
    
    # If blocking, invalidate all user sessions
    if new_blocked:
        await db.user_sessions.delete_many({"user_id": user_id})
    
    action = "bloqueado" if new_blocked else "desbloqueado"
    return {"message": f"Usuario {user_to_block['name']} {action} exitosamente", "is_blocked": new_blocked}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate, request: Request, session_token: Optional[str] = Cookie(None)):
    current_user = await get_current_user(request, session_token)
    
    # Super admin and RRHH can update users
    if current_user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="No tienes permisos para actualizar usuarios")
    
    user_to_update = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_to_update:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # RRHH cannot edit super_admin users
    if current_user.role == UserRole.RRHH.value and user_to_update.get("role") == UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No puedes editar a un Super Admin")
    
    update_data = {}
    
    if user_data.name is not None:
        update_data["name"] = user_data.name
    
    if user_data.email is not None:
        # Verificar que el email no esté en uso por otro usuario
        if user_data.email:
            existing = await db.users.find_one({"email": user_data.email, "user_id": {"$ne": user_id}}, {"_id": 0})
            if existing:
                raise HTTPException(status_code=400, detail="Email ya está en uso por otro usuario")
        update_data["email"] = user_data.email
    
    if user_data.password is not None:
        if user_data.password:
            hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            update_data["password"] = hashed_password
    
    if user_data.role is not None:
        # RRHH cannot assign super_admin role
        if current_user.role == UserRole.RRHH.value and user_data.role == UserRole.SUPER_ADMIN.value:
            raise HTTPException(status_code=403, detail="No puedes asignar el rol de Super Admin")
        update_data["role"] = user_data.role
    
    if update_data:
        await db.users.update_one({"user_id": user_id}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    user_without_password = {k: v for k, v in updated_user.items() if k != 'password'}
    return user_without_password

@api_router.get("/settings")
async def get_settings(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Only admins can view settings
    user_role = str(user.role).lower() if user.role else ""
    if user_role not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver la configuración")
    
    # Read directly from .env file to get the latest values
    env_path = Path("/app/backend/.env")
    env_values = {}
    
    if env_path.exists():
        for line in env_path.read_text().split('\n'):
            if '=' in line and not line.startswith('#'):
                key, value = line.split('=', 1)
                # Remove quotes from value
                value = value.strip().strip('"').strip("'")
                env_values[key] = value
    
    return {
        "smtp_host": env_values.get('SMTP_HOST', os.environ.get('SMTP_HOST', 'smtp.gmail.com')),
        "smtp_port": int(env_values.get('SMTP_PORT', os.environ.get('SMTP_PORT', 587))),
        "smtp_user": env_values.get('SMTP_USER', os.environ.get('SMTP_USER', '')),
        "smtp_from_email": env_values.get('SMTP_FROM_EMAIL', os.environ.get('SMTP_FROM_EMAIL', 'noreply@promanage.com')),
        "smtp_from_name": env_values.get('SMTP_FROM_NAME', os.environ.get('SMTP_FROM_NAME', 'ProManage')),
        "email_notifications_enabled": env_values.get('EMAIL_NOTIFICATIONS_ENABLED', os.environ.get('EMAIL_NOTIFICATIONS_ENABLED', 'false')).lower() == 'true',
        "app_url": env_values.get('APP_URL', os.environ.get('APP_URL', 'https://promanage.ohsmspr.com'))
    }

@api_router.put("/settings")
async def update_settings(settings: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Only admins can update settings
    user_role = str(user.role).lower() if user.role else ""
    if user_role not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="No tienes permisos para actualizar la configuración")
    
    # Update .env file
    env_path = Path("/app/backend/.env")
    env_content = env_path.read_text()
    
    if 'smtp_user' in settings:
        env_content = update_env_var(env_content, 'SMTP_USER', settings['smtp_user'])
    if 'smtp_password' in settings:
        env_content = update_env_var(env_content, 'SMTP_PASSWORD', settings['smtp_password'])
    if 'smtp_from_email' in settings:
        env_content = update_env_var(env_content, 'SMTP_FROM_EMAIL', settings['smtp_from_email'])
    if 'smtp_from_name' in settings:
        env_content = update_env_var(env_content, 'SMTP_FROM_NAME', settings['smtp_from_name'])
    if 'smtp_host' in settings:
        env_content = update_env_var(env_content, 'SMTP_HOST', settings['smtp_host'])
    if 'smtp_port' in settings:
        env_content = update_env_var(env_content, 'SMTP_PORT', str(settings['smtp_port']))
    if 'email_notifications_enabled' in settings:
        env_content = update_env_var(env_content, 'EMAIL_NOTIFICATIONS_ENABLED', str(settings['email_notifications_enabled']).lower())
    if 'app_url' in settings:
        env_content = update_env_var(env_content, 'APP_URL', settings['app_url'])
    
    env_path.write_text(env_content)
    
    # Directly update os.environ with the new values
    if 'smtp_user' in settings:
        os.environ['SMTP_USER'] = settings['smtp_user']
    if 'smtp_password' in settings:
        os.environ['SMTP_PASSWORD'] = settings['smtp_password']
    if 'smtp_from_email' in settings:
        os.environ['SMTP_FROM_EMAIL'] = settings['smtp_from_email']
    if 'smtp_from_name' in settings:
        os.environ['SMTP_FROM_NAME'] = settings['smtp_from_name']
    if 'smtp_host' in settings:
        os.environ['SMTP_HOST'] = settings['smtp_host']
    if 'smtp_port' in settings:
        os.environ['SMTP_PORT'] = str(settings['smtp_port'])
    if 'email_notifications_enabled' in settings:
        os.environ['EMAIL_NOTIFICATIONS_ENABLED'] = str(settings['email_notifications_enabled']).lower()
    if 'app_url' in settings:
        os.environ['APP_URL'] = settings['app_url']
    
    # Also reload email service variables
    import email_service
    email_service.SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    email_service.SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
    email_service.SMTP_USER = os.environ.get('SMTP_USER', '')
    email_service.SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
    email_service.SMTP_FROM_EMAIL = os.environ.get('SMTP_FROM_EMAIL', 'noreply@promanage.com')
    email_service.SMTP_FROM_NAME = os.environ.get('SMTP_FROM_NAME', 'ProManage')
    email_service.EMAIL_NOTIFICATIONS_ENABLED = os.environ.get('EMAIL_NOTIFICATIONS_ENABLED', 'false').lower() == 'true'
    
    return {"message": "Configuración de email actualizada exitosamente"}

def update_env_var(content: str, key: str, value: str) -> str:
    """Update or add an environment variable in .env content"""
    lines = content.split('\n')
    updated = False
    
    for i, line in enumerate(lines):
        if line.startswith(f"{key}="):
            lines[i] = f'{key}="{value}"'
            updated = True
            break
    
    if not updated:
        lines.append(f'{key}="{value}"')
    
    return '\n'.join(lines)

# ==================== SAFETY MODULE ====================
# Safety Checklists
class SafetyChecklistStatus(str, Enum):
    DRAFT = "draft"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    ARCHIVED = "archived"

class SafetyObservationType(str, Enum):
    POSITIVE = "positive"
    NEGATIVE = "negative"

class SafetyObservationStatus(str, Enum):
    OPEN = "open"
    IN_PROGRESS = "in_progress"
    RESOLVED = "resolved"
    CLOSED = "closed"

class IncidentSeverity(str, Enum):
    MINOR = "minor"
    MODERATE = "moderate"
    SERIOUS = "serious"
    CRITICAL = "critical"

class IncidentStatus(str, Enum):
    REPORTED = "reported"
    INVESTIGATING = "investigating"
    ACTION_TAKEN = "action_taken"
    CLOSED = "closed"

class ToolboxTalkStatus(str, Enum):
    SCHEDULED = "scheduled"
    COMPLETED = "completed"
    CANCELLED = "cancelled"

# ==================== SAFETY CHECKLISTS ====================
@api_router.get("/safety/checklists")
async def get_safety_checklists(
    request: Request, 
    session_token: Optional[str] = Cookie(None),
    project_id: Optional[str] = None,
    status: Optional[str] = None
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if status:
        query["status"] = status
    
    checklists = await db.safety_checklists.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return checklists

@api_router.get("/safety/checklists/{checklist_id}")
async def get_safety_checklist(checklist_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    checklist = await db.safety_checklists.find_one({"checklist_id": checklist_id}, {"_id": 0})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    return checklist

@api_router.post("/safety/checklists")
async def create_safety_checklist(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    checklist_id = f"checklist_{uuid4().hex[:12]}"
    
    # Process items with default values
    items = []
    for item in data.get("items", []):
        items.append({
            "item_id": f"item_{uuid4().hex[:8]}",
            "description": item.get("description", ""),
            "category": item.get("category", "general"),
            "is_checked": False,
            "notes": "",
            "checked_by": None,
            "checked_at": None
        })
    
    checklist = {
        "checklist_id": checklist_id,
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "project_id": data.get("project_id"),
        "template_id": data.get("template_id"),
        "category": data.get("category", "general"),
        "items": items,
        "status": SafetyChecklistStatus.DRAFT,
        "assigned_to": data.get("assigned_to"),
        "due_date": data.get("due_date"),
        "completion_percentage": 0,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": None
    }
    
    await db.safety_checklists.insert_one(checklist)
    del checklist["_id"]
    
    await log_audit(user.user_id, user.name, "create", "safety_checklist", checklist_id, data.get("title"), {})
    return checklist

@api_router.put("/safety/checklists/{checklist_id}")
async def update_safety_checklist(checklist_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    existing = await db.safety_checklists.find_one({"checklist_id": checklist_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    
    update_data = {
        "title": data.get("title", existing.get("title")),
        "description": data.get("description", existing.get("description")),
        "project_id": data.get("project_id", existing.get("project_id")),
        "category": data.get("category", existing.get("category")),
        "assigned_to": data.get("assigned_to", existing.get("assigned_to")),
        "due_date": data.get("due_date", existing.get("due_date")),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Handle items update
    if "items" in data:
        items = []
        for item in data["items"]:
            items.append({
                "item_id": item.get("item_id") or f"item_{uuid4().hex[:8]}",
                "description": item.get("description", ""),
                "category": item.get("category", "general"),
                "is_checked": item.get("is_checked", False),
                "notes": item.get("notes", ""),
                "checked_by": item.get("checked_by"),
                "checked_at": item.get("checked_at")
            })
        update_data["items"] = items
        
        # Calculate completion percentage
        total = len(items)
        checked = sum(1 for item in items if item.get("is_checked"))
        update_data["completion_percentage"] = round((checked / total * 100) if total > 0 else 0)
        
        # Auto-update status
        if update_data["completion_percentage"] == 100:
            update_data["status"] = SafetyChecklistStatus.COMPLETED
            update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
        elif update_data["completion_percentage"] > 0:
            update_data["status"] = SafetyChecklistStatus.IN_PROGRESS
    
    if "status" in data:
        update_data["status"] = data["status"]
        if data["status"] == SafetyChecklistStatus.COMPLETED:
            update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.safety_checklists.update_one({"checklist_id": checklist_id}, {"$set": update_data})
    
    updated = await db.safety_checklists.find_one({"checklist_id": checklist_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "safety_checklist", checklist_id, update_data.get("title"), {})
    return updated

@api_router.delete("/safety/checklists/{checklist_id}")
async def delete_safety_checklist(checklist_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.safety_checklists.delete_one({"checklist_id": checklist_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    
    await log_audit(user.user_id, user.name, "delete", "safety_checklist", checklist_id, "", {})
    return {"message": "Checklist eliminado"}

@api_router.post("/safety/checklists/{checklist_id}/check-item")
async def check_checklist_item(checklist_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    checklist = await db.safety_checklists.find_one({"checklist_id": checklist_id})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    
    item_id = data.get("item_id")
    is_checked = data.get("is_checked", True)
    notes = data.get("notes", "")
    
    items = checklist.get("items", [])
    for item in items:
        if item["item_id"] == item_id:
            item["is_checked"] = is_checked
            item["notes"] = notes
            item["checked_by"] = user.user_id if is_checked else None
            item["checked_at"] = datetime.now(timezone.utc).isoformat() if is_checked else None
            break
    
    # Calculate completion
    total = len(items)
    checked = sum(1 for item in items if item.get("is_checked"))
    completion_percentage = round((checked / total * 100) if total > 0 else 0)
    
    # Auto-update status
    status = checklist.get("status")
    completed_at = checklist.get("completed_at")
    if completion_percentage == 100:
        status = SafetyChecklistStatus.COMPLETED
        completed_at = datetime.now(timezone.utc).isoformat()
    elif completion_percentage > 0:
        status = SafetyChecklistStatus.IN_PROGRESS
    
    await db.safety_checklists.update_one(
        {"checklist_id": checklist_id},
        {"$set": {
            "items": items,
            "completion_percentage": completion_percentage,
            "status": status,
            "completed_at": completed_at,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.safety_checklists.find_one({"checklist_id": checklist_id}, {"_id": 0})
    return updated

# ==================== SAFETY CHECKLIST TEMPLATES ====================
@api_router.get("/safety/templates")
async def get_safety_templates(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    templates = await db.safety_templates.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return templates

@api_router.post("/safety/templates")
async def create_safety_template(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    template_id = f"template_{uuid4().hex[:12]}"
    
    template = {
        "template_id": template_id,
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "category": data.get("category", "general"),
        "items": data.get("items", []),
        "is_default": data.get("is_default", False),
        "created_by": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.safety_templates.insert_one(template)
    del template["_id"]
    return template

@api_router.delete("/safety/templates/{template_id}")
async def delete_safety_template(template_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.safety_templates.delete_one({"template_id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template no encontrado")
    
    return {"message": "Template eliminado"}

# ==================== SAFETY OBSERVATIONS ====================
@api_router.get("/safety/observations")
async def get_safety_observations(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    project_id: Optional[str] = None,
    observation_type: Optional[str] = None,
    status: Optional[str] = None
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if observation_type:
        query["observation_type"] = observation_type
    if status:
        query["status"] = status
    
    observations = await db.safety_observations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return observations

@api_router.get("/safety/observations/{observation_id}")
async def get_safety_observation(observation_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    observation = await db.safety_observations.find_one({"observation_id": observation_id}, {"_id": 0})
    if not observation:
        raise HTTPException(status_code=404, detail="Observación no encontrada")
    return observation

@api_router.post("/safety/observations")
async def create_safety_observation(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    observation_id = f"obs_{uuid4().hex[:12]}"
    
    observation = {
        "observation_id": observation_id,
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "project_id": data.get("project_id"),
        "location": data.get("location", ""),
        "observation_type": data.get("observation_type", SafetyObservationType.POSITIVE),
        "category": data.get("category", "general"),
        "status": SafetyObservationStatus.OPEN,
        "priority": data.get("priority", Priority.MEDIUM),
        "assigned_to": data.get("assigned_to"),
        "corrective_action": data.get("corrective_action", ""),
        "due_date": data.get("due_date"),
        "photos": data.get("photos", []),
        "media": [],  # New field for photos/videos
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "resolved_at": None
    }
    
    await db.safety_observations.insert_one(observation)
    del observation["_id"]
    
    await log_audit(user.user_id, user.name, "create", "safety_observation", observation_id, data.get("title"), {})
    return observation

@api_router.put("/safety/observations/{observation_id}")
async def update_safety_observation(observation_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    existing = await db.safety_observations.find_one({"observation_id": observation_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Observación no encontrada")
    
    update_data = {k: v for k, v in data.items() if v is not None and k != "observation_id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if data.get("status") == SafetyObservationStatus.RESOLVED:
        update_data["resolved_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.safety_observations.update_one({"observation_id": observation_id}, {"$set": update_data})
    
    updated = await db.safety_observations.find_one({"observation_id": observation_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "safety_observation", observation_id, data.get("title", ""), {})
    return updated

@api_router.delete("/safety/observations/{observation_id}")
async def delete_safety_observation(observation_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.safety_observations.delete_one({"observation_id": observation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Observación no encontrada")
    
    await log_audit(user.user_id, user.name, "delete", "safety_observation", observation_id, "", {})
    return {"message": "Observación eliminada"}

# ==================== TOOLBOX TALKS ====================
@api_router.get("/safety/toolbox-talks")
async def get_toolbox_talks(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    project_id: Optional[str] = None,
    status: Optional[str] = None
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if status:
        query["status"] = status
    
    talks = await db.toolbox_talks.find(query, {"_id": 0}).sort("scheduled_date", -1).to_list(1000)
    return talks

@api_router.get("/safety/toolbox-talks/{talk_id}")
async def get_toolbox_talk(talk_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    talk = await db.toolbox_talks.find_one({"talk_id": talk_id}, {"_id": 0})
    if not talk:
        raise HTTPException(status_code=404, detail="Toolbox Talk no encontrado")
    return talk

@api_router.post("/safety/toolbox-talks")
async def create_toolbox_talk(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    talk_id = f"talk_{uuid4().hex[:12]}"
    
    talk = {
        "talk_id": talk_id,
        "title": data.get("title", ""),
        "topic": data.get("topic", ""),
        "description": data.get("description", ""),
        "key_points": data.get("key_points", []),
        "quiz_questions": data.get("quiz_questions", []),
        "category": data.get("category", "general"),
        "project_id": data.get("project_id"),
        "scheduled_date": data.get("scheduled_date"),
        "duration_minutes": data.get("duration_minutes", 15),
        "location": data.get("location", ""),
        "presenter": data.get("presenter") or user.user_id,
        "presenter_name": data.get("presenter_name") or user.name,
        "status": ToolboxTalkStatus.SCHEDULED,
        "attendees": data.get("attendees", []),
        "attendance_records": [],
        "external_attendee_count": 0,
        "media": [],  # New field for photos/videos
        "materials": data.get("materials", []),
        "notes": data.get("notes", ""),
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": None
    }
    
    await db.toolbox_talks.insert_one(talk)
    del talk["_id"]
    
    await log_audit(user.user_id, user.name, "create", "toolbox_talk", talk_id, data.get("title"), {})
    return talk

# ==================== TOOLBOX TALK TOPICS LIBRARY ====================
TOOLBOX_TALK_TOPICS = [
    {
        "topic_id": "altura_arnes",
        "title": "Seguridad en Alturas y Uso de Arnés",
        "category": "altura",
        "duration_minutes": 15,
        "description": "Charla sobre los riesgos de trabajar en alturas y el uso correcto del equipo de protección contra caídas.",
        "key_points": [
            "Todo trabajo a más de 1.8 metros requiere protección contra caídas",
            "Inspeccionar el arnés antes de cada uso: costuras, hebillas, anillos D",
            "El punto de anclaje debe soportar mínimo 2,268 kg (5,000 lbs)",
            "Ajustar correctamente las correas: pecho, piernas y hombros",
            "Nunca usar un arnés que haya detenido una caída",
            "Mantener el cabo de vida lo más corto posible",
            "Conocer el factor de caída y la distancia de desaceleración"
        ],
        "quiz_questions": [
            "¿A partir de qué altura se requiere protección contra caídas?",
            "¿Qué elementos del arnés debemos inspeccionar antes de usarlo?",
            "¿Cuánto peso mínimo debe soportar un punto de anclaje?",
            "¿Se puede reutilizar un arnés después de una caída?"
        ]
    },
    {
        "topic_id": "epp_general",
        "title": "Uso Correcto del Equipo de Protección Personal (EPP)",
        "category": "epp",
        "duration_minutes": 15,
        "description": "Importancia y uso adecuado de los diferentes tipos de equipo de protección personal en el área de trabajo.",
        "key_points": [
            "El EPP es la última línea de defensa, no la primera",
            "Casco: debe ajustar correctamente, reemplazar si tiene grietas",
            "Gafas de seguridad: usar siempre que haya riesgo de proyección",
            "Guantes: seleccionar según el riesgo (cortes, químicos, calor)",
            "Calzado de seguridad: punta de acero, suela antideslizante",
            "Protección auditiva: obligatoria en áreas con más de 85 dB",
            "Inspeccionar el EPP antes de cada uso y reportar defectos"
        ],
        "quiz_questions": [
            "¿Es el EPP la primera o última línea de defensa?",
            "¿Cuándo debemos reemplazar un casco de seguridad?",
            "¿A partir de cuántos decibeles es obligatoria la protección auditiva?",
            "¿Qué tipo de guante usaría para manipular químicos?"
        ]
    },
    {
        "topic_id": "materiales_peligrosos",
        "title": "Manejo de Materiales Peligrosos",
        "category": "quimicos",
        "duration_minutes": 20,
        "description": "Procedimientos seguros para el manejo, almacenamiento y disposición de materiales peligrosos.",
        "key_points": [
            "Conocer las Hojas de Datos de Seguridad (SDS) de cada producto",
            "Identificar los pictogramas de peligro en las etiquetas",
            "Usar el EPP específico indicado en la SDS",
            "Nunca mezclar productos químicos sin autorización",
            "Almacenar en áreas ventiladas y separar incompatibles",
            "Conocer la ubicación de duchas de emergencia y lavaojos",
            "Reportar inmediatamente cualquier derrame"
        ],
        "quiz_questions": [
            "¿Dónde encontramos información sobre el manejo seguro de un químico?",
            "¿Qué debemos hacer antes de usar un producto químico nuevo?",
            "¿Qué hacer en caso de un derrame químico?",
            "¿Se pueden almacenar todos los químicos juntos?"
        ]
    },
    {
        "topic_id": "prevencion_caidas",
        "title": "Prevención de Caídas al Mismo Nivel",
        "category": "general",
        "duration_minutes": 10,
        "description": "Identificación y prevención de riesgos de resbalones, tropiezos y caídas en el área de trabajo.",
        "key_points": [
            "Las caídas al mismo nivel son una de las principales causas de lesiones",
            "Mantener las áreas de trabajo limpias y ordenadas",
            "Limpiar inmediatamente cualquier derrame de líquidos",
            "No correr en áreas de trabajo",
            "Usar calzado apropiado con suela antideslizante",
            "Mantener cables y mangueras organizados",
            "Reportar condiciones inseguras: pisos dañados, iluminación deficiente"
        ],
        "quiz_questions": [
            "¿Cuál es una de las principales causas de lesiones en el trabajo?",
            "¿Qué debemos hacer si vemos un derrame de líquido?",
            "¿Qué tipo de calzado debemos usar?",
            "¿Por qué es importante reportar pisos dañados?"
        ]
    },
    {
        "topic_id": "seguridad_electrica",
        "title": "Seguridad Eléctrica",
        "category": "electrico",
        "duration_minutes": 15,
        "description": "Riesgos eléctricos y medidas de prevención para evitar accidentes por contacto con electricidad.",
        "key_points": [
            "Solo personal calificado puede trabajar en sistemas eléctricos",
            "Asumir que todo circuito está energizado hasta verificar lo contrario",
            "Usar procedimientos de bloqueo/etiquetado (LOTO)",
            "No usar equipos eléctricos con cables dañados",
            "Mantener distancia de seguridad con líneas aéreas",
            "No sobrecargar tomacorrientes ni usar extensiones permanentes",
            "En caso de electrocución: no tocar a la víctima, cortar la energía"
        ],
        "quiz_questions": [
            "¿Quién puede realizar trabajos eléctricos?",
            "¿Qué es el procedimiento LOTO?",
            "¿Qué debemos hacer antes de trabajar en un circuito?",
            "¿Qué hacer si alguien está siendo electrocutado?"
        ]
    },
    {
        "topic_id": "espacios_confinados",
        "title": "Trabajo en Espacios Confinados",
        "category": "especial",
        "duration_minutes": 20,
        "description": "Procedimientos y precauciones para trabajar de forma segura en espacios confinados.",
        "key_points": [
            "Espacio confinado: entrada limitada, no diseñado para ocupación continua",
            "Siempre se requiere permiso de trabajo para entrar",
            "Monitorear atmósfera: oxígeno, gases combustibles, gases tóxicos",
            "Ventilar adecuadamente antes y durante el trabajo",
            "Tener un vigía en el exterior en todo momento",
            "Contar con equipo de rescate disponible",
            "No entrar si las condiciones no son seguras"
        ],
        "quiz_questions": [
            "¿Qué es un espacio confinado?",
            "¿Qué se debe monitorear antes de entrar?",
            "¿Qué función tiene el vigía?",
            "¿Se puede entrar sin permiso en una emergencia?"
        ]
    },
    {
        "topic_id": "manejo_cargas",
        "title": "Manejo Manual de Cargas",
        "category": "ergonomia",
        "duration_minutes": 15,
        "description": "Técnicas correctas de levantamiento y transporte manual de cargas para prevenir lesiones.",
        "key_points": [
            "Evaluar el peso antes de levantar: ¿necesito ayuda?",
            "Planificar la ruta y despejar obstáculos",
            "Pies separados al ancho de los hombros, uno ligeramente adelante",
            "Doblar las rodillas, no la espalda",
            "Mantener la carga cerca del cuerpo",
            "No girar el tronco mientras se carga",
            "Usar ayudas mecánicas cuando sea posible"
        ],
        "quiz_questions": [
            "¿Qué parte del cuerpo debemos doblar al levantar?",
            "¿Dónde debemos mantener la carga al transportarla?",
            "¿Es correcto girar el tronco mientras cargamos algo pesado?",
            "¿Cuándo debemos usar ayudas mecánicas?"
        ]
    },
    {
        "topic_id": "prevencion_incendios",
        "title": "Prevención y Combate de Incendios",
        "category": "emergencia",
        "duration_minutes": 15,
        "description": "Prevención de incendios, uso de extintores y procedimientos de evacuación.",
        "key_points": [
            "Triángulo del fuego: combustible, oxígeno, calor",
            "Mantener áreas libres de materiales combustibles innecesarios",
            "Conocer la ubicación de extintores y salidas de emergencia",
            "Técnica PASS: Tirar seguro, Apuntar, Apretar, Barrer",
            "Tipos de extintores: A (sólidos), B (líquidos), C (eléctricos)",
            "Evacuar si el fuego no se puede controlar en 30 segundos",
            "No usar ascensores durante evacuación"
        ],
        "quiz_questions": [
            "¿Cuáles son los tres elementos del triángulo del fuego?",
            "¿Qué significa la técnica PASS?",
            "¿Qué tipo de extintor se usa para fuegos eléctricos?",
            "¿Cuándo debemos evacuar en lugar de combatir el fuego?"
        ]
    },
    {
        "topic_id": "primeros_auxilios",
        "title": "Primeros Auxilios Básicos",
        "category": "emergencia",
        "duration_minutes": 15,
        "description": "Procedimientos básicos de primeros auxilios para responder a emergencias médicas.",
        "key_points": [
            "Proteger: asegurar que la escena sea segura",
            "Avisar: llamar a emergencias y reportar al supervisor",
            "Socorrer: solo si estás capacitado para hacerlo",
            "Heridas: lavar con agua limpia, aplicar presión si hay sangrado",
            "Quemaduras: enfriar con agua corriente por 10-20 minutos",
            "Conocer ubicación del botiquín y AED",
            "No mover a víctimas con posibles lesiones de columna"
        ],
        "quiz_questions": [
            "¿Cuál es el primer paso ante una emergencia?",
            "¿Cómo debemos tratar una quemadura?",
            "¿Cuándo NO debemos mover a una persona lesionada?",
            "¿Dónde está el botiquín más cercano a tu área de trabajo?"
        ]
    },
    {
        "topic_id": "senalizacion",
        "title": "Señalización de Seguridad",
        "category": "general",
        "duration_minutes": 10,
        "description": "Tipos de señales de seguridad, su significado y la importancia de respetarlas.",
        "key_points": [
            "Rojo: prohibición, peligro, equipo contra incendios",
            "Amarillo: advertencia, precaución",
            "Verde: seguridad, primeros auxilios, salidas",
            "Azul: obligación, información",
            "Las señales deben estar visibles y en buen estado",
            "Reportar señales dañadas o faltantes",
            "Respetar todas las señales, incluso las temporales"
        ],
        "quiz_questions": [
            "¿Qué indica una señal de color rojo?",
            "¿Qué color indica obligación?",
            "¿De qué color son las señales de salida de emergencia?",
            "¿Qué hacer si vemos una señal dañada?"
        ]
    },
    {
        "topic_id": "excavaciones",
        "title": "Seguridad en Excavaciones y Zanjas",
        "category": "especial",
        "duration_minutes": 20,
        "description": "Riesgos y medidas de seguridad para trabajos en excavaciones y zanjas.",
        "key_points": [
            "Excavaciones de más de 1.2m requieren protección contra derrumbes",
            "Identificar servicios subterráneos antes de excavar",
            "Inspeccionar diariamente y después de lluvias",
            "Mantener materiales y equipo alejados del borde (mínimo 60cm)",
            "Proveer medios de acceso cada 7.5 metros",
            "No trabajar bajo cargas suspendidas",
            "Usar sistemas de entibación, talud o escudo"
        ],
        "quiz_questions": [
            "¿A partir de qué profundidad se requiere protección?",
            "¿Qué debemos hacer antes de empezar a excavar?",
            "¿A qué distancia del borde deben estar los materiales?",
            "¿Cada cuántos metros debe haber un acceso a la zanja?"
        ]
    },
    {
        "topic_id": "trabajo_caliente",
        "title": "Trabajo en Caliente",
        "category": "especial",
        "duration_minutes": 15,
        "description": "Procedimientos seguros para soldadura, corte y otras operaciones con llama o chispa.",
        "key_points": [
            "Trabajo en caliente: cualquier operación que genere llama, chispa o calor",
            "Requiere permiso de trabajo en caliente",
            "Retirar o cubrir materiales combustibles en radio de 10 metros",
            "Tener extintor disponible en el área",
            "Usar pantallas protectoras para proteger a otros",
            "Vigía de fuego durante y 30 minutos después del trabajo",
            "Verificar que cilindros estén asegurados y válvulas cerradas al terminar"
        ],
        "quiz_questions": [
            "¿Qué se considera trabajo en caliente?",
            "¿Qué radio debe estar libre de combustibles?",
            "¿Cuánto tiempo debe permanecer el vigía después del trabajo?",
            "¿Qué equipo de emergencia debe estar disponible?"
        ]
    },
    {
        "topic_id": "orden_limpieza",
        "title": "Orden y Limpieza (5S)",
        "category": "general",
        "duration_minutes": 10,
        "description": "Metodología 5S para mantener un área de trabajo segura, limpia y organizada.",
        "key_points": [
            "Seiri (Clasificar): separar lo necesario de lo innecesario",
            "Seiton (Ordenar): un lugar para cada cosa y cada cosa en su lugar",
            "Seiso (Limpiar): mantener el área de trabajo limpia",
            "Seiketsu (Estandarizar): crear procedimientos y mantenerlos",
            "Shitsuke (Disciplina): convertir las 5S en un hábito",
            "El desorden causa accidentes: tropiezos, golpes, caídas",
            "Limpiar es inspeccionar: detectar anomalías mientras limpiamos"
        ],
        "quiz_questions": [
            "¿Qué significan las 5S?",
            "¿Por qué el orden y limpieza previenen accidentes?",
            "¿Qué hacemos en la etapa de 'Clasificar'?",
            "¿Qué significa 'limpiar es inspeccionar'?"
        ]
    },
    {
        "topic_id": "andamios",
        "title": "Seguridad en Andamios",
        "category": "altura",
        "duration_minutes": 15,
        "description": "Inspección, montaje y uso seguro de andamios en el lugar de trabajo.",
        "key_points": [
            "Solo personal competente puede montar/desmontar andamios",
            "Inspeccionar antes de cada uso: base, marcos, plataformas",
            "Base firme y nivelada, usar placas base",
            "Barandillas en los 4 lados si altura supera 2 metros",
            "Acceso seguro mediante escalera integrada",
            "Capacidad de carga: nunca exceder el límite",
            "No usar andamios durante tormentas eléctricas o vientos fuertes"
        ],
        "quiz_questions": [
            "¿Quién puede montar y desmontar andamios?",
            "¿A partir de qué altura se requieren barandillas?",
            "¿Qué debemos verificar en la base del andamio?",
            "¿Cuándo NO debemos usar andamios?"
        ]
    },
    {
        "topic_id": "ruido_ocupacional",
        "title": "Protección Contra el Ruido",
        "category": "epp",
        "duration_minutes": 10,
        "description": "Efectos del ruido en la salud y uso correcto de protección auditiva.",
        "key_points": [
            "La pérdida auditiva por ruido es permanente e irreversible",
            "Límite permisible: 85 dB por 8 horas de exposición",
            "Regla práctica: si hay que gritar para comunicarse, hay exceso de ruido",
            "Tipos de protección: tapones, orejeras, tapones moldeados",
            "Insertar tapones correctamente: tirar oreja hacia arriba y atrás",
            "NRR (Nivel de Reducción de Ruido): verificar en el empaque",
            "Usar protección doble en ruidos extremos"
        ],
        "quiz_questions": [
            "¿Es reversible la pérdida auditiva por ruido?",
            "¿Cuál es el límite de ruido permisible?",
            "¿Cómo sabemos si hay exceso de ruido sin medirlo?",
            "¿Cómo se insertan correctamente los tapones?"
        ]
    }
]

@api_router.get("/safety/toolbox-topics")
async def get_toolbox_talk_topics(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    category: Optional[str] = None
):
    """Get predefined toolbox talk topics library"""
    user = await get_current_user(request, session_token)
    
    topics = TOOLBOX_TALK_TOPICS.copy()
    
    if category:
        topics = [t for t in topics if t.get("category") == category]
    
    return topics

@api_router.get("/safety/toolbox-topics/{topic_id}")
async def get_toolbox_talk_topic(
    topic_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get a specific toolbox talk topic by ID"""
    user = await get_current_user(request, session_token)
    
    topic = next((t for t in TOOLBOX_TALK_TOPICS if t["topic_id"] == topic_id), None)
    if not topic:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    
    return topic

@api_router.get("/safety/toolbox-categories")
async def get_toolbox_categories(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get list of toolbox talk categories"""
    user = await get_current_user(request, session_token)
    
    categories = {
        "general": "General",
        "altura": "Trabajo en Alturas",
        "epp": "Equipo de Protección Personal",
        "quimicos": "Materiales Peligrosos",
        "electrico": "Seguridad Eléctrica",
        "especial": "Trabajos Especiales",
        "ergonomia": "Ergonomía",
        "emergencia": "Emergencias"
    }
    
    return categories

@api_router.put("/safety/toolbox-talks/{talk_id}")
async def update_toolbox_talk(talk_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    existing = await db.toolbox_talks.find_one({"talk_id": talk_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Toolbox Talk no encontrado")
    
    update_data = {k: v for k, v in data.items() if v is not None and k != "talk_id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if data.get("status") == ToolboxTalkStatus.COMPLETED:
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.toolbox_talks.update_one({"talk_id": talk_id}, {"$set": update_data})
    
    updated = await db.toolbox_talks.find_one({"talk_id": talk_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "toolbox_talk", talk_id, data.get("title", ""), {})
    return updated

@api_router.post("/safety/toolbox-talks/{talk_id}/attendance")
async def record_toolbox_attendance(talk_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    talk = await db.toolbox_talks.find_one({"talk_id": talk_id})
    if not talk:
        raise HTTPException(status_code=404, detail="Toolbox Talk no encontrado")
    
    attendance_record = {
        "user_id": data.get("user_id"),
        "user_name": data.get("user_name"),
        "signature": data.get("signature", ""),
        "attended_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.toolbox_talks.update_one(
        {"talk_id": talk_id},
        {"$push": {"attendance_records": attendance_record}}
    )
    
    updated = await db.toolbox_talks.find_one({"talk_id": talk_id}, {"_id": 0})
    return updated

@api_router.delete("/safety/toolbox-talks/{talk_id}")
async def delete_toolbox_talk(talk_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.toolbox_talks.delete_one({"talk_id": talk_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Toolbox Talk no encontrado")
    
    await log_audit(user.user_id, user.name, "delete", "toolbox_talk", talk_id, "", {})
    return {"message": "Toolbox Talk eliminado"}

# ==================== INCIDENTS ====================
@api_router.get("/safety/incidents")
async def get_incidents(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if status:
        query["status"] = status
    if severity:
        query["severity"] = severity
    
    incidents = await db.safety_incidents.find(query, {"_id": 0}).sort("incident_date", -1).to_list(1000)
    return incidents

@api_router.get("/safety/incidents/{incident_id}")
async def get_incident(incident_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    incident = await db.safety_incidents.find_one({"incident_id": incident_id}, {"_id": 0})
    if not incident:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    return incident

@api_router.post("/safety/incidents")
async def create_incident(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    incident_id = f"incident_{uuid4().hex[:12]}"
    
    incident = {
        "incident_id": incident_id,
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "project_id": data.get("project_id"),
        "incident_date": data.get("incident_date", datetime.now(timezone.utc).isoformat()),
        "incident_time": data.get("incident_time", ""),
        "location": data.get("location", ""),
        "incident_type": data.get("incident_type", "other"),
        "severity": data.get("severity", IncidentSeverity.MINOR),
        "status": IncidentStatus.REPORTED,
        "persons_involved": data.get("persons_involved", []),
        "witnesses": data.get("witnesses", []),
        "injuries_description": data.get("injuries_description", ""),
        "property_damage": data.get("property_damage", ""),
        "immediate_actions": data.get("immediate_actions", ""),
        "root_cause": data.get("root_cause", ""),
        "corrective_actions": data.get("corrective_actions", []),
        "preventive_actions": data.get("preventive_actions", []),
        "photos": data.get("photos", []),
        "documents": data.get("documents", []),
        "media": [],  # New field for photos/videos
        "reported_by": user.user_id,
        "reported_by_name": user.name,
        "assigned_to": data.get("assigned_to"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "closed_at": None
    }
    
    await db.safety_incidents.insert_one(incident)
    del incident["_id"]
    
    await log_audit(user.user_id, user.name, "create", "safety_incident", incident_id, data.get("title"), {})
    return incident

@api_router.put("/safety/incidents/{incident_id}")
async def update_incident(incident_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    existing = await db.safety_incidents.find_one({"incident_id": incident_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    
    update_data = {k: v for k, v in data.items() if v is not None and k != "incident_id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if data.get("status") == IncidentStatus.CLOSED:
        update_data["closed_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.safety_incidents.update_one({"incident_id": incident_id}, {"$set": update_data})
    
    updated = await db.safety_incidents.find_one({"incident_id": incident_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "safety_incident", incident_id, data.get("title", ""), {})
    return updated

@api_router.delete("/safety/incidents/{incident_id}")
async def delete_incident(incident_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    result = await db.safety_incidents.delete_one({"incident_id": incident_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    
    await log_audit(user.user_id, user.name, "delete", "safety_incident", incident_id, "", {})
    return {"message": "Incidente eliminado"}

# ==================== SAFETY MEDIA UPLOAD ====================
SAFETY_UPLOAD_DIR = Path("/app/backend/uploads/safety")
SAFETY_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/safety/upload")
async def upload_safety_media(
    file: UploadFile = File(...),
    entity_type: str = Query(...),  # observation, toolbox_talk, incident
    entity_id: str = Query(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    """Upload photos/videos for safety module entities"""
    user = await get_current_user(request, session_token)
    
    # Validate file type
    allowed_types = [
        "image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp",
        "video/mp4", "video/quicktime", "video/x-msvideo", "video/webm"
    ]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Solo imágenes (JPG, PNG, GIF, WEBP) y videos (MP4, MOV, AVI, WEBM)")
    
    # Max file size: 50MB
    MAX_SIZE = 50 * 1024 * 1024
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="El archivo excede el límite de 50MB")
    
    # Generate unique filename
    ext = Path(file.filename).suffix or ".jpg"
    filename = f"{entity_type}_{entity_id}_{uuid4().hex[:8]}{ext}"
    file_path = SAFETY_UPLOAD_DIR / filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Create media record
    media_record = {
        "media_id": f"media_{uuid4().hex[:12]}",
        "filename": filename,
        "original_filename": file.filename,
        "file_size": len(content),
        "file_type": file.content_type,
        "media_type": "photo" if file.content_type.startswith("image/") else "video",
        "url": f"/api/safety/media/{filename}",
        "uploaded_by": user.user_id,
        "uploaded_by_name": user.name,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update the corresponding entity
    collection_map = {
        "observation": "safety_observations",
        "toolbox_talk": "toolbox_talks",
        "incident": "safety_incidents"
    }
    id_field_map = {
        "observation": "observation_id",
        "toolbox_talk": "talk_id",
        "incident": "incident_id"
    }
    
    collection_name = collection_map.get(entity_type)
    id_field = id_field_map.get(entity_type)
    
    if not collection_name or not id_field:
        raise HTTPException(status_code=400, detail="Tipo de entidad no válido")
    
    collection = db[collection_name]
    entity = await collection.find_one({id_field: entity_id})
    
    if not entity:
        # Remove uploaded file if entity not found
        file_path.unlink(missing_ok=True)
        raise HTTPException(status_code=404, detail=f"Entidad {entity_type} no encontrada")
    
    # Add media to entity's media list
    await collection.update_one(
        {id_field: entity_id},
        {"$push": {"media": media_record}}
    )
    
    return media_record

@api_router.get("/safety/media/{filename}")
async def get_safety_media(filename: str):
    """Serve safety media files"""
    file_path = SAFETY_UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    return FileResponse(file_path, media_type=media_type)

@api_router.delete("/safety/media/{filename}")
async def delete_safety_media(
    filename: str,
    entity_type: str = Query(...),
    entity_id: str = Query(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    """Delete a safety media file"""
    user = await get_current_user(request, session_token)
    
    file_path = SAFETY_UPLOAD_DIR / filename
    if file_path.exists():
        file_path.unlink()
    
    # Remove from entity
    collection_map = {
        "observation": "safety_observations",
        "toolbox_talk": "toolbox_talks",
        "incident": "safety_incidents"
    }
    id_field_map = {
        "observation": "observation_id",
        "toolbox_talk": "talk_id",
        "incident": "incident_id"
    }
    
    collection_name = collection_map.get(entity_type)
    id_field = id_field_map.get(entity_type)
    
    if collection_name and id_field:
        collection = db[collection_name]
        await collection.update_one(
            {id_field: entity_id},
            {"$pull": {"media": {"filename": filename}}}
        )
    
    return {"message": "Archivo eliminado"}

# ==================== TOOLBOX ATTENDANCE WITH EMPLOYEES ====================
@api_router.post("/safety/toolbox-talks/{talk_id}/attendance-bulk")
async def record_toolbox_attendance_bulk(
    talk_id: str,
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Record attendance for multiple employees and external attendees"""
    user = await get_current_user(request, session_token)
    
    talk = await db.toolbox_talks.find_one({"talk_id": talk_id})
    if not talk:
        raise HTTPException(status_code=404, detail="Toolbox Talk no encontrado")
    
    employee_ids = data.get("employee_ids", [])
    external_count = data.get("external_count", 0)
    external_names = data.get("external_names", [])
    
    attendance_records = []
    
    # Add employee attendance
    for emp_id in employee_ids:
        emp = await db.users.find_one({"user_id": emp_id}, {"_id": 0})
        if emp:
            attendance_records.append({
                "user_id": emp_id,
                "user_name": emp.get("name", "Desconocido"),
                "type": "employee",
                "attended_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Add external attendees
    for i, name in enumerate(external_names):
        attendance_records.append({
            "user_id": f"external_{i}",
            "user_name": name,
            "type": "external",
            "attended_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Add remaining external count without names
    for i in range(len(external_names), external_count):
        attendance_records.append({
            "user_id": f"external_{i}",
            "user_name": f"Externo {i + 1}",
            "type": "external",
            "attended_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Update talk with new attendance records
    await db.toolbox_talks.update_one(
        {"talk_id": talk_id},
        {
            "$set": {
                "attendance_records": attendance_records,
                "external_attendee_count": external_count
            }
        }
    )
    
    updated = await db.toolbox_talks.find_one({"talk_id": talk_id}, {"_id": 0})
    return updated

# ==================== SAFETY DASHBOARD STATS ====================
@api_router.get("/safety/dashboard")
async def get_safety_dashboard(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    project_id: Optional[str] = None
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    
    # Get counts
    checklists = await db.safety_checklists.find(query, {"_id": 0}).to_list(1000)
    observations = await db.safety_observations.find(query, {"_id": 0}).to_list(1000)
    toolbox_talks = await db.toolbox_talks.find(query, {"_id": 0}).to_list(1000)
    incidents = await db.safety_incidents.find(query, {"_id": 0}).to_list(1000)
    
    # Calculate stats
    total_checklists = len(checklists)
    completed_checklists = len([c for c in checklists if c.get("status") == "completed"])
    
    positive_observations = len([o for o in observations if o.get("type") == "positive"])
    negative_observations = len([o for o in observations if o.get("type") == "negative"])
    open_observations = len([o for o in observations if o.get("status") != "resolved"])
    
    scheduled_talks = len([t for t in toolbox_talks if t.get("status") == "scheduled"])
    completed_talks = len([t for t in toolbox_talks if t.get("status") == "completed"])
    
    total_incidents = len(incidents)
    open_incidents = len([i for i in incidents if i.get("status") != "closed"])
    critical_incidents = len([i for i in incidents if i.get("severity") == "critical"])
    
    # Days without incidents
    if incidents:
        try:
            # Get the most recent incident
            last_incident = max(incidents, key=lambda x: x.get("incident_date", ""))
            incident_date_str = last_incident.get("incident_date", "")
            
            # Parse the date string - handle different formats
            if incident_date_str:
                # Try to parse as date only first (YYYY-MM-DD)
                try:
                    last_incident_date = datetime.strptime(incident_date_str, "%Y-%m-%d")
                    last_incident_date = last_incident_date.replace(tzinfo=timezone.utc)
                except ValueError:
                    # Try to parse as ISO format
                    try:
                        last_incident_date = datetime.fromisoformat(incident_date_str.replace("Z", "+00:00"))
                    except ValueError:
                        # Default to 365 days if parsing fails
                        last_incident_date = datetime.now(timezone.utc) - timedelta(days=365)
                
                days_without_incident = (datetime.now(timezone.utc) - last_incident_date).days
            else:
                days_without_incident = 365
        except Exception:
            days_without_incident = 365
    else:
        days_without_incident = 365  # Default if no incidents
    
    return {
        "days_without_incidents": days_without_incident,
        "total_checklists": total_checklists,
        "total_observations": len(observations),
        "total_toolbox_talks": len(toolbox_talks),
        "total_incidents": total_incidents,
        "checklists": {
            "total": total_checklists,
            "completed": completed_checklists,
            "completion_rate": round((completed_checklists / total_checklists * 100) if total_checklists > 0 else 0)
        },
        "observations": {
            "total": len(observations),
            "positive": positive_observations,
            "negative": negative_observations,
            "open": open_observations
        },
        "toolbox_talks": {
            "total": len(toolbox_talks),
            "scheduled": scheduled_talks,
            "completed": completed_talks
        },
        "incidents": {
            "total": total_incidents,
            "open": open_incidents,
            "critical": critical_incidents,
            "days_without_incident": days_without_incident
        },
        "recent_checklists": checklists[:5],
        "recent_observations": observations[:5],
        "recent_incidents": incidents[:5],
        "upcoming_talks": [t for t in toolbox_talks if t.get("status") == "scheduled"][:5]
    }

# ==================== CLIENT PORTAL ====================
@api_router.post("/clients")
async def create_client(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Check if email exists
    existing = await db.users.find_one({"email": data.get("email")})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    
    client_id = f"client_{uuid4().hex[:12]}"
    password_hash = bcrypt.hashpw(data.get("password", "client123").encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    client = {
        "user_id": client_id,
        "name": data.get("nombre_contacto") or data.get("name", ""),
        "email": data.get("email"),
        "password": password_hash,
        "role": UserRole.CLIENT.value,
        "company_name": data.get("empresa") or data.get("company_name", ""),
        "company_address": data.get("company_address", ""),
        "company_phone": data.get("company_phone", ""),
        "company_email": data.get("company_email", ""),
        "contact_person": data.get("contact_person", ""),
        "tax_id": data.get("tax_id", ""),
        "notes": data.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.user_id
    }
    
    await db.users.insert_one(client)
    await log_audit(user.user_id, user.name, "create", "client", client_id, data.get("name"), {})
    return {"message": "Cliente creado", "client_id": client_id}

@api_router.get("/clients")
async def get_clients(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value:
        # Client can only see their own profile
        client = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "password": 0})
        return [client] if client else []
    
    clients = await db.users.find({"role": UserRole.CLIENT.value}, {"_id": 0, "password": 0}).to_list(1000)
    return clients

@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    client = await db.users.find_one({"user_id": client_id, "role": UserRole.CLIENT.value}, {"_id": 0, "password": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return client

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    update_data = {
        "company_name": data.get("company_name"),
        "company_address": data.get("company_address"),
        "company_phone": data.get("company_phone"),
        "company_email": data.get("company_email"),
        "contact_person": data.get("contact_person"),
        "tax_id": data.get("tax_id"),
        "notes": data.get("notes"),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    update_data = {k: v for k, v in update_data.items() if v is not None}
    
    await db.users.update_one({"user_id": client_id}, {"$set": update_data})
    return {"message": "Cliente actualizado"}

# ==================== CLIENT PROFILES (Reusable commercial contacts) ====================
@api_router.get("/client-profiles")
async def get_client_profiles(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all client profiles for reuse in estimates/invoices"""
    user = await get_current_user(request, session_token)
    profiles = await db.client_profiles.find({}, {"_id": 0}).sort("company_name", 1).to_list(1000)
    return profiles

@api_router.post("/client-profiles")
async def create_client_profile(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Create a new client profile for reuse"""
    user = await get_current_user(request, session_token)
    
    profile_id = f"cp_{uuid4().hex[:12]}"
    
    profile = {
        "profile_id": profile_id,
        "company_name": data.get("company_name", ""),
        "contact_name": data.get("contact_name", ""),
        "email": data.get("email", ""),
        "phone": data.get("phone", ""),
        "address": data.get("address", ""),
        "notes": data.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.user_id
    }
    
    await db.client_profiles.insert_one(profile)
    # Remove MongoDB _id before returning
    profile.pop("_id", None)
    return {"message": "Perfil de cliente creado", "profile_id": profile_id, "profile": profile}

@api_router.post("/client-profiles/find-or-create")
async def find_or_create_client_profile(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Find existing client profile by email/company or create new one"""
    user = await get_current_user(request, session_token)
    
    email = data.get("email", "").strip().lower()
    company_name = data.get("company_name", "").strip()
    contact_name = data.get("contact_name", "").strip()
    
    # Try to find by email first (most specific)
    existing = None
    if email:
        existing = await db.client_profiles.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}}, {"_id": 0})
    
    # If not found by email, try by company name
    if not existing and company_name:
        existing = await db.client_profiles.find_one({"company_name": {"$regex": f"^{company_name}$", "$options": "i"}}, {"_id": 0})
    
    if existing:
        return {"found": True, "profile": existing}
    
    # Create new profile
    profile_id = f"cp_{uuid4().hex[:12]}"
    profile = {
        "profile_id": profile_id,
        "company_name": company_name,
        "contact_name": contact_name,
        "email": email,
        "phone": data.get("phone", ""),
        "address": data.get("address", ""),
        "notes": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.user_id
    }
    
    await db.client_profiles.insert_one(profile)
    # Remove MongoDB _id before returning
    profile.pop("_id", None)
    return {"found": False, "profile": profile, "message": "Perfil de cliente creado automáticamente"}

# ============== CLIENT PROFILES DETAIL ENDPOINTS ==============

@api_router.get("/client-profiles/{profile_id}")
async def get_client_profile(profile_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Get a single client profile by ID"""
    user = await get_current_user(request, session_token)
    profile = await db.client_profiles.find_one({"profile_id": profile_id}, {"_id": 0})
    if not profile:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return profile

@api_router.put("/client-profiles/{profile_id}")
async def update_client_profile(profile_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Update a client profile"""
    user = await get_current_user(request, session_token)
    
    existing = await db.client_profiles.find_one({"profile_id": profile_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    update_data = {
        "company_name": data.get("company_name", existing.get("company_name", "")),
        "contact_name": data.get("contact_name", existing.get("contact_name", "")),
        "email": data.get("email", existing.get("email", "")),
        "phone": data.get("phone", existing.get("phone", "")),
        "address": data.get("address", existing.get("address", "")),
        "tax_id": data.get("tax_id", existing.get("tax_id", "")),
        "notes": data.get("notes", existing.get("notes", "")),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user.user_id
    }
    
    await db.client_profiles.update_one({"profile_id": profile_id}, {"$set": update_data})
    return {"message": "Cliente actualizado"}

@api_router.delete("/client-profiles/{profile_id}")
async def delete_client_profile(profile_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Delete a client profile"""
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.ADMIN.value]:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar clientes")
    
    await db.client_profiles.delete_one({"profile_id": profile_id})
    return {"message": "Cliente eliminado"}

# ============== MIGRATION ENDPOINT ==============

@api_router.post("/migrate/estimates-to-client-profiles")
async def migrate_estimates_to_client_profiles(request: Request, session_token: Optional[str] = Cookie(None)):
    """
    Migration endpoint: Links existing estimates to client profiles.
    Creates client profiles if they don't exist based on estimate client info.
    Only super_admin can run this migration.
    """
    import logging
    logger = logging.getLogger("migration")
    
    try:
        logger.info("=== MIGRATION START ===")
        user = await get_current_user(request, session_token)
        logger.info(f"User: {user.name} (role: {user.role})")
        
        if user.role != UserRole.SUPER_ADMIN.value:
            logger.warning(f"Access denied for role: {user.role}")
            raise HTTPException(status_code=403, detail="Solo super administradores pueden ejecutar migraciones")
        
        # Get all estimates without client_profile_id or with null/empty client_profile_id
        logger.info("Fetching estimates without client_profile_id...")
        estimates = await db.estimates.find({
            "$or": [
                {"client_profile_id": {"$exists": False}},
                {"client_profile_id": None},
                {"client_profile_id": ""}
            ]
        }).to_list(10000)
        
        logger.info(f"Found {len(estimates)} estimates to migrate")
        
        migrated_count = 0
        created_profiles = 0
        errors = []
        
        for estimate in estimates:
            try:
                estimate_id = estimate.get("estimate_id")
                client_email = (estimate.get("client_email") or "").strip().lower()
                client_company = (estimate.get("client_company") or "").strip()
                client_name = (estimate.get("client_name") or "").strip()
                
                logger.info(f"Processing estimate {estimate_id}: email={client_email}, company={client_company}, name={client_name}")
                
                if not client_email and not client_company and not client_name:
                    error_msg = f"Estimate {estimate_id}: No client info to match"
                    logger.warning(error_msg)
                    errors.append(error_msg)
                    continue
                
                # Try to find existing client profile
                existing_profile = None
                
                # First try by email (most specific)
                if client_email:
                    existing_profile = await db.client_profiles.find_one({
                        "email": {"$regex": f"^{client_email}$", "$options": "i"}
                    })
                    if existing_profile:
                        logger.info(f"Found profile by email: {existing_profile.get('profile_id')}")
                
                # If not found, try by company name
                if not existing_profile and client_company:
                    existing_profile = await db.client_profiles.find_one({
                        "company_name": {"$regex": f"^{client_company}$", "$options": "i"}
                    })
                    if existing_profile:
                        logger.info(f"Found profile by company: {existing_profile.get('profile_id')}")
                
                # If still not found, create new profile
                if existing_profile:
                    profile_id = existing_profile.get("profile_id")
                else:
                    # Create new client profile
                    profile_id = f"cp_{uuid4().hex[:12]}"
                    new_profile = {
                        "profile_id": profile_id,
                        "company_name": client_company or "",
                        "contact_name": client_name or "",
                        "email": client_email or "",
                        "phone": estimate.get("client_phone") or "",
                        "address": estimate.get("client_address") or "",
                        "notes": "",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "created_by": user.user_id,
                        "migrated_from_estimate": estimate_id
                    }
                    logger.info(f"Creating new profile: {profile_id}")
                    await db.client_profiles.insert_one(new_profile)
                    created_profiles += 1
                
                # Update the estimate with the client_profile_id
                logger.info(f"Updating estimate {estimate_id} with profile_id {profile_id}")
                await db.estimates.update_one(
                    {"estimate_id": estimate_id},
                    {"$set": {"client_profile_id": profile_id}}
                )
                migrated_count += 1
                
            except Exception as e:
                error_msg = f"Estimate {estimate.get('estimate_id', 'unknown')}: {str(e)}"
                logger.error(error_msg)
                import traceback
                logger.error(traceback.format_exc())
                errors.append(error_msg)
        
        logger.info(f"=== MIGRATION COMPLETE: {migrated_count} migrated, {created_profiles} profiles created ===")
        
        return {
            "message": "Migración completada",
            "total_estimates_processed": len(estimates),
            "estimates_migrated": migrated_count,
            "new_profiles_created": created_profiles,
            "errors": errors[:20] if errors else []  # Limit errors to first 20
        }
        
    except HTTPException as he:
        logger.error(f"HTTP Exception: {he.detail}")
        raise he
    except Exception as e:
        logger.error(f"=== MIGRATION ERROR ===")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error message: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error en migración: {str(e)}")

# ============== CLIENT PROFILE ESTIMATES ==============

@api_router.get("/client-profiles/{profile_id}/estimates")
async def get_client_profile_estimates(profile_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all estimates for a client profile"""
    user = await get_current_user(request, session_token)
    estimates = await db.estimates.find({"client_profile_id": profile_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return estimates

# ============== CLIENT PROFILE DOCUMENTS ==============

@api_router.post("/client-profiles/{profile_id}/documents")
async def upload_client_profile_document(profile_id: str, document_type: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Upload a document to a client profile"""
    user = await get_current_user(request, session_token)
    
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No file provided")
    
    content = await file.read()
    file_b64 = base64.b64encode(content).decode()
    
    doc = {
        "document_id": str(uuid4()),
        "profile_id": profile_id,
        "filename": file.filename,
        "document_type": document_type,
        "file_data": file_b64,
        "uploaded_by": user.user_id,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.client_profile_documents.insert_one(doc)
    return {"message": "Documento subido", "document_id": doc["document_id"]}

@api_router.get("/client-profiles/{profile_id}/documents")
async def get_client_profile_documents(profile_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all documents for a client profile"""
    user = await get_current_user(request, session_token)
    docs = await db.client_profile_documents.find({"profile_id": profile_id}, {"_id": 0, "file_data": 0}).to_list(100)
    return docs

@api_router.get("/client-profiles/{profile_id}/documents/{doc_id}/download")
async def download_client_profile_document(profile_id: str, doc_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Download a client profile document"""
    from fastapi.responses import Response
    user = await get_current_user(request, session_token)
    
    doc = await db.client_profile_documents.find_one({"document_id": doc_id, "profile_id": profile_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    content = base64.b64decode(doc["file_data"])
    return Response(content=content, media_type="application/octet-stream", headers={"Content-Disposition": f"attachment; filename={doc['filename']}"})

@api_router.delete("/client-profiles/{profile_id}/documents/{doc_id}")
async def delete_client_profile_document(profile_id: str, doc_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Delete a client profile document"""
    user = await get_current_user(request, session_token)
    await db.client_profile_documents.delete_one({"document_id": doc_id, "profile_id": profile_id})
    return {"message": "Documento eliminado"}

@api_router.get("/clients/{client_id}/projects")
async def get_client_projects(client_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    # Clientes NO pueden ver proyectos - los proyectos son privados de OHSMS
    if user.role == UserRole.CLIENT.value:
        return []
    
    projects = await db.projects.find({"client_id": client_id}, {"_id": 0}).to_list(100)
    return projects

@api_router.post("/clients/{client_id}/documents")
async def upload_client_document(client_id: str, document_type: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No file provided")
    
    content = await file.read()
    file_b64 = base64.b64encode(content).decode()
    
    doc = {
        "document_id": str(uuid4()),
        "client_id": client_id,
        "filename": file.filename,
        "document_type": document_type,
        "file_data": file_b64,
        "uploaded_by": user.user_id,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.client_documents.insert_one(doc)
    return {"message": "Documento subido", "document_id": doc["document_id"]}

@api_router.get("/clients/{client_id}/documents")
async def get_client_documents(client_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    docs = await db.client_documents.find({"client_id": client_id}, {"_id": 0, "file_data": 0}).to_list(100)
    return docs

@api_router.get("/clients/{client_id}/documents/{doc_id}/download")
async def download_client_document(client_id: str, doc_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    from fastapi.responses import Response
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    doc = await db.client_documents.find_one({"document_id": doc_id, "client_id": client_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    content = base64.b64decode(doc["file_data"])
    return Response(content=content, media_type="application/octet-stream", headers={"Content-Disposition": f"attachment; filename={doc['filename']}"})

@api_router.delete("/clients/{client_id}/documents/{doc_id}")
async def delete_client_document(client_id: str, doc_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    if user.role == UserRole.CLIENT.value and user.user_id != client_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    await db.client_documents.delete_one({"document_id": doc_id, "client_id": client_id})
    return {"message": "Documento eliminado"}

# ============== COMPANY SETTINGS ENDPOINTS ==============

@api_router.get("/company")
async def get_company_settings(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    company = await db.company_settings.find_one({}, {"_id": 0})
    if not company:
        company = {
            "company_name": "",
            "company_logo": None,
            "address": "",
            "city": "",
            "state": "",
            "zip_code": "",
            "country": "",
            "phone": "",
            "email": "",
            "website": "",
            "tax_id": "",
            "currency": "USD",
            "footer_text": "",
            "next_invoice_number": 1,
            "next_estimate_number": 1,
            "next_po_number": 1,
            "location_latitude": None,
            "location_longitude": None,
            "geofence_radius": 100,
            "geofence_enabled": False,
            "minimum_margin_percent": 15,
            "default_b2b_percentage": 4,
            "default_cfse_percentage": 7,
            "default_liability_percentage": 7,
            "default_municipal_patent_percentage": 1
        }
    # Asegurar que exista minimum_margin_percent
    if "minimum_margin_percent" not in company:
        company["minimum_margin_percent"] = 15
    # Asegurar que existan los porcentajes de estimación de costos
    if "default_b2b_percentage" not in company:
        company["default_b2b_percentage"] = 4
    if "default_cfse_percentage" not in company:
        company["default_cfse_percentage"] = 7
    if "default_liability_percentage" not in company:
        company["default_liability_percentage"] = 7
    if "default_municipal_patent_percentage" not in company:
        company["default_municipal_patent_percentage"] = 1
    return company

@api_router.put("/company")
async def update_company_settings(
    company_data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo los administradores pueden modificar la configuración de empresa")
    
    existing = await db.company_settings.find_one({})
    
    update_data = {
        "company_name": company_data.get("company_name", ""),
        "address": company_data.get("address", ""),
        "city": company_data.get("city", ""),
        "state": company_data.get("state", ""),
        "zip_code": company_data.get("zip_code", ""),
        "country": company_data.get("country", ""),
        "phone": company_data.get("phone", ""),
        "email": company_data.get("email", ""),
        "website": company_data.get("website", ""),
        "tax_id": company_data.get("tax_id", ""),
        "currency": company_data.get("currency", "USD"),
        "footer_text": company_data.get("footer_text", ""),
        "next_invoice_number": company_data.get("next_invoice_number", 1),
        "next_estimate_number": company_data.get("next_estimate_number", 1),
        "next_po_number": company_data.get("next_po_number", 1),
        "location_latitude": company_data.get("location_latitude"),
        "location_longitude": company_data.get("location_longitude"),
        "geofence_radius": company_data.get("geofence_radius", 100),
        "geofence_enabled": company_data.get("geofence_enabled", False),
        "minimum_margin_percent": company_data.get("minimum_margin_percent", 15),
        "default_b2b_percentage": company_data.get("default_b2b_percentage", 4),
        "default_cfse_percentage": company_data.get("default_cfse_percentage", 7),
        "default_liability_percentage": company_data.get("default_liability_percentage", 7),
        "default_municipal_patent_percentage": company_data.get("default_municipal_patent_percentage", 1),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if existing:
        # Mantener el logo existente si no se envía uno nuevo
        if "company_logo" not in company_data:
            update_data["company_logo"] = existing.get("company_logo")
        else:
            update_data["company_logo"] = company_data.get("company_logo")
        await db.company_settings.update_one({}, {"$set": update_data})
    else:
        update_data["company_logo"] = company_data.get("company_logo")
        update_data["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.company_settings.insert_one(update_data)
    
    await log_audit(user.user_id, user.name, "update", "company_settings", "company", "Configuración de empresa", {})
    
    return {"message": "Configuración de empresa actualizada"}

@api_router.post("/company/logo")
async def upload_company_logo(
    file: UploadFile = File(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Check for admin roles (handle both enum value and string)
    user_role = str(user.role).lower() if user.role else ""
    if user_role not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Solo los administradores pueden subir el logo")
    
    # Validar tipo de archivo
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp", "image/svg+xml"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPG, PNG, GIF, WebP o SVG")
    
    # Crear directorio si no existe
    upload_dir = Path("/app/uploads/logos")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Generar nombre único
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
    filename = f"company_logo_{uuid4().hex[:8]}.{ext}"
    file_path = upload_dir / filename
    
    # Guardar archivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    logo_url = f"/uploads/logos/{filename}"
    
    # Actualizar en base de datos
    await db.company_settings.update_one(
        {},
        {"$set": {"company_logo": logo_url, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"logo_url": logo_url, "message": "Logo subido exitosamente"}

UPLOAD_DIR = Path("/app/backend/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/documents/upload", response_model=Document)
async def upload_document(
    project_id: str,
    file: UploadFile = File(...),
    folder_id: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # PM y Admin tienen acceso completo, otros solo si son creadores o del equipo
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    
    # Validate folder if provided
    if folder_id:
        folder = await db.document_folders.find_one({
            "folder_id": folder_id,
            "project_id": project_id
        })
        if not folder:
            raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    
    max_size = 10 * 1024 * 1024
    file_content = await file.read()
    if len(file_content) > max_size:
        raise HTTPException(status_code=400, detail="El archivo es demasiado grande (máximo 10MB)")
    
    document_id = f"doc_{uuid.uuid4().hex[:12]}"
    file_extension = Path(file.filename).suffix
    filename = f"{document_id}{file_extension}"
    file_path = UPLOAD_DIR / filename
    
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    mime_type = mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    
    document_doc = {
        "document_id": document_id,
        "project_id": project_id,
        "folder_id": folder_id,
        "filename": filename,
        "original_filename": file.filename,
        "file_size": len(file_content),
        "file_type": mime_type,
        "uploaded_by": user.user_id,
        "uploaded_by_name": user.name,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.documents.insert_one(document_doc)
    return Document(**document_doc)

@api_router.get("/documents", response_model=List[Document])
async def get_documents(project_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # PM y Admin tienen acceso completo, otros solo si son creadores o del equipo
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    
    documents = await db.documents.find({"project_id": project_id}, {"_id": 0}).sort("uploaded_at", -1).to_list(1000)
    return [Document(**d) for d in documents]

@api_router.get("/documents/{document_id}/download")
async def download_document(document_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    document_doc = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not document_doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    project_doc = await db.projects.find_one({"project_id": document_doc['project_id']}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este documento")
    
    file_path = UPLOAD_DIR / document_doc['filename']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el servidor")
    
    return FileResponse(
        path=file_path,
        filename=document_doc['original_filename'],
        media_type=document_doc['file_type']
    )

@api_router.delete("/documents/{document_id}")
async def delete_document(document_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    document_doc = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not document_doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    project_doc = await db.projects.find_one({"project_id": document_doc['project_id']}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and document_doc['uploaded_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar este documento")
    
    file_path = UPLOAD_DIR / document_doc['filename']
    if file_path.exists():
        file_path.unlink()
    
    await db.documents.delete_one({"document_id": document_id})
    return {"message": "Documento eliminado exitosamente"}

# ============== DOCUMENT FOLDERS ==============

DEFAULT_FOLDERS = ["Safety", "PO", "Estimados", "Facturas", "RFI", "RFC"]

@api_router.post("/document-folders", response_model=DocumentFolder)
async def create_document_folder(
    folder_data: DocumentFolderCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Create a new folder for project documents"""
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": folder_data.project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Check access
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    
    # Validate parent folder if provided
    if folder_data.parent_folder_id:
        parent_folder = await db.document_folders.find_one({
            "folder_id": folder_data.parent_folder_id,
            "project_id": folder_data.project_id
        })
        if not parent_folder:
            raise HTTPException(status_code=404, detail="Carpeta padre no encontrada")
    
    # Check for duplicate folder name in same level
    existing = await db.document_folders.find_one({
        "project_id": folder_data.project_id,
        "name": folder_data.name,
        "parent_folder_id": folder_data.parent_folder_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una carpeta con ese nombre en esta ubicación")
    
    folder_doc = {
        "folder_id": f"folder_{uuid.uuid4().hex[:12]}",
        "project_id": folder_data.project_id,
        "name": folder_data.name,
        "parent_folder_id": folder_data.parent_folder_id,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.document_folders.insert_one(folder_doc)
    return DocumentFolder(**folder_doc)

@api_router.get("/document-folders", response_model=List[DocumentFolder])
async def get_document_folders(
    project_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get all folders for a project"""
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    
    folders = await db.document_folders.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    return [DocumentFolder(**f) for f in folders]

@api_router.put("/document-folders/{folder_id}")
async def update_document_folder(
    folder_id: str,
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Rename a folder"""
    user = await get_current_user(request, session_token)
    
    folder_doc = await db.document_folders.find_one({"folder_id": folder_id}, {"_id": 0})
    if not folder_doc:
        raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    
    project_doc = await db.projects.find_one({"project_id": folder_doc['project_id']}, {"_id": 0})
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para modificar esta carpeta")
    
    new_name = data.get("name")
    if new_name:
        # Check for duplicate name
        existing = await db.document_folders.find_one({
            "project_id": folder_doc['project_id'],
            "name": new_name,
            "parent_folder_id": folder_doc['parent_folder_id'],
            "folder_id": {"$ne": folder_id}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe una carpeta con ese nombre en esta ubicación")
        
        await db.document_folders.update_one(
            {"folder_id": folder_id},
            {"$set": {"name": new_name}}
        )
    
    return {"message": "Carpeta actualizada"}

@api_router.delete("/document-folders/{folder_id}")
async def delete_document_folder(
    folder_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Delete a folder and optionally its contents"""
    user = await get_current_user(request, session_token)
    
    folder_doc = await db.document_folders.find_one({"folder_id": folder_id}, {"_id": 0})
    if not folder_doc:
        raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    
    project_doc = await db.projects.find_one({"project_id": folder_doc['project_id']}, {"_id": 0})
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar esta carpeta")
    
    # Get all subfolders recursively
    async def get_all_subfolder_ids(parent_id):
        ids = [parent_id]
        subfolders = await db.document_folders.find({"parent_folder_id": parent_id}, {"folder_id": 1}).to_list(1000)
        for sf in subfolders:
            ids.extend(await get_all_subfolder_ids(sf['folder_id']))
        return ids
    
    folder_ids_to_delete = await get_all_subfolder_ids(folder_id)
    
    # Move documents in these folders to root (folder_id = None)
    await db.documents.update_many(
        {"folder_id": {"$in": folder_ids_to_delete}},
        {"$set": {"folder_id": None}}
    )
    
    # Delete all folders
    await db.document_folders.delete_many({"folder_id": {"$in": folder_ids_to_delete}})
    
    return {"message": "Carpeta eliminada. Los documentos fueron movidos a la raíz."}

@api_router.post("/document-folders/initialize-defaults/{project_id}")
async def initialize_default_folders(
    project_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Initialize default folders for a project"""
    user = await get_current_user(request, session_token)
    
    project_doc = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id:
        raise HTTPException(status_code=403, detail="No tienes permisos")
    
    created_folders = []
    for folder_name in DEFAULT_FOLDERS:
        existing = await db.document_folders.find_one({
            "project_id": project_id,
            "name": folder_name,
            "parent_folder_id": None
        })
        if not existing:
            folder_doc = {
                "folder_id": f"folder_{uuid.uuid4().hex[:12]}",
                "project_id": project_id,
                "name": folder_name,
                "parent_folder_id": None,
                "created_by": user.user_id,
                "created_by_name": user.name,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.document_folders.insert_one(folder_doc)
            created_folders.append(folder_name)
    
    return {"message": f"Carpetas creadas: {', '.join(created_folders)}" if created_folders else "Las carpetas predeterminadas ya existen"}

@api_router.put("/documents/{document_id}/move")
async def move_document_to_folder(
    document_id: str,
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Move a document to a different folder"""
    user = await get_current_user(request, session_token)
    
    document_doc = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not document_doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    project_doc = await db.projects.find_one({"project_id": document_doc['project_id']}, {"_id": 0})
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value] and project_doc['created_by'] != user.user_id and user.user_id not in project_doc.get('team_members', []):
        raise HTTPException(status_code=403, detail="No tienes acceso")
    
    folder_id = data.get("folder_id")  # Can be None to move to root
    
    if folder_id:
        folder = await db.document_folders.find_one({
            "folder_id": folder_id,
            "project_id": document_doc['project_id']
        })
        if not folder:
            raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    
    await db.documents.update_one(
        {"document_id": document_id},
        {"$set": {"folder_id": folder_id}}
    )
    
    return {"message": "Documento movido exitosamente"}

@api_router.post("/invoices/generate", response_model=Invoice)
async def generate_invoice_from_timesheet(
    invoice_data: InvoiceCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Get project
    project = await db.projects.find_one({"project_id": invoice_data.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Get timesheet entries for this project
    timesheet_entries = await db.timesheet.find(
        {"project_id": invoice_data.project_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get ALL employee profiles to lookup hourly rates
    employee_profiles = await db.employee_profiles.find({}, {"_id": 0, "user_id": 1, "hourly_rate": 1}).to_list(1000)
    employee_rates = {ep.get('user_id'): ep.get('hourly_rate', 0) for ep in employee_profiles}
    
    # Also get labor entries as fallback for project-specific rates
    labor_entries = await db.labor.find(
        {"project_id": invoice_data.project_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Create invoice items from timesheet
    items = []
    total_hours = 0
    
    # Group timesheet by user (using user_id as key for accurate rate lookup)
    user_hours = {}
    user_names = {}
    user_ids_list = []
    
    for entry in timesheet_entries:
        user_id = entry.get('user_id', '')
        user_name = entry.get('user_name', 'Unknown')
        hours = entry.get('hours_worked', 0)
        
        if user_id not in user_hours:
            user_hours[user_id] = 0
            user_names[user_id] = user_name
            user_ids_list.append(user_id)
        user_hours[user_id] += hours
        total_hours += hours
    
    # Default rate if no employee profile data
    default_rate = 50.0
    
    # Create items - now using employee profile rates
    for user_id in user_ids_list:
        hours = user_hours[user_id]
        user_name = user_names[user_id]
        
        # PRIORITY 1: Get rate from employee profile (RH data)
        rate = employee_rates.get(user_id, 0)
        
        # PRIORITY 2: If no rate in profile, try project-specific labor data
        if rate <= 0:
            for labor in labor_entries:
                if labor.get('user_id') == user_id or labor.get('labor_category') == user_name:
                    rate = labor.get('hourly_rate', 0)
                    if rate > 0:
                        break
        
        # PRIORITY 3: Use default rate as last resort
        if rate <= 0:
            rate = default_rate
            logger.warning(f"Using default rate for user {user_name} ({user_id}) - no rate found in employee profile or labor data")
        
        items.append(InvoiceItem(
            description=f"Horas trabajadas - {user_name}",
            hours=round(hours, 2),
            rate=rate,
            amount=round(hours * rate, 2)
        ))
    
    # Calculate totals
    subtotal = sum(item.amount for item in items)
    tax_amount = round(subtotal * (invoice_data.tax_rate / 100), 2)
    total = subtotal + tax_amount
    
    # Generate invoice number (use custom if provided)
    if invoice_data.custom_number and invoice_data.custom_number.strip():
        invoice_number = invoice_data.custom_number.strip()
    else:
        # Get next number from company settings or count
        company_settings = await db.company_settings.find_one({}, {"_id": 0})
        next_num = company_settings.get("next_invoice_number", 1) if company_settings else 1
        invoice_number = f"INV-{datetime.now().year}-{str(next_num).zfill(4)}"
        # Increment the counter
        await db.company_settings.update_one({}, {"$inc": {"next_invoice_number": 1}}, upsert=True)
    
    # Create invoice
    invoice_id = f"inv_{uuid4().hex[:16]}"
    due_date = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    
    # Get sponsor from project
    sponsor_name = invoice_data.sponsor_name or project.get('sponsor', '')
    
    invoice_doc = {
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "project_id": invoice_data.project_id,
        "project_name": project.get('name', ''),
        "client_name": invoice_data.client_name,
        "client_email": invoice_data.client_email,
        "client_phone": invoice_data.client_phone,
        "client_address": invoice_data.client_address,
        "sponsor_name": sponsor_name,
        "items": [item.model_dump() for item in items],
        "subtotal": subtotal,
        "tax_rate": invoice_data.tax_rate,
        "tax_amount": tax_amount,
        "tax_type_name": invoice_data.tax_type_name,
        "total": total,
        "amount_paid": 0.0,
        "balance_due": total,
        "status": "draft",
        "notes": invoice_data.notes,
        "created_by": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "due_date": due_date,
        "sent_date": None,
        "paid_date": None
    }
    
    await db.invoices.insert_one(invoice_doc)
    
    # Save client for future autocomplete
    if invoice_data.client_name:
        await db.saved_clients.update_one(
            {"name": invoice_data.client_name},
            {"$set": {
                "name": invoice_data.client_name,
                "email": invoice_data.client_email or "",
                "phone": invoice_data.client_phone or "",
                "address": invoice_data.client_address or "",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }, "$setOnInsert": {
                "id": f"saved_client_{uuid4().hex[:12]}",
                "created_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "create",
        "invoice",
        invoice_id,
        invoice_number,
        {"project": project.get('name'), "total": total}
    )
    
    return Invoice(**invoice_doc)

# Modelo para crear factura manual con items personalizados
class ManualInvoiceItem(BaseModel):
    description: str
    quantity: float = 1.0
    unit_price: float
    amount: float

class ManualInvoiceCreate(BaseModel):
    project_id: Optional[str] = None
    client_name: str
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    sponsor_name: Optional[str] = None
    items: List[ManualInvoiceItem]
    tax_rate: float = 0.0
    selected_taxes: Optional[List[dict]] = []
    discount_percent: float = 0.0
    notes: Optional[str] = None
    terms: Optional[str] = None
    custom_number: Optional[str] = None
    price_breakdown: Optional[dict] = None  # {material_equipment: float, labor: float, total: float}

@api_router.post("/invoices/manual")
async def create_manual_invoice(
    invoice_data: ManualInvoiceCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Crear factura manual con items personalizados (formato Task)"""
    user = await get_current_user(request, session_token)
    
    # Get project if provided
    project_name = ""
    if invoice_data.project_id:
        project = await db.projects.find_one({"project_id": invoice_data.project_id}, {"_id": 0})
        if project:
            project_name = project.get('name', '')
    
    # Calculate totals
    subtotal = sum(item.amount for item in invoice_data.items)
    discount_amount = subtotal * (invoice_data.discount_percent / 100)
    taxable = subtotal - discount_amount
    tax_amount = round(taxable * (invoice_data.tax_rate / 100), 2)
    total = taxable + tax_amount
    
    # Generate invoice number
    if invoice_data.custom_number and invoice_data.custom_number.strip():
        invoice_number = invoice_data.custom_number.strip()
    else:
        company_settings = await db.company_settings.find_one({}, {"_id": 0})
        next_num = company_settings.get("next_invoice_number", 1) if company_settings else 1
        invoice_number = f"INV-{datetime.now().year}-{str(next_num).zfill(4)}"
        await db.company_settings.update_one({}, {"$inc": {"next_invoice_number": 1}}, upsert=True)
    
    invoice_id = f"inv_{uuid4().hex[:16]}"
    due_date = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    
    # Convert items to invoice format (hours=quantity, rate=unit_price)
    invoice_items = []
    for item in invoice_data.items:
        invoice_items.append({
            "description": item.description,
            "hours": item.quantity,
            "rate": item.unit_price,
            "amount": item.amount
        })
    
    # Get sponsor from project if available
    sponsor_name = invoice_data.sponsor_name or ""
    if invoice_data.project_id and not sponsor_name:
        project = await db.projects.find_one({"project_id": invoice_data.project_id}, {"_id": 0})
        if project:
            project_name = project.get('name', '')
            sponsor_name = project.get('sponsor', '')

    invoice_doc = {
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "project_id": invoice_data.project_id or "",
        "project_name": project_name,
        "client_name": invoice_data.client_name,
        "client_email": invoice_data.client_email,
        "client_phone": invoice_data.client_phone,
        "client_address": invoice_data.client_address,
        "sponsor_name": sponsor_name,
        "items": invoice_items,
        "subtotal": subtotal,
        "discount_percent": invoice_data.discount_percent,
        "discount_amount": discount_amount,
        "tax_rate": invoice_data.tax_rate,
        "tax_amount": tax_amount,
        "selected_taxes": invoice_data.selected_taxes or [],
        "total": total,
        "amount_paid": 0.0,
        "balance_due": total,
        "status": "draft",
        "notes": invoice_data.notes,
        "terms": invoice_data.terms,
        "price_breakdown": invoice_data.price_breakdown,
        "created_by": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "due_date": due_date,
        "sent_date": None,
        "paid_date": None
    }
    
    await db.invoices.insert_one(invoice_doc)
    
    # Save client for future autocomplete
    if invoice_data.client_name:
        await db.saved_clients.update_one(
            {"name": invoice_data.client_name},
            {"$set": {
                "name": invoice_data.client_name,
                "email": invoice_data.client_email or "",
                "phone": invoice_data.client_phone or "",
                "address": invoice_data.client_address or "",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }, "$setOnInsert": {
                "id": f"saved_client_{uuid4().hex[:12]}",
                "created_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    # Remove _id if present (added by MongoDB)
    invoice_doc.pop('_id', None)
    
    await log_audit(user.user_id, user.name, "create", "invoice", invoice_id, invoice_number, {"client": invoice_data.client_name, "total": total})
    
    return invoice_doc

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(
    project_id: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [Invoice(**inv) for inv in invoices]

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(
    invoice_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    return Invoice(**invoice)

@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(
    invoice_id: str,
    invoice_data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    # Get project name if project_id provided
    project_name = None
    if invoice_data.get('project_id'):
        project = await db.projects.find_one({"project_id": invoice_data['project_id']}, {"_id": 0, "name": 1})
        if project:
            project_name = project.get('name')
    
    update_data = {
        "project_id": invoice_data.get('project_id', invoice.get('project_id')),
        "project_name": project_name or invoice.get('project_name'),
        "client_name": invoice_data.get('client_name', invoice.get('client_name')),
        "client_email": invoice_data.get('client_email', invoice.get('client_email')),
        "client_phone": invoice_data.get('client_phone', invoice.get('client_phone')),
        "client_address": invoice_data.get('client_address', invoice.get('client_address')),
        "sponsor_name": invoice_data.get('sponsor_name', invoice.get('sponsor_name')),
        "items": invoice_data.get('items', invoice.get('items')),
        "subtotal": invoice_data.get('subtotal', invoice.get('subtotal')),
        "discount_percent": invoice_data.get('discount_percent', invoice.get('discount_percent', 0)),
        "discount_amount": invoice_data.get('discount_amount', invoice.get('discount_amount', 0)),
        "tax_rate": invoice_data.get('tax_rate', invoice.get('tax_rate')),
        "tax_amount": invoice_data.get('tax_amount', invoice.get('tax_amount')),
        "total": invoice_data.get('total', invoice.get('total')),
        "notes": invoice_data.get('notes', invoice.get('notes')),
        "terms": invoice_data.get('terms', invoice.get('terms')),
        "selected_taxes": invoice_data.get('selected_taxes', invoice.get('selected_taxes', [])),
        "price_breakdown": invoice_data.get('price_breakdown', invoice.get('price_breakdown'))
    }
    
    # Update invoice_number if custom_number provided
    if invoice_data.get('custom_number') and invoice_data['custom_number'].strip():
        update_data["invoice_number"] = invoice_data['custom_number'].strip()
    
    # Recalculate balance_due
    update_data["balance_due"] = update_data["total"] - invoice.get('amount_paid', 0)
    
    await db.invoices.update_one({"invoice_id": invoice_id}, {"$set": update_data})
    
    await log_audit(user.user_id, user.name, "update", "invoice", invoice_id, invoice.get('invoice_number'), {"total": update_data["total"]})
    
    updated = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    return Invoice(**updated)

@api_router.put("/invoices/{invoice_id}/status", response_model=Invoice)
async def update_invoice_status(
    invoice_id: str,
    status: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if status not in ["draft", "sent", "paid"]:
        raise HTTPException(status_code=400, detail="Estado inválido")
    
    await db.invoices.update_one(
        {"invoice_id": invoice_id},
        {"$set": {"status": status}}
    )
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "update",
        "invoice",
        invoice_id,
        invoice.get('invoice_number', ''),
        {"status": status}
    )
    
    return Invoice(**invoice)

@api_router.put("/invoices/{invoice_id}/mark-sent")
async def mark_invoice_sent(
    invoice_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Mark invoice as sent and set sent_date"""
    user = await get_current_user(request, session_token)
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    await db.invoices.update_one(
        {"invoice_id": invoice_id},
        {"$set": {
            "status": "sent",
            "sent_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "invoice", invoice_id, updated.get('invoice_number', ''), {"action": "marked_sent"})
    
    return updated

# ============= SAVED CLIENTS (for autocomplete) =============
@api_router.get("/saved-clients")
async def get_saved_clients(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all saved clients for autocomplete"""
    user = await get_current_user(request, session_token)
    clients = await db.saved_clients.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return clients

@api_router.post("/saved-clients")
async def create_saved_client(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Save a new client for future use"""
    user = await get_current_user(request, session_token)
    
    # Check if client already exists by name
    existing = await db.saved_clients.find_one({"name": data.get("name")})
    if existing:
        # Update existing client
        await db.saved_clients.update_one(
            {"name": data.get("name")},
            {"$set": {
                "email": data.get("email", ""),
                "phone": data.get("phone", ""),
                "address": data.get("address", ""),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated = await db.saved_clients.find_one({"name": data.get("name")}, {"_id": 0})
        return updated
    
    client_id = f"saved_client_{uuid4().hex[:12]}"
    client_doc = {
        "id": client_id,
        "name": data.get("name", ""),
        "email": data.get("email", ""),
        "phone": data.get("phone", ""),
        "address": data.get("address", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.saved_clients.insert_one(client_doc)
    client_doc.pop('_id', None)
    return client_doc

@api_router.delete("/saved-clients/{client_id}")
async def delete_saved_client(client_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Delete a saved client"""
    user = await get_current_user(request, session_token)
    await db.saved_clients.delete_one({"id": client_id})
    return {"message": "Cliente eliminado"}

# ============= TAX TYPES CONFIGURATION =============
@api_router.get("/tax-types")
async def get_tax_types(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all configured tax types"""
    user = await get_current_user(request, session_token)
    tax_types = await db.tax_types.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return tax_types

@api_router.post("/tax-types")
async def create_tax_type(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Create a new tax type"""
    user = await get_current_user(request, session_token)
    
    tax_id = f"tax_{uuid4().hex[:8]}"
    tax_doc = {
        "id": tax_id,
        "name": data.get("name", ""),
        "percentage": float(data.get("percentage", 0)),
        "description": data.get("description", ""),
        "is_active": data.get("is_active", True),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.tax_types.insert_one(tax_doc)
    tax_doc.pop('_id', None)
    await log_audit(user.user_id, user.name, "create", "tax_type", tax_id, data.get("name"), {})
    return tax_doc

@api_router.put("/tax-types/{tax_id}")
async def update_tax_type(tax_id: str, data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Update a tax type"""
    user = await get_current_user(request, session_token)
    
    await db.tax_types.update_one(
        {"id": tax_id},
        {"$set": {
            "name": data.get("name"),
            "percentage": float(data.get("percentage", 0)),
            "description": data.get("description", ""),
            "is_active": data.get("is_active", True)
        }}
    )
    
    updated = await db.tax_types.find_one({"id": tax_id}, {"_id": 0})
    return updated

@api_router.delete("/tax-types/{tax_id}")
async def delete_tax_type(tax_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Delete a tax type"""
    user = await get_current_user(request, session_token)
    await db.tax_types.delete_one({"id": tax_id})
    await log_audit(user.user_id, user.name, "delete", "tax_type", tax_id, "", {})
    return {"message": "Tipo de impuesto eliminado"}

@api_router.get("/projects/{project_id}/financial-summary")
async def get_project_financial_summary(
    project_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get financial summary for a project including invoices and payments"""
    user = await get_current_user(request, session_token)
    
    # Get all invoices for the project
    invoices = await db.invoices.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Calculate totals
    total_invoiced = sum(inv.get('total', 0) for inv in invoices)
    total_paid = sum(inv.get('amount_paid', 0) for inv in invoices)
    total_pending = total_invoiced - total_paid
    
    # Count by status
    status_counts = {
        'paid': 0,
        'partial': 0,
        'pending': 0,
        'sent': 0,
        'draft': 0,
        'overdue': 0
    }
    
    for inv in invoices:
        status = inv.get('status', 'pending')
        if status in status_counts:
            status_counts[status] += 1
    
    return {
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "invoice_count": len(invoices),
        "status_counts": status_counts
    }

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    await db.invoices.delete_one({"invoice_id": invoice_id})
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "delete",
        "invoice",
        invoice_id,
        invoice.get('invoice_number', ''),
        {}
    )
    
    return {"message": "Factura eliminada exitosamente"}

@api_router.post("/invoices/{invoice_id}/payments", response_model=Payment)
async def add_payment_to_invoice(
    invoice_id: str,
    payment_data: PaymentCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Get invoice
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    # Calculate new amounts
    current_paid = invoice.get('amount_paid', 0)
    new_paid = current_paid + payment_data.amount
    balance = invoice['total'] - new_paid
    
    # Determine new status
    if balance <= 0:
        new_status = 'paid'
        paid_date = datetime.now(timezone.utc).isoformat()
    elif new_paid > 0:
        new_status = 'partial'
        paid_date = None
    else:
        new_status = invoice['status']
        paid_date = None
    
    # Create payment record
    payment_id = f"pay_{uuid4().hex[:16]}"
    payment_doc = {
        "payment_id": payment_id,
        "invoice_id": invoice_id,
        "amount": payment_data.amount,
        "payment_method": payment_data.payment_method,
        "reference": payment_data.reference,
        "notes": payment_data.notes,
        "created_by": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.payments.insert_one(payment_doc)
    
    # Update invoice
    update_data = {
        "amount_paid": new_paid,
        "balance_due": balance,
        "status": new_status
    }
    if paid_date:
        update_data["paid_date"] = paid_date
    
    await db.invoices.update_one(
        {"invoice_id": invoice_id},
        {"$set": update_data}
    )
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "create",
        "payment",
        payment_id,
        f"Pago de ${payment_data.amount} para {invoice.get('invoice_number')}",
        {"invoice": invoice.get('invoice_number'), "amount": payment_data.amount}
    )
    
    # Send Slack notification
    event_type = "invoice_paid" if new_status == 'paid' else "payment_received"
    await notify_slack_event(event_type, {
        "invoice": invoice.get('invoice_number'),
        "amount": payment_data.amount
    })
    
    return Payment(**payment_doc)

@api_router.get("/invoices/{invoice_id}/payments", response_model=List[Payment])
async def get_invoice_payments(
    invoice_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    payments = await db.payments.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(1000)
    return [Payment(**p) for p in payments]

@api_router.post("/invoices/{invoice_id}/send")
async def send_invoice_email(
    invoice_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    invoice = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    
    if not invoice.get('client_email'):
        raise HTTPException(status_code=400, detail="La factura no tiene email del cliente")
    
    # Update status and sent date
    await db.invoices.update_one(
        {"invoice_id": invoice_id},
        {"$set": {
            "status": "sent",
            "sent_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # TODO: Implement actual email sending using SMTP settings
    # For now, just mark as sent
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "update",
        "invoice",
        invoice_id,
        invoice.get('invoice_number', ''),
        {"action": "sent_email", "to": invoice.get('client_email')}
    )
    
    return {"message": f"Factura enviada a {invoice.get('client_email')}", "status": "sent"}

# ============== ESTIMATES ENDPOINTS ==============

@api_router.post("/estimates", response_model=Estimate)
async def create_estimate(
    estimate_data: EstimateCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Generate estimate number (use custom if provided)
    if estimate_data.custom_number and estimate_data.custom_number.strip():
        estimate_number = estimate_data.custom_number.strip()
    else:
        company_settings = await db.company_settings.find_one({}, {"_id": 0})
        next_num = company_settings.get("next_estimate_number", 1) if company_settings else 1
        estimate_number = f"EST-{datetime.now().year}-{str(next_num).zfill(4)}"
        await db.company_settings.update_one({}, {"$inc": {"next_estimate_number": 1}}, upsert=True)
    
    estimate_id = f"est_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    # Get project name if project_id provided
    project_name = None
    if estimate_data.project_id:
        project = await db.projects.find_one({"project_id": estimate_data.project_id}, {"_id": 0, "name": 1})
        if project:
            project_name = project.get('name')
    
    # Calculate totals
    subtotal = sum(item.amount for item in estimate_data.items)
    discount_amount = subtotal * (estimate_data.discount_percent / 100)
    taxable_amount = subtotal - discount_amount
    tax_amount = taxable_amount * (estimate_data.tax_rate / 100)
    total = taxable_amount + tax_amount
    
    estimate_doc = {
        "estimate_id": estimate_id,
        "estimate_number": estimate_number,
        "project_id": estimate_data.project_id,
        "project_name": project_name,
        "client_profile_id": estimate_data.client_profile_id,
        "client_company": estimate_data.client_company,
        "client_name": estimate_data.client_name,
        "client_email": estimate_data.client_email,
        "client_phone": estimate_data.client_phone,
        "client_address": estimate_data.client_address,
        "title": estimate_data.title,
        "description": estimate_data.description,
        "items": [item.model_dump() for item in estimate_data.items],
        "subtotal": round(subtotal, 2),
        "discount_percent": estimate_data.discount_percent,
        "discount_amount": round(discount_amount, 2),
        "tax_rate": estimate_data.tax_rate,
        "tax_amount": round(tax_amount, 2),
        "selected_taxes": estimate_data.selected_taxes or [],
        "total": round(total, 2),
        "status": "draft",
        "notes": estimate_data.notes,
        "terms": estimate_data.terms,
        "valid_until": estimate_data.valid_until,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now,
        "sent_date": None,
        "approved_date": None,
        "converted_invoice_id": None
    }
    
    await db.estimates.insert_one(estimate_doc)
    
    await log_audit(user.user_id, user.name, "create", "estimate", estimate_id, estimate_number, {})
    
    return Estimate(**estimate_doc)

@api_router.get("/estimates", response_model=List[Estimate])
async def get_estimates(
    status: Optional[str] = None,
    project_id: Optional[str] = None,
    unassigned: Optional[bool] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    query = {}
    if status:
        query["status"] = status
    if project_id:
        query["project_id"] = project_id
    if unassigned:
        query["client_profile_id"] = {"$in": [None, ""]}
    
    estimates = await db.estimates.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [Estimate(**e) for e in estimates]

@api_router.put("/estimates/{estimate_id}/assign")
async def assign_estimate_to_client(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    body = await request.json()
    client_profile_id = body.get("client_profile_id")
    
    if not client_profile_id:
        raise HTTPException(status_code=400, detail="client_profile_id es requerido")
    
    # Verify client exists
    client = await db.client_profiles.find_one({"profile_id": client_profile_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Update estimate
    result = await db.estimates.update_one(
        {"estimate_id": estimate_id},
        {"$set": {
            "client_profile_id": client_profile_id,
            "client_company": client.get("company_name"),
            "client_name": client.get("contact_name"),
            "client_email": client.get("email"),
            "client_phone": client.get("phone"),
            "client_address": client.get("address")
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    await log_audit(user.user_id, user.name, "update", "estimate", estimate_id, None, {"action": "assigned_to_client", "client_profile_id": client_profile_id})
    
    return {"message": "Estimado asignado exitosamente"}

@api_router.get("/estimates/{estimate_id}", response_model=Estimate)
async def get_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    return Estimate(**estimate)

@api_router.put("/estimates/{estimate_id}", response_model=Estimate)
async def update_estimate(
    estimate_id: str,
    estimate_data: EstimateCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    if estimate.get('status') == 'converted':
        raise HTTPException(status_code=400, detail="No se puede editar un estimado convertido a factura")
    
    # Get project name if project_id provided
    project_name = None
    if estimate_data.project_id:
        project = await db.projects.find_one({"project_id": estimate_data.project_id}, {"_id": 0, "name": 1})
        if project:
            project_name = project.get('name')
    
    # Calculate totals - use price_breakdown.total if available, otherwise calculate from items
    if estimate_data.price_breakdown and estimate_data.price_breakdown.get('total'):
        subtotal = estimate_data.price_breakdown.get('total', 0)
    else:
        subtotal = sum(item.amount for item in estimate_data.items)
    discount_amount = subtotal * (estimate_data.discount_percent / 100)
    taxable_amount = subtotal - discount_amount
    tax_amount = taxable_amount * (estimate_data.tax_rate / 100)
    total = taxable_amount + tax_amount
    
    update_data = {
        "project_id": estimate_data.project_id,
        "project_name": project_name,
        "client_company": estimate_data.client_company,
        "client_name": estimate_data.client_name,
        "client_email": estimate_data.client_email,
        "client_phone": estimate_data.client_phone,
        "client_address": estimate_data.client_address,
        "title": estimate_data.title,
        "description": estimate_data.description,
        "items": [item.model_dump() for item in estimate_data.items],
        "subtotal": round(subtotal, 2),
        "discount_percent": estimate_data.discount_percent,
        "discount_amount": round(discount_amount, 2),
        "tax_rate": estimate_data.tax_rate,
        "tax_amount": round(tax_amount, 2),
        "selected_taxes": estimate_data.selected_taxes or [],
        "total": round(total, 2),
        "notes": estimate_data.notes,
        "terms": estimate_data.terms,
        "valid_until": estimate_data.valid_until,
        "price_breakdown": estimate_data.price_breakdown
    }
    
    # Update estimate_number if custom_number is provided
    if estimate_data.custom_number and estimate_data.custom_number.strip():
        update_data["estimate_number"] = estimate_data.custom_number.strip()
    
    await db.estimates.update_one({"estimate_id": estimate_id}, {"$set": update_data})
    
    updated = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    await log_audit(user.user_id, user.name, "update", "estimate", estimate_id, estimate.get('estimate_number'), {})
    
    return Estimate(**updated)

@api_router.put("/estimates/{estimate_id}/status")
async def update_estimate_status(
    estimate_id: str,
    status: str = Query(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    valid_statuses = ['draft', 'sent', 'approved', 'rejected', 'converted']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Debe ser uno de: {valid_statuses}")
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    update_data = {"status": status}
    now = datetime.now(timezone.utc).isoformat()
    
    if status == 'sent':
        update_data["sent_date"] = now
    elif status == 'approved':
        update_data["approved_date"] = now
    
    await db.estimates.update_one({"estimate_id": estimate_id}, {"$set": update_data})
    
    await log_audit(user.user_id, user.name, "update", "estimate", estimate_id, estimate.get('estimate_number'), {"status": status})
    
    return {"message": f"Estado actualizado a {status}"}

@api_router.post("/estimates/{estimate_id}/convert")
async def convert_estimate_to_invoice(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    if estimate.get('status') == 'converted':
        raise HTTPException(status_code=400, detail="Este estimado ya fue convertido a factura")
    
    # Generate invoice
    count = await db.invoices.count_documents({})
    invoice_number = f"INV-{datetime.now().year}-{str(count + 1).zfill(4)}"
    invoice_id = f"inv_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    # Convert estimate items to invoice items
    invoice_items = []
    for item in estimate.get('items', []):
        invoice_items.append({
            "description": item.get('description'),
            "hours": item.get('quantity', 1),
            "rate": item.get('unit_price'),
            "amount": item.get('amount')
        })
    
    invoice_doc = {
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "project_id": estimate.get('project_id') or '',
        "project_name": estimate.get('project_name') or estimate.get('title'),
        "client_name": estimate.get('client_name'),
        "client_email": estimate.get('client_email'),
        "items": invoice_items,
        "subtotal": estimate.get('subtotal'),
        "tax_rate": estimate.get('tax_rate'),
        "tax_amount": estimate.get('tax_amount'),
        "total": estimate.get('total'),
        "amount_paid": 0,
        "balance_due": estimate.get('total'),
        "status": "draft",
        "notes": estimate.get('notes'),
        "created_by": user.user_id,
        "created_at": now,
        "due_date": None,
        "sent_date": None,
        "paid_date": None
    }
    
    await db.invoices.insert_one(invoice_doc)
    
    # Update estimate status
    await db.estimates.update_one(
        {"estimate_id": estimate_id},
        {"$set": {"status": "converted", "converted_invoice_id": invoice_id}}
    )
    
    await log_audit(user.user_id, user.name, "convert", "estimate", estimate_id, estimate.get('estimate_number'), {"invoice_id": invoice_id})
    
    return {"message": "Estimado convertido a factura exitosamente", "invoice_id": invoice_id, "invoice_number": invoice_number}

@api_router.post("/estimates/{estimate_id}/duplicate", response_model=Estimate)
async def duplicate_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    # Generate new estimate
    count = await db.estimates.count_documents({})
    new_estimate_number = f"EST-{datetime.now().year}-{str(count + 1).zfill(4)}"
    new_estimate_id = f"est_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    new_estimate = {
        **estimate,
        "estimate_id": new_estimate_id,
        "estimate_number": new_estimate_number,
        "status": "draft",
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now,
        "sent_date": None,
        "approved_date": None,
        "converted_invoice_id": None
    }
    
    await db.estimates.insert_one(new_estimate)
    
    await log_audit(user.user_id, user.name, "duplicate", "estimate", new_estimate_id, new_estimate_number, {"original": estimate_id})
    
    return Estimate(**new_estimate)

@api_router.delete("/estimates/{estimate_id}")
async def delete_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    await db.estimates.delete_one({"estimate_id": estimate_id})
    
    await log_audit(user.user_id, user.name, "delete", "estimate", estimate_id, estimate.get('estimate_number'), {})
    
    return {"message": "Estimado eliminado exitosamente"}

@api_router.post("/estimates/{estimate_id}/send")
async def send_estimate_email(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimado no encontrado")
    
    if not estimate.get('client_email'):
        raise HTTPException(status_code=400, detail="El estimado no tiene email del cliente")
    
    # Update status and sent date
    await db.estimates.update_one(
        {"estimate_id": estimate_id},
        {"$set": {
            "status": "sent",
            "sent_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_audit(user.user_id, user.name, "update", "estimate", estimate_id, estimate.get('estimate_number'), {"action": "sent_email", "to": estimate.get('client_email')})
    
    return {"message": f"Estimado enviado a {estimate.get('client_email')}", "status": "sent"}

# ============== PURCHASE ORDERS ENDPOINTS ==============

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(
    po_data: PurchaseOrderCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Generate PO number (use custom if provided)
    if po_data.custom_number and po_data.custom_number.strip():
        po_number = po_data.custom_number.strip()
    else:
        company_settings = await db.company_settings.find_one({}, {"_id": 0})
        next_num = company_settings.get("next_po_number", 1) if company_settings else 1
        po_number = f"PO-{datetime.now(timezone.utc).year}-{str(next_num).zfill(4)}"
        await db.company_settings.update_one({}, {"$inc": {"next_po_number": 1}}, upsert=True)
    
    # Get project name if project_id provided
    project_name = None
    if po_data.project_id:
        project = await db.projects.find_one({"project_id": po_data.project_id}, {"_id": 0, "name": 1})
        if project:
            project_name = project.get('name')
    
    # Calculate totals
    subtotal = sum(item.amount for item in po_data.items)
    discount_amount = subtotal * (po_data.discount_percent / 100)
    taxable = subtotal - discount_amount
    tax_amount = taxable * (po_data.tax_rate / 100)
    total = taxable + tax_amount
    
    po_id = f"po_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    po_doc = {
        "po_id": po_id,
        "po_number": po_number,
        "project_id": po_data.project_id,
        "project_name": project_name,
        "supplier_name": po_data.supplier_name,
        "supplier_email": po_data.supplier_email,
        "supplier_phone": po_data.supplier_phone,
        "supplier_address": po_data.supplier_address,
        "title": po_data.title,
        "description": po_data.description,
        "items": [item.model_dump() for item in po_data.items],
        "subtotal": subtotal,
        "discount_percent": po_data.discount_percent,
        "discount_amount": discount_amount,
        "tax_rate": po_data.tax_rate,
        "tax_amount": tax_amount,
        "total": total,
        "status": "draft",
        "notes": po_data.notes,
        "terms": po_data.terms,
        "expected_delivery_date": po_data.expected_delivery_date,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now,
        "approved_date": None,
        "sent_date": None,
        "received_date": None,
        "linked_expense_id": None
    }
    
    await db.purchase_orders.insert_one(po_doc)
    await log_audit(user.user_id, user.name, "create", "purchase_order", po_id, po_number, {"total": total})
    
    return PurchaseOrder(**po_doc)

@api_router.get("/purchase-orders", response_model=List[PurchaseOrder])
async def get_purchase_orders(
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    if status:
        query["status"] = status
    
    pos = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [PurchaseOrder(**po) for po in pos]

@api_router.get("/purchase-orders/{po_id}", response_model=PurchaseOrder)
async def get_purchase_order(
    po_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    return PurchaseOrder(**po)

@api_router.put("/purchase-orders/{po_id}", response_model=PurchaseOrder)
async def update_purchase_order(
    po_id: str,
    po_data: PurchaseOrderCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    if po.get('status') in ['completed', 'cancelled']:
        raise HTTPException(status_code=400, detail="No se puede editar una orden completada o cancelada")
    
    # Get project name if project_id provided
    project_name = None
    if po_data.project_id:
        project = await db.projects.find_one({"project_id": po_data.project_id}, {"_id": 0, "name": 1})
        if project:
            project_name = project.get('name')
    
    # Calculate totals
    subtotal = sum(item.amount for item in po_data.items)
    discount_amount = subtotal * (po_data.discount_percent / 100)
    taxable = subtotal - discount_amount
    tax_amount = taxable * (po_data.tax_rate / 100)
    total = taxable + tax_amount
    
    update_data = {
        "project_id": po_data.project_id,
        "project_name": project_name,
        "supplier_name": po_data.supplier_name,
        "supplier_email": po_data.supplier_email,
        "supplier_phone": po_data.supplier_phone,
        "supplier_address": po_data.supplier_address,
        "title": po_data.title,
        "description": po_data.description,
        "items": [item.model_dump() for item in po_data.items],
        "subtotal": subtotal,
        "discount_percent": po_data.discount_percent,
        "discount_amount": discount_amount,
        "tax_rate": po_data.tax_rate,
        "tax_amount": tax_amount,
        "total": total,
        "notes": po_data.notes,
        "terms": po_data.terms,
        "expected_delivery_date": po_data.expected_delivery_date
    }
    
    # Update po_number if custom_number provided
    if po_data.custom_number and po_data.custom_number.strip():
        update_data["po_number"] = po_data.custom_number.strip()
    
    await db.purchase_orders.update_one({"po_id": po_id}, {"$set": update_data})
    await log_audit(user.user_id, user.name, "update", "purchase_order", po_id, po.get('po_number'), {"total": total})
    
    updated = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    return PurchaseOrder(**updated)

@api_router.put("/purchase-orders/{po_id}/status")
async def update_purchase_order_status(
    po_id: str,
    status: str = Query(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    valid_statuses = ['draft', 'approved', 'sent', 'partially_received', 'completed', 'cancelled']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Estados válidos: {', '.join(valid_statuses)}")
    
    update_data = {"status": status}
    now = datetime.now(timezone.utc).isoformat()
    
    if status == 'approved':
        update_data["approved_date"] = now
    elif status == 'sent':
        update_data["sent_date"] = now
    elif status in ['completed', 'partially_received']:
        update_data["received_date"] = now
        
        # Create expense if linked to project and status is completed
        if status == 'completed' and po.get('project_id') and not po.get('linked_expense_id'):
            # Find or create a budget category for purchase orders
            category = await db.budget_categories.find_one(
                {"project_id": po.get('project_id'), "name": "Órdenes de Compra"},
                {"_id": 0}
            )
            if not category:
                category_id = f"cat_{uuid4().hex[:12]}"
                category = {
                    "category_id": category_id,
                    "project_id": po.get('project_id'),
                    "name": "Órdenes de Compra",
                    "allocated_amount": 0,
                    "spent_amount": 0,
                    "created_at": now
                }
                await db.budget_categories.insert_one(category)
            else:
                category_id = category.get('category_id')
            
            # Create expense
            expense_id = f"exp_{uuid4().hex[:12]}"
            expense_doc = {
                "expense_id": expense_id,
                "project_id": po.get('project_id'),
                "category_id": category_id,
                "description": f"OC: {po.get('po_number')} - {po.get('title')}",
                "amount": po.get('total'),
                "date": now[:10],
                "created_by": user.user_id,
                "created_at": now
            }
            await db.expenses.insert_one(expense_doc)
            
            # Update category spent amount
            await db.budget_categories.update_one(
                {"category_id": category_id},
                {"$inc": {"spent_amount": po.get('total')}}
            )
            
            # Update project budget_spent
            await db.projects.update_one(
                {"project_id": po.get('project_id')},
                {"$inc": {"budget_spent": po.get('total')}}
            )
            
            update_data["linked_expense_id"] = expense_id
    
    await db.purchase_orders.update_one({"po_id": po_id}, {"$set": update_data})
    await log_audit(user.user_id, user.name, "update", "purchase_order", po_id, po.get('po_number'), {"status": status})
    
    return {"message": f"Estado actualizado a {status}"}

@api_router.delete("/purchase-orders/{po_id}")
async def delete_purchase_order(
    po_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    await db.purchase_orders.delete_one({"po_id": po_id})
    await log_audit(user.user_id, user.name, "delete", "purchase_order", po_id, po.get('po_number'), {})
    
    return {"message": "Orden de compra eliminada exitosamente"}

@api_router.post("/purchase-orders/{po_id}/duplicate", response_model=PurchaseOrder)
async def duplicate_purchase_order(
    po_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    # Generate PO number for duplicate (always auto-generate)
    company_settings = await db.company_settings.find_one({}, {"_id": 0})
    next_num = company_settings.get("next_po_number", 1) if company_settings else 1
    po_number = f"PO-{datetime.now(timezone.utc).year}-{str(next_num).zfill(4)}"
    await db.company_settings.update_one({}, {"$inc": {"next_po_number": 1}}, upsert=True)
    
    new_po_id = f"po_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    new_po = {
        **po,
        "po_id": new_po_id,
        "po_number": po_number,
        "status": "draft",
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now,
        "approved_date": None,
        "sent_date": None,
        "received_date": None,
        "linked_expense_id": None
    }
    
    await db.purchase_orders.insert_one(new_po)
    await log_audit(user.user_id, user.name, "create", "purchase_order", new_po_id, po_number, {"duplicated_from": po_id})
    
    return PurchaseOrder(**new_po)

@api_router.post("/purchase-orders/{po_id}/send")
async def send_purchase_order_email(
    po_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    
    if not po.get('supplier_email'):
        raise HTTPException(status_code=400, detail="La orden no tiene email de proveedor")
    
    # Get PDF attachment from request body if provided
    body = await request.json() if request.headers.get('content-type') == 'application/json' else {}
    pdf_base64 = body.get('pdf_base64')
    
    # Build email content
    items_html = ""
    for item in po.get('items', []):
        items_html += f"<tr><td>{item['description']}</td><td>{item['quantity']}</td><td>${item['unit_price']:.2f}</td><td>${item['amount']:.2f}</td></tr>"
    
    html_content = f"""
    <h2>Orden de Compra {po.get('po_number')}</h2>
    <p><strong>Título:</strong> {po.get('title')}</p>
    <p><strong>Fecha:</strong> {po.get('created_at')[:10]}</p>
    {f"<p><strong>Entrega esperada:</strong> {po.get('expected_delivery_date')}</p>" if po.get('expected_delivery_date') else ""}
    <table border="1" cellpadding="5">
        <tr><th>Descripción</th><th>Cantidad</th><th>Precio Unit.</th><th>Total</th></tr>
        {items_html}
    </table>
    <p><strong>Subtotal:</strong> ${po.get('subtotal', 0):.2f}</p>
    {f"<p><strong>Descuento ({po.get('discount_percent')}%):</strong> -${po.get('discount_amount', 0):.2f}</p>" if po.get('discount_amount') else ""}
    {f"<p><strong>Impuesto ({po.get('tax_rate')}%):</strong> ${po.get('tax_amount', 0):.2f}</p>" if po.get('tax_amount') else ""}
    <p><strong>Total:</strong> ${po.get('total', 0):.2f}</p>
    {f"<p><strong>Notas:</strong> {po.get('notes')}</p>" if po.get('notes') else ""}
    {f"<p><strong>Términos:</strong> {po.get('terms')}</p>" if po.get('terms') else ""}
    <p><em>Adjunto: PDF de la Orden de Compra</em></p>
    """
    
    text_content = f"Orden de Compra {po.get('po_number')} - Total: ${po.get('total', 0):.2f}"
    
    # Prepare attachments if PDF provided
    attachments = None
    if pdf_base64:
        attachments = [{
            'filename': f"PO_{po.get('po_number')}.pdf",
            'content': pdf_base64,
            'content_type': 'application/pdf'
        }]
    
    await send_email(
        po.get('supplier_email'),
        f"Orden de Compra {po.get('po_number')} - {po.get('title')}",
        html_content,
        text_content,
        attachments
    )
    
    # Update status to sent
    await db.purchase_orders.update_one(
        {"po_id": po_id},
        {"$set": {"status": "sent", "sent_date": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_audit(user.user_id, user.name, "update", "purchase_order", po_id, po.get('po_number'), {"action": "sent_email", "to": po.get('supplier_email')})
    
    return {"message": f"Orden enviada a {po.get('supplier_email')}", "status": "sent"}

@api_router.get("/audit-logs", response_model=List[AuditLog])
async def get_audit_logs(
    limit: int = 100,
    entity_type: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Only admins can view audit logs
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver el historial")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Clean up any corrupted data before returning
    clean_logs = []
    for log in logs:
        # Fix entity_name if it's not a string
        if not isinstance(log.get('entity_name'), str):
            if isinstance(log.get('entity_name'), dict):
                # Extract meaningful info from dict or use default
                log['details'] = log.get('details') or log.get('entity_name')
                log['entity_name'] = "N/A"
            else:
                log['entity_name'] = str(log.get('entity_name', 'N/A'))
        
        try:
            clean_logs.append(AuditLog(**log))
        except Exception:
            # Skip invalid entries
            continue
    
    return clean_logs

class ClearAuditLogsRequest(BaseModel):
    password: str

@api_router.post("/audit-logs/clear")
async def clear_audit_logs(
    data: ClearAuditLogsRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Only super admins can clear audit logs
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo el Super Admin puede eliminar el historial de auditoría")
    
    # Verify password
    user_doc = await db.users.find_one({"id": user.id}, {"_id": 0})
    if not user_doc or not bcrypt.checkpw(data.password.encode('utf-8'), user_doc.get('password', '').encode('utf-8')):
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    
    # Get count before deletion
    count_before = await db.audit_logs.count_documents({})
    
    # Delete all audit logs
    result = await db.audit_logs.delete_many({})
    
    # Get all super admins to notify
    super_admins = await db.users.find(
        {"role": UserRole.SUPER_ADMIN.value, "email": {"$exists": True, "$ne": ""}},
        {"_id": 0, "email": 1, "name": 1}
    ).to_list(100)
    
    # Send notification email to all super admins
    from email_service import send_email
    
    for admin in super_admins:
        if admin.get('email'):
            subject = "⚠️ ALERTA: Historial de Auditoría Eliminado"
            body = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background: #dc2626; color: white; padding: 20px; text-align: center;">
                    <h1 style="margin: 0;">⚠️ Alerta de Seguridad</h1>
                </div>
                <div style="padding: 30px; background: #f8fafc;">
                    <h2 style="color: #1e293b;">Historial de Auditoría Eliminado</h2>
                    <p style="color: #475569;">Se ha realizado una acción crítica en el sistema:</p>
                    <div style="background: white; padding: 20px; border-radius: 8px; border-left: 4px solid #dc2626;">
                        <p><strong>Acción:</strong> Eliminación completa del historial de auditoría</p>
                        <p><strong>Ejecutado por:</strong> {user.name} ({user.email})</p>
                        <p><strong>Registros eliminados:</strong> {result.deleted_count}</p>
                        <p><strong>Fecha y hora:</strong> {datetime.now(PUERTO_RICO_TZ).strftime('%d/%m/%Y %H:%M:%S')}</p>
                    </div>
                    <p style="color: #64748b; font-size: 12px; margin-top: 20px;">
                        Este es un mensaje automático de seguridad. Si no autorizaste esta acción, contacta inmediatamente al administrador del sistema.
                    </p>
                </div>
            </div>
            """
            background_tasks.add_task(send_email, admin['email'], subject, body)
    
    # Log this critical action
    await log_audit(
        user_id=user.id,
        user_name=user.name,
        action="delete",
        entity_type="audit_logs",
        entity_id="all",
        entity_name="Historial de Auditoría",
        details={"records_deleted": result.deleted_count, "notified_admins": len(super_admins)}
    )
    
    return {"message": f"Se eliminaron {result.deleted_count} registros del historial. Se notificó a {len(super_admins)} administradores."}

# ==================== DATA EXPORT/IMPORT ====================

@api_router.get("/data/export")
async def export_all_data(request: Request, session_token: Optional[str] = Cookie(None)):
    """Export all application data as JSON for migration"""
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo Super Admin puede exportar datos")
    
    # Collections to export
    collections_to_export = [
        'users', 'projects', 'tasks', 'clients', 'invoices', 'estimates',
        'expenses', 'budget_categories', 'labor_entries', 'timesheets',
        'purchase_orders', 'change_orders', 'clock_entries', 'payroll_runs',
        'pay_stubs', 'requests', 'approvals', 'safety_checklists',
        'safety_checklist_templates', 'safety_observations', 'safety_incidents',
        'toolbox_talks', 'toolbox_topics', 'company', 'settings',
        'labor_rates', 'nomenclatures', 'tax_types', 'payroll_settings',
        'documents_from_client', 'documents_to_client', 'notifications'
    ]
    
    export_data = {
        "export_date": datetime.now(PUERTO_RICO_TZ).isoformat(),
        "exported_by": user.name,
        "version": "1.0",
        "collections": {}
    }
    
    for collection_name in collections_to_export:
        try:
            collection = db[collection_name]
            docs = await collection.find({}, {"_id": 0}).to_list(10000)
            export_data["collections"][collection_name] = docs
        except Exception as e:
            export_data["collections"][collection_name] = []
    
    return export_data

@api_router.post("/data/import")
async def import_all_data(
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Import application data from JSON export"""
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo Super Admin puede importar datos")
    
    if "collections" not in data:
        raise HTTPException(status_code=400, detail="Formato de archivo inválido")
    
    results = {}
    
    for collection_name, documents in data["collections"].items():
        if not documents or not isinstance(documents, list):
            results[collection_name] = {"imported": 0, "skipped": True}
            continue
        
        try:
            collection = db[collection_name]
            
            # For each document, try to insert or update
            imported_count = 0
            for doc in documents:
                if not doc:
                    continue
                
                # Find unique identifier
                id_field = None
                for field in ['id', 'user_id', 'project_id', 'task_id', 'invoice_id', 
                             'estimate_id', 'expense_id', 'clock_id', 'rate_id',
                             'checklist_id', 'observation_id', 'incident_id', 'talk_id',
                             'order_id', 'po_id', 'category_id', 'nomenclature_id']:
                    if field in doc:
                        id_field = field
                        break
                
                if id_field:
                    # Upsert: update if exists, insert if not
                    await collection.update_one(
                        {id_field: doc[id_field]},
                        {"$set": doc},
                        upsert=True
                    )
                else:
                    # Just insert for collections without clear ID
                    await collection.insert_one(doc)
                
                imported_count += 1
            
            results[collection_name] = {"imported": imported_count}
        except Exception as e:
            results[collection_name] = {"error": str(e)}
    
    # Log the import action
    await log_audit(
        user_id=user.id,
        user_name=user.name,
        action="create",
        entity_type="data_import",
        entity_id="bulk",
        entity_name="Importación de Datos",
        details={"results": results}
    )
    
    return {"message": "Importación completada", "results": results}

@api_router.post("/data/clear-all")
async def clear_all_data(
    request: Request,
    body: dict,
    session_token: Optional[str] = Cookie(None)
):
    """Clear all test data from the application - requires super admin password"""
    user = await get_current_user(request, session_token)
    
    # Check for super_admin role (handle both enum value and string)
    user_role = str(user.role).lower() if user.role else ""
    if user_role not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail=f"Solo Super Admin puede limpiar datos. Tu rol: {user.role}")
    
    # Verify password for extra security
    password = body.get("password", "")
    if not password:
        raise HTTPException(status_code=400, detail="Se requiere contraseña para confirmar")
    
    # Verify the super admin's password - try both user_id and email for robustness
    db_user = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    if not db_user:
        # Fallback: try finding by email
        db_user = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not db_user or not bcrypt.checkpw(password.encode('utf-8'), db_user.get("password", db_user.get("password_hash", "")).encode('utf-8')):
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    
    # Collections to clear (excluding users, company settings, and system config)
    collections_to_clear = [
        'projects', 'tasks', 'clients', 'invoices', 'estimates',
        'expenses', 'budget_categories', 'labor', 'timesheet',
        'purchase_orders', 'change_orders', 'clock_entries', 'payroll_runs',
        'pay_stubs', 'requests', 'approvals', 'safety_checklists',
        'safety_templates', 'safety_observations', 'safety_incidents',
        'toolbox_talks', 'notifications', 'project_logs',
        'documents', 'comments', 'audit_logs', 'cost_estimates',
        'client_documents', 'employee_documents', 'employee_profiles',
        'project_document_status', 'payments', 'saved_clients',
        'required_documents'
    ]
    
    results = {}
    total_deleted = 0
    
    for collection_name in collections_to_clear:
        try:
            collection = db[collection_name]
            result = await collection.delete_many({})
            deleted = result.deleted_count
            results[collection_name] = {"deleted": deleted}
            total_deleted += deleted
        except Exception as e:
            results[collection_name] = {"error": str(e)}
    
    # Log the action (create new audit log entry after clearing)
    await log_audit(
        user_id=user.user_id,
        user_name=user.name,
        action="delete",
        entity_type="data_clear",
        entity_id="all",
        entity_name="Limpieza Total de Datos",
        details={"results": results, "total_deleted": total_deleted}
    )
    
    return {
        "message": f"Datos limpiados correctamente. {total_deleted} registros eliminados.",
        "results": results,
        "total_deleted": total_deleted
    }

@api_router.get("/integrations", response_model=List[IntegrationConfig])
async def get_integrations(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver integraciones")
    
    integrations = await db.integrations.find({}, {"_id": 0}).to_list(100)
    return [IntegrationConfig(**i) for i in integrations]

@api_router.post("/integrations", response_model=IntegrationConfig)
async def create_or_update_integration(
    integration_type: str,
    enabled: bool,
    config: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden configurar integraciones")
    
    # Check if integration already exists
    existing = await db.integrations.find_one({"integration_type": integration_type}, {"_id": 0})
    
    if existing:
        # Update
        await db.integrations.update_one(
            {"integration_type": integration_type},
            {"$set": {
                "enabled": enabled,
                "config": config,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        integration_id = existing['integration_id']
    else:
        # Create
        integration_id = f"int_{uuid4().hex[:16]}"
        integration_doc = {
            "integration_id": integration_id,
            "integration_type": integration_type,
            "enabled": enabled,
            "config": config,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.integrations.insert_one(integration_doc)
    
    # Test Slack webhook if it's a Slack integration
    if integration_type == "slack" and enabled:
        webhook_url = config.get('webhook_url')
        if webhook_url:
            await send_slack_notification(
                webhook_url,
                "✅ Integración de Slack configurada exitosamente!",
                "Prueba de Conexión",
                "good"
            )
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "update" if existing else "create",
        "integration",
        integration_id,
        f"{integration_type} integration",
        {"enabled": enabled}
    )
    
    integration = await db.integrations.find_one({"integration_id": integration_id}, {"_id": 0})
    return IntegrationConfig(**integration)

@api_router.delete("/integrations/{integration_id}")
async def delete_integration(
    integration_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar integraciones")
    
    integration = await db.integrations.find_one({"integration_id": integration_id}, {"_id": 0})
    if not integration:
        raise HTTPException(status_code=404, detail="Integración no encontrada")
    
    await db.integrations.delete_one({"integration_id": integration_id})
    
    # Log audit
    await log_audit(
        user.user_id,
        user.name,
        "delete",
        "integration",
        integration_id,
        f"{integration.get('integration_type')} integration",
        {}
    )
    
    return {"message": "Integración eliminada exitosamente"}

@api_router.post("/integrations/test-slack")
async def test_slack_integration(
    webhook_url: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden probar integraciones")
    
    try:
        await send_slack_notification(
            webhook_url,
            "🎉 La conexión con Slack funciona correctamente!",
            "Prueba de Integración",
            "good"
        )
        return {"message": "Notificación de prueba enviada exitosamente"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al enviar notificación: {str(e)}")

# ==================== LABOR RATES MODELS ====================
class LaborRate(BaseModel):
    rate_id: str
    role_name: str
    quoted_rate: float
    assumed_rate: float
    overtime_rate: float
    created_at: str
    updated_at: str

class LaborRateCreate(BaseModel):
    role_name: str
    quoted_rate: float
    assumed_rate: float
    overtime_rate: float

class LaborRateUpdate(BaseModel):
    role_name: Optional[str] = None
    quoted_rate: Optional[float] = None
    assumed_rate: Optional[float] = None
    overtime_rate: Optional[float] = None

# ==================== COST ESTIMATES MODELS ====================
class LaborCostItem(BaseModel):
    role_name: str
    qty_personnel: int = 0
    regular_hours: float = 0
    overtime_hours: float = 0
    rate: float = 0
    overtime_rate: float = 0
    subtotal: float = 0

class SubcontractorItem(BaseModel):
    trade: str  # Civil, Mechanical, Electrical
    description: str = ""
    cost: float = 0

class MaterialItem(BaseModel):
    description: str
    quantity: float = 0
    unit_cost: float = 0
    total: float = 0

class EquipmentItem(BaseModel):
    description: str
    quantity: int = 0
    days: int = 0
    rate: float = 0
    total: float = 0

class TransportationItem(BaseModel):
    description: str
    city_town: str = ""
    roundtrip_miles: float = 0
    cost_per_mile: float = 0
    days: int = 0
    total: float = 0

class GeneralConditionItem(BaseModel):
    description: str
    quantity: float = 0
    unit_cost: float = 0
    total: float = 0

class CostEstimate(BaseModel):
    estimate_id: str
    project_id: Optional[str] = None
    project_name: Optional[str] = ""
    estimate_name: str
    labor_costs: List[LaborCostItem] = []
    subcontractors: List[SubcontractorItem] = []
    materials: List[MaterialItem] = []
    equipment: List[EquipmentItem] = []
    general_conditions: List[GeneralConditionItem] = []
    transportation: List[TransportationItem] = []
    overhead_percentage: float = 0
    profit_percentage: float = 0
    contingency_percentage: float = 0
    tax_percentage: float = 0
    b2b_percentage: float = 0
    cfse_percentage: float = 0
    liability_percentage: float = 0
    municipal_patent_percentage: float = 0
    total_labor: float = 0
    total_subcontractors: float = 0
    total_materials: float = 0
    total_equipment: float = 0
    total_transportation: float = 0
    total_general_conditions: float = 0
    subtotal: float = 0
    grand_total: float = 0
    created_by: str
    created_at: str
    updated_at: str

class CostEstimateCreate(BaseModel):
    project_id: Optional[str] = None
    estimate_name: str
    labor_costs: List[LaborCostItem] = []
    subcontractors: List[SubcontractorItem] = []
    materials: List[MaterialItem] = []
    equipment: List[EquipmentItem] = []
    transportation: List[TransportationItem] = []
    general_conditions: List[GeneralConditionItem] = []
    overhead_percentage: float = 0
    profit_percentage: float = 0
    contingency_percentage: float = 0
    tax_percentage: float = 0
    b2b_percentage: float = 0
    cfse_percentage: float = 0
    liability_percentage: float = 0
    municipal_patent_percentage: float = 0

# ==================== LABOR RATES ENDPOINTS ====================
@api_router.get("/labor-rates", response_model=List[LaborRate])
async def get_labor_rates(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    rates = await db.labor_rates.find({}, {"_id": 0}).to_list(1000)
    return rates

@api_router.post("/labor-rates", response_model=LaborRate)
async def create_labor_rate(
    rate_data: LaborRateCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear tarifas")
    
    rate_id = f"lr_{uuid4().hex[:16]}"
    now = datetime.now(PUERTO_RICO_TZ).isoformat()
    
    rate_doc = {
        "rate_id": rate_id,
        "role_name": rate_data.role_name,
        "quoted_rate": rate_data.quoted_rate,
        "assumed_rate": rate_data.assumed_rate,
        "overtime_rate": rate_data.overtime_rate,
        "created_at": now,
        "updated_at": now
    }
    
    await db.labor_rates.insert_one(rate_doc)
    return LaborRate(**rate_doc)

@api_router.put("/labor-rates/{rate_id}", response_model=LaborRate)
async def update_labor_rate(
    rate_id: str,
    rate_data: LaborRateUpdate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden actualizar tarifas")
    
    rate = await db.labor_rates.find_one({"rate_id": rate_id}, {"_id": 0})
    if not rate:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    
    update_data = {"updated_at": datetime.now(PUERTO_RICO_TZ).isoformat()}
    
    if rate_data.role_name is not None:
        update_data["role_name"] = rate_data.role_name
    if rate_data.quoted_rate is not None:
        update_data["quoted_rate"] = rate_data.quoted_rate
    if rate_data.assumed_rate is not None:
        update_data["assumed_rate"] = rate_data.assumed_rate
    if rate_data.overtime_rate is not None:
        update_data["overtime_rate"] = rate_data.overtime_rate
    
    await db.labor_rates.update_one({"rate_id": rate_id}, {"$set": update_data})
    
    updated_rate = await db.labor_rates.find_one({"rate_id": rate_id}, {"_id": 0})
    return LaborRate(**updated_rate)

@api_router.delete("/labor-rates/{rate_id}")
async def delete_labor_rate(
    rate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar tarifas")
    
    result = await db.labor_rates.delete_one({"rate_id": rate_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    
    return {"message": "Tarifa eliminada exitosamente"}

# ==================== COST ESTIMATES ENDPOINTS ====================
@api_router.get("/cost-estimates", response_model=List[CostEstimate])
async def get_cost_estimates(
    project_id: Optional[str] = None,
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    query = {}
    if project_id:
        query["project_id"] = project_id
    
    estimates = await db.cost_estimates.find(query, {"_id": 0}).to_list(1000)
    return estimates

@api_router.get("/cost-estimates/{estimate_id}", response_model=CostEstimate)
async def get_cost_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    return CostEstimate(**estimate)

@api_router.post("/cost-estimates", response_model=CostEstimate)
async def create_cost_estimate(
    estimate_data: CostEstimateCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Get project name if project_id is provided
    project_name = ""
    if estimate_data.project_id:
        project = await db.projects.find_one({"project_id": estimate_data.project_id}, {"_id": 0})
        if project:
            project_name = project.get("name", "")
    
    estimate_id = f"ce_{uuid4().hex[:16]}"
    now = datetime.now(PUERTO_RICO_TZ).isoformat()
    
    # Calculate totals
    total_labor = sum(item.subtotal for item in estimate_data.labor_costs)
    total_subcontractors = sum(item.cost for item in estimate_data.subcontractors)
    total_materials = sum(item.total for item in estimate_data.materials)
    total_equipment = sum(item.total for item in estimate_data.equipment)
    total_transportation = sum(item.total for item in estimate_data.transportation)
    total_general_conditions = sum(item.total for item in estimate_data.general_conditions)
    
    subtotal = (
        total_labor + 
        total_subcontractors + 
        total_materials + 
        total_equipment + 
        total_transportation +
        total_general_conditions
    )
    
    # Apply B2B percentage only to subcontractors
    b2b_amount = total_subcontractors * (estimate_data.b2b_percentage / 100)
    
    # Apply other percentages to subtotal
    grand_total = subtotal * (
        1 + estimate_data.overhead_percentage / 100 +
        estimate_data.profit_percentage / 100 +
        estimate_data.contingency_percentage / 100 +
        estimate_data.tax_percentage / 100 +
        estimate_data.cfse_percentage / 100 +
        estimate_data.liability_percentage / 100 +
        estimate_data.municipal_patent_percentage / 100
    ) + b2b_amount
    
    estimate_doc = {
        "estimate_id": estimate_id,
        "project_id": estimate_data.project_id or "",
        "project_name": project_name,
        "estimate_name": estimate_data.estimate_name,
        "labor_costs": [item.dict() for item in estimate_data.labor_costs],
        "subcontractors": [item.dict() for item in estimate_data.subcontractors],
        "materials": [item.dict() for item in estimate_data.materials],
        "equipment": [item.dict() for item in estimate_data.equipment],
        "transportation": [item.dict() for item in estimate_data.transportation],
        "general_conditions": [item.dict() for item in estimate_data.general_conditions],
        "overhead_percentage": estimate_data.overhead_percentage,
        "profit_percentage": estimate_data.profit_percentage,
        "contingency_percentage": estimate_data.contingency_percentage,
        "tax_percentage": estimate_data.tax_percentage,
        "b2b_percentage": estimate_data.b2b_percentage,
        "cfse_percentage": estimate_data.cfse_percentage,
        "liability_percentage": estimate_data.liability_percentage,
        "municipal_patent_percentage": estimate_data.municipal_patent_percentage,
        "total_labor": round(total_labor, 2),
        "total_subcontractors": round(total_subcontractors, 2),
        "total_materials": round(total_materials, 2),
        "total_equipment": round(total_equipment, 2),
        "total_transportation": round(total_transportation, 2),
        "total_general_conditions": round(total_general_conditions, 2),
        "subtotal": round(subtotal, 2),
        "grand_total": round(grand_total, 2),
        "created_by": user.user_id,
        "created_at": now,
        "updated_at": now
    }
    
    await db.cost_estimates.insert_one(estimate_doc)
    return CostEstimate(**estimate_doc)

@api_router.put("/cost-estimates/{estimate_id}", response_model=CostEstimate)
async def update_cost_estimate(
    estimate_id: str,
    estimate_data: CostEstimateCreate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    # Calculate totals
    total_labor = sum(item.subtotal for item in estimate_data.labor_costs)
    total_subcontractors = sum(item.cost for item in estimate_data.subcontractors)
    total_materials = sum(item.total for item in estimate_data.materials)
    total_equipment = sum(item.total for item in estimate_data.equipment)
    total_transportation = sum(item.total for item in estimate_data.transportation)
    total_general_conditions = sum(item.total for item in estimate_data.general_conditions)
    
    subtotal = (
        total_labor + 
        total_subcontractors + 
        total_materials + 
        total_equipment + 
        total_transportation +
        total_general_conditions
    )
    
    # Apply B2B percentage only to subcontractors
    b2b_amount = total_subcontractors * (estimate_data.b2b_percentage / 100)
    
    # Apply other percentages to subtotal
    grand_total = subtotal * (
        1 + estimate_data.overhead_percentage / 100 +
        estimate_data.profit_percentage / 100 +
        estimate_data.contingency_percentage / 100 +
        estimate_data.tax_percentage / 100 +
        estimate_data.cfse_percentage / 100 +
        estimate_data.liability_percentage / 100 +
        estimate_data.municipal_patent_percentage / 100
    ) + b2b_amount
    
    # Get project name if project_id is provided
    project_name = ""
    if estimate_data.project_id:
        project = await db.projects.find_one({"project_id": estimate_data.project_id}, {"_id": 0})
        if project:
            project_name = project.get("name", "")
    
    update_doc = {
        "project_id": estimate_data.project_id or "",
        "project_name": project_name,
        "estimate_name": estimate_data.estimate_name,
        "labor_costs": [item.dict() for item in estimate_data.labor_costs],
        "subcontractors": [item.dict() for item in estimate_data.subcontractors],
        "materials": [item.dict() for item in estimate_data.materials],
        "equipment": [item.dict() for item in estimate_data.equipment],
        "transportation": [item.dict() for item in estimate_data.transportation],
        "general_conditions": [item.dict() for item in estimate_data.general_conditions],
        "overhead_percentage": estimate_data.overhead_percentage,
        "profit_percentage": estimate_data.profit_percentage,
        "contingency_percentage": estimate_data.contingency_percentage,
        "tax_percentage": estimate_data.tax_percentage,
        "b2b_percentage": estimate_data.b2b_percentage,
        "cfse_percentage": estimate_data.cfse_percentage,
        "liability_percentage": estimate_data.liability_percentage,
        "municipal_patent_percentage": estimate_data.municipal_patent_percentage,
        "total_labor": round(total_labor, 2),
        "total_subcontractors": round(total_subcontractors, 2),
        "total_materials": round(total_materials, 2),
        "total_equipment": round(total_equipment, 2),
        "total_transportation": round(total_transportation, 2),
        "total_general_conditions": round(total_general_conditions, 2),
        "subtotal": round(subtotal, 2),
        "grand_total": round(grand_total, 2),
        "updated_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.cost_estimates.update_one({"estimate_id": estimate_id}, {"$set": update_doc})
    
    updated_estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    return CostEstimate(**updated_estimate)

@api_router.delete("/cost-estimates/{estimate_id}")
async def delete_cost_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    result = await db.cost_estimates.delete_one({"estimate_id": estimate_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    return {"message": "Estimación eliminada exitosamente"}

# Convert Cost Estimate to Formal Estimate
@api_router.post("/cost-estimates/{estimate_id}/convert-to-estimate", response_model=Estimate)
async def convert_cost_estimate_to_estimate(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Convert a cost estimate into a formal estimate (Estimado)"""
    user = await get_current_user(request, session_token)
    
    # Get the cost estimate
    cost_estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not cost_estimate:
        raise HTTPException(status_code=404, detail="Estimación de costo no encontrada")
    
    if not cost_estimate.get("project_id"):
        raise HTTPException(status_code=400, detail="La estimación de costo debe tener un proyecto asignado")
    
    # Get project info
    project = await db.projects.find_one({"project_id": cost_estimate["project_id"]}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Helper function to round
    def round2(num):
        return round(num, 2)
    
    # Calculate totals from cost estimate data
    labor_costs = cost_estimate.get("labor_costs", [])
    subcontractors = cost_estimate.get("subcontractors", [])
    materials = cost_estimate.get("materials", [])
    equipment = cost_estimate.get("equipment", [])
    transportation = cost_estimate.get("transportation", [])
    general_conditions = cost_estimate.get("general_conditions", [])
    
    # Get percentages (with defaults for fixed values)
    profit_pct = cost_estimate.get("profit_percentage", 0)
    overhead_pct = cost_estimate.get("overhead_percentage", 0)
    cfse_pct = cost_estimate.get("cfse_percentage", 7)  # Fixed 7%
    liability_pct = cost_estimate.get("liability_percentage", 7)  # Fixed 7%
    municipal_pct = cost_estimate.get("municipal_patent_percentage", 1)  # Fixed 1%
    contingency_pct = cost_estimate.get("contingency_percentage", 6)  # Fixed 6%
    b2b_ohsms_pct = cost_estimate.get("b2b_ohsms_percentage", 0)
    b2b_ohsms_labor_pct = cost_estimate.get("b2b_ohsms_labor_percentage", 4)  # Fixed 4%
    b2b_subcontractor_pct = cost_estimate.get("b2b_subcontractor_percentage", 0)
    
    # Calculate subtotals
    total_labor = round2(sum(float(item.get("subtotal", 0)) for item in labor_costs))
    total_subcontractors = round2(sum(float(item.get("cost", 0)) for item in subcontractors))
    total_subcontractor_labor = round2(sum(float(item.get("labor_cost", 0)) for item in subcontractors))
    total_materials = round2(sum(float(item.get("total", 0)) for item in materials))
    total_equipment = round2(sum(float(item.get("total", 0)) for item in equipment))
    total_transportation = round2(sum(float(item.get("total", 0)) for item in transportation))
    total_gc = round2(sum(float(item.get("total", 0)) for item in general_conditions))
    
    subtotal = round2(total_labor + total_subcontractors + total_materials + 
                      total_equipment + total_transportation + total_gc)
    
    # Calculate cascading percentages
    # Step 1: Subtotal x (1 + Profit%) = s
    after_profit = round2(subtotal * (1 + profit_pct / 100))
    profit_amount = round2(after_profit - subtotal)
    
    # Step 2: s x (1 + Overhead%) = w
    after_overhead = round2(after_profit * (1 + overhead_pct / 100))
    overhead_amount = round2(after_overhead - after_profit)
    
    # Step 3: CFSE on labor only
    cfse_amount = round2(total_labor * (cfse_pct / 100))
    
    # Step 4: w + cfseAmount = combined
    combined_total = round2(after_overhead + cfse_amount)
    
    # Step 5: combined x (1 + Liability%) = M
    after_liability = round2(combined_total * (1 + liability_pct / 100))
    liability_amount = round2(after_liability - combined_total)
    
    # Step 6: M x (1 + Municipal%) = C
    after_municipal = round2(after_liability * (1 + municipal_pct / 100))
    municipal_amount = round2(after_municipal - after_liability)
    
    # Step 7: C x (1 + Contingency%) = U
    after_contingency = round2(after_municipal * (1 + contingency_pct / 100))
    contingency_amount = round2(after_contingency - after_municipal)
    
    # Step 8: B2B OHSMS Global
    b2b_ohsms_amount = round2(after_contingency * (b2b_ohsms_pct / 100))
    after_b2b_ohsms = round2(after_contingency + b2b_ohsms_amount)
    
    # B2B Subcontractor
    b2b_subcontractor_amount = round2(total_subcontractor_labor * (b2b_subcontractor_pct / 100))
    
    # Calculate Price Breakdown (Material/Equipment vs Labor)
    total_material_equipment = round2(total_subcontractors + total_materials + total_equipment + total_transportation + total_gc)
    
    # Calculate Labor ratio and Labor for Price Breakdown
    labor_ratio = total_labor / subtotal if subtotal > 0 else 0
    mat_equip_ratio = total_material_equipment / subtotal if subtotal > 0 else 0
    
    # Labor del Price Breakdown = after_b2b_ohsms * labor_ratio (CFSE ya está incluido en cascade)
    labor_for_price_breakdown = round2(after_b2b_ohsms * labor_ratio)
    
    # B2B M.O. = Labor (del Price Breakdown) × 4%
    b2b_ohsms_labor_amount = round2(labor_for_price_breakdown * (b2b_ohsms_labor_pct / 100))
    
    # Grand total
    grand_total = round2(after_b2b_ohsms + b2b_subcontractor_amount + b2b_ohsms_labor_amount)
    
    # Labor with percentages = Labor proporción del cascade + B2B OHSMS Labor
    labor_with_percentages = round2(labor_for_price_breakdown + b2b_ohsms_labor_amount)
    # Material/Equipment with percentages = proporción mat/equip del cascade + B2B Subcontractor
    mat_equip_with_percentages = round2((after_b2b_ohsms * mat_equip_ratio) + b2b_subcontractor_amount)
    
    # Total percentage amounts
    total_pct_amounts = round2(
        profit_amount + overhead_amount + cfse_amount + liability_amount +
        municipal_amount + contingency_amount + b2b_ohsms_amount +
        b2b_ohsms_labor_amount + b2b_subcontractor_amount
    )
    
    # Build estimate items - ONLY Price Breakdown (2 items: Material/Equipment and Labor)
    estimate_items = [
        {
            "description": "Material/Equipment",
            "quantity": 1,
            "unit_price": mat_equip_with_percentages,
            "amount": mat_equip_with_percentages
        },
        {
            "description": "Labor",
            "quantity": 1,
            "unit_price": labor_with_percentages,
            "amount": labor_with_percentages
        }
    ]
    
    # Price Breakdown data
    price_breakdown_data = {
        "material_equipment": mat_equip_with_percentages,
        "labor": labor_with_percentages,
        "total": grand_total
    }
    
    # Generate estimate number
    company_settings = await db.company_settings.find_one({}, {"_id": 0})
    next_num = company_settings.get("next_estimate_number", 1) if company_settings else 1
    estimate_number = f"EST-{datetime.now().year}-{str(next_num).zfill(4)}"
    await db.company_settings.update_one({}, {"$inc": {"next_estimate_number": 1}}, upsert=True)
    
    new_estimate_id = f"est_{uuid4().hex[:16]}"
    now = datetime.now(timezone.utc).isoformat()
    
    # Create the estimate document
    estimate_doc = {
        "estimate_id": new_estimate_id,
        "estimate_number": estimate_number,
        "project_id": cost_estimate["project_id"],
        "project_name": project.get("name"),
        "client_profile_id": None,
        "client_company": project.get("client_name", ""),
        "client_name": project.get("client_name", "Cliente"),
        "client_email": None,
        "client_phone": None,
        "client_address": None,
        "title": f"Estimado - {cost_estimate.get('estimate_name', 'Sin nombre')}",
        "description": f"Generado desde Estimación de Costo: {cost_estimate.get('estimate_name', '')}",
        "items": estimate_items,
        "subtotal": grand_total,
        "discount_percent": 0,
        "discount_amount": 0,
        "tax_rate": 0,
        "tax_amount": 0,
        "selected_taxes": [],
        "total": grand_total,
        "status": "draft",
        "notes": cost_estimate.get("notes") or f"Price Breakdown desde estimación de costo: {cost_estimate.get('estimate_name', '')}",
        "terms": cost_estimate.get("terms"),
        "valid_until": None,
        "created_by": user.user_id,
        "created_by_name": user.name,
        "created_at": now,
        "sent_date": None,
        "approved_date": None,
        "converted_invoice_id": None,
        "price_breakdown": price_breakdown_data
    }
    
    await db.estimates.insert_one(estimate_doc)
    await log_audit(user.user_id, user.name, "create", "estimate", new_estimate_id, estimate_number, 
                   {"source": "cost_estimate", "cost_estimate_id": estimate_id})
    
    return Estimate(**estimate_doc)

# ==================== COST ESTIMATE EXPORT ENDPOINTS ====================
@api_router.get("/cost-estimates/{estimate_id}/export/pdf")
async def export_cost_estimate_pdf(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_RIGHT
    import io
    import base64
    
    user = await get_current_user(request, session_token)
    
    estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    # Get company info
    company = await db.company.find_one({}, {"_id": 0}) or {}
    
    # Define corporate colors matching frontend/invoice
    PRIMARY_COLOR = colors.HexColor('#f97316')  # Orange
    SECONDARY_COLOR = colors.HexColor('#475569')  # Slate
    TEXT_COLOR = colors.HexColor('#1e293b')  # Dark
    LIGHT_BG = colors.HexColor('#f8fafc')  # Light gray
    
    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    
    # Custom styles matching invoice CSS
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=TEXT_COLOR, spaceAfter=5, alignment=2)  # Right aligned
    company_style = ParagraphStyle('Company', parent=styles['Normal'], fontSize=10, textColor=PRIMARY_COLOR, fontName='Helvetica-Bold')
    company_detail_style = ParagraphStyle('CompanyDetail', parent=styles['Normal'], fontSize=8, textColor=SECONDARY_COLOR)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=11, textColor=TEXT_COLOR, spaceAfter=8, spaceBefore=15, fontName='Helvetica-Bold')
    normal_right = ParagraphStyle('NormalRight', parent=styles['Normal'], alignment=2)
    subtitle_right = ParagraphStyle('SubtitleRight', parent=styles['Normal'], fontSize=9, textColor=SECONDARY_COLOR, alignment=2)
    
    elements = []
    
    # Build header table with logo on left, doc info on right (matching invoice style)
    # Left side: Company logo and info
    left_elements = []
    
    # Try to load company logo
    logo_path = None
    if company.get('company_logo'):
        logo_url = company.get('company_logo', '')
        if logo_url.startswith('/api/uploads/'):
            logo_path = ROOT_DIR / logo_url.replace('/api/uploads/', 'uploads/')
        elif logo_url.startswith('/uploads/'):
            logo_path = ROOT_DIR / logo_url.lstrip('/')
    
    if logo_path and logo_path.exists():
        try:
            logo_img = Image(str(logo_path), width=1.5*inch, height=0.75*inch)
            left_elements.append(logo_img)
            left_elements.append(Spacer(1, 5))
        except:
            pass
    
    # Company name and details
    company_name = company.get('company_name', 'OHSMS PR')
    left_elements.append(Paragraph(company_name, company_style))
    
    if company.get('address'):
        left_elements.append(Paragraph(company.get('address', ''), company_detail_style))
    if company.get('city') or company.get('state'):
        left_elements.append(Paragraph(f"{company.get('city', '')}, {company.get('state', '')} {company.get('zip_code', '')}", company_detail_style))
    if company.get('phone'):
        left_elements.append(Paragraph(f"Tel: {company.get('phone', '')}", company_detail_style))
    if company.get('email'):
        left_elements.append(Paragraph(company.get('email', ''), company_detail_style))
    
    # Right side: Document title and info
    right_elements = []
    right_elements.append(Paragraph("ESTIMACIÓN DE COSTOS", title_style))
    right_elements.append(Paragraph(f"#{estimate.get('estimate_name', 'Sin nombre')}", subtitle_right))
    right_elements.append(Paragraph(f"Proyecto: {estimate.get('project_name', 'N/A')}", subtitle_right))
    right_elements.append(Paragraph(f"Fecha: {datetime.now(PUERTO_RICO_TZ).strftime('%d/%m/%Y')}", subtitle_right))
    
    # Create header table with two columns
    from reportlab.platypus import KeepTogether
    header_table_data = [[left_elements, right_elements]]
    header_table = Table(header_table_data, colWidths=[3.5*inch, 3.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    
    # Orange separator line
    elements.append(Spacer(1, 10))
    line_table = Table([['']], colWidths=[7*inch])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 1, PRIMARY_COLOR),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 15))
    
    # Table style matching invoice CSS
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_COLOR),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_COLOR),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, PRIMARY_COLOR),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fcfcfd')]),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
    ])
    
    footer_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_COLOR),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (-1, 0), (-1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
    ])
    
    # Labor Costs Section
    labor_costs = estimate.get('labor_costs', [])
    if labor_costs:
        elements.append(Paragraph("Mano de Obra", heading_style))
        labor_data = [['Rol', 'Cant.', 'Hrs Reg.', 'Hrs OT', 'Tarifa', 'Tarifa OT', 'Subtotal']]
        for item in labor_costs:
            qty = item.get('qty_personnel', 1)
            reg_hrs = item.get('regular_hours', 0)
            ot_hrs = item.get('overtime_hours', 0)
            rate = item.get('rate', 0)
            ot_rate = item.get('overtime_rate', 0)
            subtotal = item.get('subtotal', 0)
            labor_data.append([
                item.get('role_name', ''),
                str(qty),
                str(reg_hrs),
                str(ot_hrs),
                f"${float(rate):,.2f}",
                f"${float(ot_rate):,.2f}",
                f"${float(subtotal):,.2f}"
            ])
        t = Table(labor_data, colWidths=[1.5*inch, 0.5*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.9*inch, 1.3*inch])
        t.setStyle(table_style)
        elements.append(t)
        # Footer row
        total_labor_sum = sum(float(item.get('subtotal', 0)) for item in labor_costs)
        footer_data = [['', '', '', '', '', 'Total:', f"${total_labor_sum:,.2f}"]]
        tf = Table(footer_data, colWidths=[1.5*inch, 0.5*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.9*inch, 1.3*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # Subcontractors Section
    subcontractors = estimate.get('subcontractors', [])
    if subcontractors:
        elements.append(Paragraph("Subcontratistas", heading_style))
        sub_data = [['Tipo', 'Descripción', 'Costo Total', 'Mano de Obra']]
        for item in subcontractors:
            sub_data.append([
                item.get('trade', ''),
                item.get('description', '')[:35],
                f"${float(item.get('cost', 0)):,.2f}",
                f"${float(item.get('labor_cost', 0)):,.2f}"
            ])
        t = Table(sub_data, colWidths=[1.3*inch, 2.7*inch, 1.25*inch, 1.25*inch])
        t.setStyle(table_style)
        elements.append(t)
        total_sub_sum = sum(float(item.get('cost', 0)) for item in subcontractors)
        total_sub_labor = sum(float(item.get('labor_cost', 0)) for item in subcontractors)
        footer_data = [['', 'Total:', f"${total_sub_sum:,.2f}", f"${total_sub_labor:,.2f}"]]
        tf = Table(footer_data, colWidths=[1.3*inch, 2.7*inch, 1.25*inch, 1.25*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # Materials Section
    materials = estimate.get('materials', [])
    if materials:
        elements.append(Paragraph("Materiales", heading_style))
        mat_data = [['Descripción', 'Cantidad', 'Precio Unitario', 'Total']]
        for item in materials:
            mat_data.append([
                item.get('description', '')[:30],
                str(item.get('quantity', 0)),
                f"${float(item.get('unit_cost', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}"
            ])
        t = Table(mat_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        t.setStyle(table_style)
        elements.append(t)
        total_mat_sum = sum(float(item.get('total', 0)) for item in materials)
        footer_data = [['', '', 'Total:', f"${total_mat_sum:,.2f}"]]
        tf = Table(footer_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # Equipment Section
    equipment = estimate.get('equipment', [])
    if equipment:
        elements.append(Paragraph("Equipos", heading_style))
        eq_data = [['Descripción', 'Cantidad', 'Días', 'Tarifa/Día', 'Total']]
        for item in equipment:
            eq_data.append([
                item.get('description', '')[:25],
                str(item.get('quantity', 0)),
                str(item.get('days', 0)),
                f"${float(item.get('rate', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}"
            ])
        t = Table(eq_data, colWidths=[2*inch, 1*inch, 0.8*inch, 1.2*inch, 1.5*inch])
        t.setStyle(table_style)
        elements.append(t)
        total_eq_sum = sum(float(item.get('total', 0)) for item in equipment)
        footer_data = [['', '', '', 'Total:', f"${total_eq_sum:,.2f}"]]
        tf = Table(footer_data, colWidths=[2*inch, 1*inch, 0.8*inch, 1.2*inch, 1.5*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # Transportation Section
    transportation = estimate.get('transportation', [])
    if transportation:
        elements.append(Paragraph("Transporte", heading_style))
        trans_data = [['Descripción', 'Ciudad', 'Millas', 'Costo/Milla', 'Días', 'Total']]
        for item in transportation:
            trans_data.append([
                item.get('description', '')[:20],
                item.get('city_town', '')[:15],
                str(item.get('roundtrip_miles', 0)),
                f"${float(item.get('cost_per_mile', 0)):,.2f}",
                str(item.get('days', 0)),
                f"${float(item.get('total', 0)):,.2f}"
            ])
        t = Table(trans_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 0.7*inch, 1.3*inch])
        t.setStyle(table_style)
        elements.append(t)
        total_trans_sum = sum(float(item.get('total', 0)) for item in transportation)
        footer_data = [['', '', '', '', 'Total:', f"${total_trans_sum:,.2f}"]]
        tf = Table(footer_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 0.7*inch, 1.3*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # General Conditions Section
    general_conditions = estimate.get('general_conditions', [])
    if general_conditions:
        elements.append(Paragraph("Condiciones Generales", heading_style))
        gc_data = [['Descripción', 'Cantidad', 'Costo Unitario', 'Total']]
        for item in general_conditions:
            gc_data.append([
                item.get('description', '')[:30],
                str(item.get('quantity', 0)),
                f"${float(item.get('unit_cost', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}"
            ])
        t = Table(gc_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        t.setStyle(table_style)
        elements.append(t)
        total_gc_sum = sum(float(item.get('total', 0)) for item in general_conditions)
        footer_data = [['', '', 'Total:', f"${total_gc_sum:,.2f}"]]
        tf = Table(footer_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        tf.setStyle(footer_style)
        elements.append(tf)
    
    # Summary Section with orange accent - CASCADING CALCULATION
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Resumen", heading_style))
    
    # Helper function to round to 2 decimals
    def round2(num):
        return round(num * 100) / 100
    
    # Get raw data from estimate
    labor_costs = estimate.get('labor_costs', [])
    subcontractors = estimate.get('subcontractors', [])
    materials = estimate.get('materials', [])
    equipment = estimate.get('equipment', [])
    transportation = estimate.get('transportation', [])
    general_conditions = estimate.get('general_conditions', [])
    
    # Calculate totals from line items
    total_labor = round2(sum(float(item.get('subtotal', 0)) for item in labor_costs))
    total_subcontractors = round2(sum(float(item.get('cost', 0)) for item in subcontractors))
    total_subcontractor_labor = round2(sum(float(item.get('labor_cost', 0)) for item in subcontractors))
    total_materials = round2(sum(float(item.get('total', 0)) for item in materials))
    total_equipment = round2(sum(float(item.get('total', 0)) for item in equipment))
    total_transportation = round2(sum(float(item.get('total', 0)) for item in transportation))
    total_gc = round2(sum(float(item.get('total', 0)) for item in general_conditions))
    
    subtotal = round2(total_labor + total_subcontractors + total_materials + 
                      total_equipment + total_transportation + total_gc)
    
    # Get percentages
    profit_pct = float(estimate.get('profit_percentage', 0))
    overhead_pct = float(estimate.get('overhead_percentage', 0))
    cfse_pct = float(estimate.get('cfse_percentage', 0))
    liability_pct = float(estimate.get('liability_percentage', 0))
    municipal_patent_pct = float(estimate.get('municipal_patent_percentage', 0))
    contingency_pct = float(estimate.get('contingency_percentage', 0))
    b2b_ohsms_pct = float(estimate.get('b2b_ohsms_percentage', 0))
    b2b_ohsms_labor_pct = float(estimate.get('b2b_ohsms_labor_percentage', 0))
    b2b_subcontractor_pct = float(estimate.get('b2b_subcontractor_percentage', 0))
    
    # CASCADING CALCULATION matching frontend exactly:
    # B2B Subcontractor - applies only to subcontractor's LABOR COST (added at the end)
    b2b_subcontractor_amount = round2(total_subcontractor_labor * (b2b_subcontractor_pct / 100))
    
    # Step 1: Subtotal x (1 + Profit%) = s
    profit_multiplier = 1 + (profit_pct / 100)
    after_profit = round2(subtotal * profit_multiplier)
    profit_amount = round2(after_profit - subtotal)
    
    # Step 2: s x (1 + Overhead%) = w
    overhead_multiplier = 1 + (overhead_pct / 100)
    after_overhead = round2(after_profit * overhead_multiplier)
    overhead_amount = round2(after_overhead - after_profit)
    
    # Step 3: Mano de Obra x CFSE% = cfseAmount (only the increment)
    cfse_amount = round2(total_labor * (cfse_pct / 100))
    
    # Step 4: w + cfseAmount = qq
    combined_total = round2(after_overhead + cfse_amount)
    
    # Step 5: qq x (1 + Liability%) = M
    liability_multiplier = 1 + (liability_pct / 100)
    after_liability = round2(combined_total * liability_multiplier)
    liability_amount = round2(after_liability - combined_total)
    
    # Step 6: M x (1 + Municipal Patent%) = C
    municipal_patent_multiplier = 1 + (municipal_patent_pct / 100)
    after_municipal_patent = round2(after_liability * municipal_patent_multiplier)
    municipal_patent_amount = round2(after_municipal_patent - after_liability)
    
    # Step 7: C x (1 + Contingency%) = U
    contingency_multiplier = 1 + (contingency_pct / 100)
    after_contingency = round2(after_municipal_patent * contingency_multiplier)
    contingency_amount = round2(after_contingency - after_municipal_patent)
    
    # Step 8: U x 0.35 x B2B OHSMS% = B2B OHSMS Amount
    b2b_ohsms_base = round2(after_contingency * 0.35)
    b2b_ohsms_amount = round2(b2b_ohsms_base * (b2b_ohsms_pct / 100))
    after_b2b_ohsms = round2(after_contingency + b2b_ohsms_amount)
    
    # Calculate Material/Equipment breakdown
    total_material_equipment = round2(total_subcontractors + total_materials + total_equipment + total_transportation + total_gc)
    
    # Calculate Labor ratio and Labor for Price Breakdown
    labor_ratio = total_labor / subtotal if subtotal > 0 else 0
    mat_equip_ratio = total_material_equipment / subtotal if subtotal > 0 else 0
    
    # Labor del Price Breakdown = after_b2b_ohsms * labor_ratio (CFSE ya está en cascade)
    labor_for_price_breakdown = round2(after_b2b_ohsms * labor_ratio)
    
    # B2B OHSMS Labor = Labor (del Price Breakdown) × 4%
    b2b_ohsms_labor_amount = round2(labor_for_price_breakdown * (b2b_ohsms_labor_pct / 100))
    
    # Final total = cascaded total + B2B subcontractor (labor) + B2B OHSMS (labor)
    grand_total = round2(after_b2b_ohsms + b2b_subcontractor_amount + b2b_ohsms_labor_amount)
    
    # Labor with percentages = Labor proporción + B2B OHSMS Labor
    labor_with_percentages = round2(labor_for_price_breakdown + b2b_ohsms_labor_amount)
    # Material/Equipment with percentages = proporción mat/equip del cascade + B2B Subcontractor
    mat_equip_with_percentages = round2((after_b2b_ohsms * mat_equip_ratio) + b2b_subcontractor_amount)
    
    # Build summary table
    summary_data = [
        ['Concepto', 'Base', 'Monto'],
        ['Mano de Obra', '', f"${total_labor:,.2f}"],
        ['Subcontratistas', '', f"${total_subcontractors:,.2f}"],
        ['Materiales', '', f"${total_materials:,.2f}"],
        ['Equipos', '', f"${total_equipment:,.2f}"],
        ['Transporte', '', f"${total_transportation:,.2f}"],
        ['Condiciones Generales', '', f"${total_gc:,.2f}"],
        ['SUBTOTAL', '', f"${subtotal:,.2f}"],
    ]
    
    if profit_pct > 0:
        summary_data.append([f"Profit ({profit_pct}%)", f"${subtotal:,.2f}", f"${profit_amount:,.2f}"])
    
    if overhead_pct > 0:
        summary_data.append([f"Overhead ({overhead_pct}%)", f"${after_profit:,.2f}", f"${overhead_amount:,.2f}"])
    
    if cfse_pct > 0:
        summary_data.append([f"CFSE ({cfse_pct}%) - Solo M.O.", f"${total_labor:,.2f}", f"${cfse_amount:,.2f}"])
    
    if liability_pct > 0:
        summary_data.append([f"Liability ({liability_pct}%)", f"${combined_total:,.2f}", f"${liability_amount:,.2f}"])
    
    if municipal_patent_pct > 0:
        summary_data.append([f"Municipal Patent ({municipal_patent_pct}%)", f"${after_liability:,.2f}", f"${municipal_patent_amount:,.2f}"])
    
    if contingency_pct > 0:
        summary_data.append([f"Contingency ({contingency_pct}%)", f"${after_municipal_patent:,.2f}", f"${contingency_amount:,.2f}"])
    
    if b2b_ohsms_pct > 0:
        summary_data.append([f"B2B OHSMS Global ({b2b_ohsms_pct}%)", f"${b2b_ohsms_base:,.2f} (35%)", f"${b2b_ohsms_amount:,.2f}"])
    
    if b2b_ohsms_labor_pct > 0:
        summary_data.append([f"B2B OHSMS M.O. ({b2b_ohsms_labor_pct}%)", f"${total_labor:,.2f}", f"${b2b_ohsms_labor_amount:,.2f}"])
    
    if b2b_subcontractor_pct > 0 and total_subcontractor_labor > 0:
        summary_data.append([f"B2B Subcontratista ({b2b_subcontractor_pct}%)", f"${total_subcontractor_labor:,.2f}", f"${b2b_subcontractor_amount:,.2f}"])
    
    summary_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_COLOR),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, PRIMARY_COLOR),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fcfcfd')]),
    ])
    
    t = Table(summary_data, colWidths=[3*inch, 1.75*inch, 1.75*inch])
    t.setStyle(summary_style)
    elements.append(t)
    
    # Material/Equipment | Labor | Total breakdown
    elements.append(Spacer(1, 15))
    breakdown_header = [['Material/Equipment', 'Labor', 'Total']]
    breakdown_data = [[f"${mat_equip_with_percentages:,.2f}", f"${labor_with_percentages:,.2f}", f"${grand_total:,.2f}"]]
    
    breakdown_header_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
    ])
    
    breakdown_data_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_COLOR),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
    ])
    
    th = Table(breakdown_header, colWidths=[2.17*inch, 2.17*inch, 2.17*inch])
    th.setStyle(breakdown_header_style)
    elements.append(th)
    
    td = Table(breakdown_data, colWidths=[2.17*inch, 2.17*inch, 2.17*inch])
    td.setStyle(breakdown_data_style)
    elements.append(td)
    
    # Grand total with orange background
    elements.append(Spacer(1, 10))
    total_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
    ])
    
    total_data = [['GRAN TOTAL', f"${grand_total:,.2f}"]]
    tt = Table(total_data, colWidths=[4*inch, 2.5*inch])
    tt.setStyle(total_style)
    elements.append(tt)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"estimacion_{estimate_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/cost-estimates/{estimate_id}/export/excel")
async def export_cost_estimate_excel(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    user = await get_current_user(request, session_token)
    
    estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    ws_summary['A1'] = "Estimación de Costos"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Nombre: {estimate.get('estimate_name', '')}"
    ws_summary['A3'] = f"Proyecto: {estimate.get('project_name', '')}"
    ws_summary['A4'] = f"Fecha: {datetime.now(PUERTO_RICO_TZ).strftime('%d/%m/%Y')}"
    
    # Summary table starting at row 6
    summary_headers = ['Categoría', 'Total']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    summary_data = [
        ('Mano de Obra', estimate.get('total_labor', 0)),
        ('Subcontratistas', estimate.get('total_subcontractors', 0)),
        ('Materiales', estimate.get('total_materials', 0)),
        ('Equipos', estimate.get('total_equipment', 0)),
        ('Transporte', estimate.get('total_transportation', 0)),
        ('Condiciones Generales', estimate.get('total_general_conditions', 0)),
        ('Subtotal', estimate.get('subtotal', 0)),
        (f"Gastos Generales ({estimate.get('overhead_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('overhead_percentage', 0) / 100),
        (f"Utilidad ({estimate.get('profit_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('profit_percentage', 0) / 100),
        (f"Contingencia ({estimate.get('contingency_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('contingency_percentage', 0) / 100),
        (f"Impuestos ({estimate.get('tax_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('tax_percentage', 0) / 100),
        ('GRAN TOTAL', estimate.get('grand_total', 0)),
    ]
    
    for row_idx, (category, total) in enumerate(summary_data, 7):
        ws_summary.cell(row=row_idx, column=1, value=category).border = border
        cell = ws_summary.cell(row=row_idx, column=2, value=total)
        cell.border = border
        cell.number_format = currency_format
    
    # Make last row bold
    ws_summary.cell(row=18, column=1).font = Font(bold=True)
    ws_summary.cell(row=18, column=2).font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    
    # Labor Costs Sheet
    labor_costs = estimate.get('labor_costs', [])
    if labor_costs:
        ws_labor = wb.create_sheet("Mano de Obra")
        headers = ['Rol', 'Horas', 'Tarifa/Hora', 'Subtotal']
        for col, header in enumerate(headers, 1):
            cell = ws_labor.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(labor_costs, 2):
            ws_labor.cell(row=row_idx, column=1, value=item.get('role_name', '')).border = border
            ws_labor.cell(row=row_idx, column=2, value=item.get('hours', 0)).border = border
            cell = ws_labor.cell(row=row_idx, column=3, value=item.get('hourly_rate', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_labor.cell(row=row_idx, column=4, value=item.get('subtotal', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_labor.column_dimensions[get_column_letter(col)].width = 18
    
    # Subcontractors Sheet
    subcontractors = estimate.get('subcontractors', [])
    if subcontractors:
        ws_sub = wb.create_sheet("Subcontratistas")
        headers = ['Oficio', 'Descripción', 'Costo']
        for col, header in enumerate(headers, 1):
            cell = ws_sub.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(subcontractors, 2):
            ws_sub.cell(row=row_idx, column=1, value=item.get('trade', '')).border = border
            ws_sub.cell(row=row_idx, column=2, value=item.get('description', '')).border = border
            cell = ws_sub.cell(row=row_idx, column=3, value=item.get('cost', 0))
            cell.border = border
            cell.number_format = currency_format
        
        ws_sub.column_dimensions['A'].width = 20
        ws_sub.column_dimensions['B'].width = 40
        ws_sub.column_dimensions['C'].width = 18
    
    # Materials Sheet
    materials = estimate.get('materials', [])
    if materials:
        ws_mat = wb.create_sheet("Materiales")
        headers = ['Descripción', 'Cantidad', 'Precio Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_mat.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(materials, 2):
            ws_mat.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_mat.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_mat.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_mat.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_mat.column_dimensions[get_column_letter(col)].width = 18
    
    # Equipment Sheet
    equipment = estimate.get('equipment', [])
    if equipment:
        ws_eq = wb.create_sheet("Equipos")
        headers = ['Descripción', 'Cantidad', 'Días', 'Tarifa/Día', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_eq.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(equipment, 2):
            ws_eq.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_eq.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            ws_eq.cell(row=row_idx, column=3, value=item.get('days', 0)).border = border
            cell = ws_eq.cell(row=row_idx, column=4, value=item.get('rate_per_day', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_eq.cell(row=row_idx, column=5, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 6):
            ws_eq.column_dimensions[get_column_letter(col)].width = 18
    
    # Transportation Sheet
    transportation = estimate.get('transportation', [])
    if transportation:
        ws_trans = wb.create_sheet("Transporte")
        headers = ['Descripción', 'Cantidad', 'Costo Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(transportation, 2):
            ws_trans.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_trans.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_trans.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_trans.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_trans.column_dimensions[get_column_letter(col)].width = 18
    
    # General Conditions Sheet
    general_conditions = estimate.get('general_conditions', [])
    if general_conditions:
        ws_gc = wb.create_sheet("Condiciones Generales")
        headers = ['Descripción', 'Cantidad', 'Costo Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_gc.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(general_conditions, 2):
            ws_gc.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_gc.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_gc.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_gc.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_gc.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"estimacion_{estimate_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== REQUIRED DOCUMENTS ENDPOINTS ====================
@api_router.get("/required-documents")
async def get_required_documents(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    from_client = await db.required_documents.find({"direction": "from_client"}, {"_id": 0}).to_list(1000)
    to_client = await db.required_documents.find({"direction": "to_client"}, {"_id": 0}).to_list(1000)
    
    return {"from_client": from_client, "to_client": to_client}

@api_router.post("/required-documents/from-client")
async def add_document_from_client(
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    doc_id = f"doc_{uuid4().hex[:16]}"
    doc = {
        "document_id": doc_id,
        "document_name": data["document_name"],
        "direction": "from_client",
        "created_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.required_documents.insert_one(doc)
    # Return document without MongoDB _id field
    return {
        "document_id": doc_id,
        "document_name": data["document_name"],
        "direction": "from_client",
        "created_at": doc["created_at"]
    }

@api_router.post("/required-documents/to-client")
async def add_document_to_client(
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    doc_id = f"doc_{uuid4().hex[:16]}"
    doc = {
        "document_id": doc_id,
        "document_name": data["document_name"],
        "direction": "to_client",
        "created_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.required_documents.insert_one(doc)
    # Return document without MongoDB _id field
    return {
        "document_id": doc_id,
        "document_name": data["document_name"],
        "direction": "to_client",
        "created_at": doc["created_at"]
    }

@api_router.delete("/required-documents/{doc_id}")
async def delete_required_document(
    doc_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    await db.required_documents.delete_one({"document_id": doc_id})
    return {"message": "Documento eliminado"}

# ==================== PROJECT DOCUMENT STATUS ====================
@api_router.get("/projects/{project_id}/document-status")
async def get_project_document_status(
    project_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    status_docs = await db.project_document_status.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Convert to dict with document_id as key
    status_dict = {doc["document_id"]: doc["status"] for doc in status_docs}
    
    return status_dict

@api_router.post("/projects/{project_id}/document-status")
async def update_project_document_status(
    project_id: str,
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    # Update or insert status
    await db.project_document_status.update_one(
        {
            "project_id": project_id,
            "document_id": data["document_id"]
        },
        {
            "$set": {
                "project_id": project_id,
                "document_id": data["document_id"],
                "direction": data["direction"],
                "status": data["status"],
                "updated_at": datetime.now(PUERTO_RICO_TZ).isoformat(),
                "updated_by": user.user_id
            }
        },
        upsert=True
    )
    
    return {"message": "Estado actualizado"}

# ==================== NOMENCLATURES ENDPOINTS ====================
@api_router.get("/nomenclatures")
async def get_nomenclatures(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    
    nomenclatures = await db.nomenclatures.find({}, {"_id": 0}).to_list(1000)
    return nomenclatures

@api_router.post("/nomenclatures")
async def create_nomenclature(
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    nomenclature_id = f"nom_{uuid4().hex[:16]}"
    current_year = datetime.now(PUERTO_RICO_TZ).year
    
    nom = {
        "nomenclature_id": nomenclature_id,
        "name": data["name"],
        "prefix": data["prefix"].upper(),
        "department_number": data.get("department_number", ""),
        "starting_number": data.get("starting_number", 1),
        "current_number": 1,
        "current_year": current_year,
        "created_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.nomenclatures.insert_one(nom)
    # Return without _id
    return {
        "nomenclature_id": nomenclature_id,
        "name": nom["name"],
        "prefix": nom["prefix"],
        "department_number": nom["department_number"],
        "starting_number": nom["starting_number"],
        "current_number": nom["current_number"],
        "current_year": nom["current_year"],
        "created_at": nom["created_at"]
    }

@api_router.put("/nomenclatures/{nomenclature_id}")
async def update_nomenclature(
    nomenclature_id: str,
    data: dict,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    update_data = {
        "name": data["name"],
        "prefix": data["prefix"].upper(),
        "department_number": data.get("department_number", ""),
        "starting_number": data.get("starting_number", 1)
    }
    
    await db.nomenclatures.update_one({"nomenclature_id": nomenclature_id}, {"$set": update_data})
    return {"message": "Nomenclatura actualizada"}

@api_router.delete("/nomenclatures/{nomenclature_id}")
async def delete_nomenclature(
    nomenclature_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    await db.nomenclatures.delete_one({"nomenclature_id": nomenclature_id})
    return {"message": "Nomenclatura eliminada"}

@api_router.get("/nomenclatures/next-number/{nomenclature_id}")
async def get_next_number(
    nomenclature_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    nomenclature = await db.nomenclatures.find_one({"nomenclature_id": nomenclature_id}, {"_id": 0})
    if not nomenclature:
        raise HTTPException(status_code=404, detail="Nomenclatura no encontrada")
    
    current_year = datetime.now(PUERTO_RICO_TZ).year
    
    # Si cambió el año, resetear el contador
    if nomenclature.get("current_year") != current_year:
        await db.nomenclatures.update_one(
            {"nomenclature_id": nomenclature_id},
            {"$set": {"current_number": 1, "current_year": current_year}}
        )
        current_number = 1
    else:
        current_number = nomenclature.get("current_number", 1)
    
    # Incrementar para el próximo
    await db.nomenclatures.update_one(
        {"nomenclature_id": nomenclature_id},
        {"$set": {"current_number": current_number + 1}}
    )
    
    # Generar número completo: PREFIX-YEAR-DEPARTMENT-CURRENT
    department = nomenclature.get('department_number', '')
    if department:
        full_number = f"{nomenclature['prefix']}-{current_year}-{department}-{current_number}"
    else:
        full_number = f"{nomenclature['prefix']}-{current_year}-{current_number}"
    
    return {"number": full_number, "nomenclature": nomenclature}

# ==================== PAYROLL SETTINGS ====================
@api_router.get("/payroll-settings")
async def get_payroll_settings(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    settings = await db.payroll_settings.find_one({}, {"_id": 0})
    return settings or {
        "hacienda_percent": 0,
        "social_security_percent": 6.2,
        "medicare_percent": 1.45,
        "contractor_percent": 10
    }

@api_router.put("/payroll-settings")
async def update_payroll_settings(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    await db.payroll_settings.update_one({}, {"$set": data}, upsert=True)
    return {"message": "Configuración guardada"}

@api_router.post("/payroll/process")
async def process_payroll(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="Solo RRHH o administradores")
    
    payroll_id = str(uuid4())
    period_start = data.get("period_start")
    period_end = data.get("period_end")
    employees = data.get("employees", [])
    
    payroll_run = {
        "id": payroll_id,
        "period_start": period_start,
        "period_end": period_end,
        "processed_by": user.user_id,
        "processed_by_name": user.name or "Usuario",
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "status": "completed",
        "employees": employees,
        "totals": data.get("totals", {})
    }
    
    await db.payroll_runs.insert_one(payroll_run)
    
    # Generar talonarios para cada empleado
    for emp in employees:
        pay_stub = {
            "id": str(uuid4()),
            "payroll_id": payroll_id,
            "employee_id": emp.get("user_id"),
            "employee_name": emp.get("name"),
            "period_start": period_start,
            "period_end": period_end,
            "hours_worked": emp.get("hours", 0),
            "hourly_rate": emp.get("rate", 0),
            "gross_pay": emp.get("grossPay", 0),
            "deductions": {
                "hacienda": emp.get("hacienda", 0),
                "social_security": emp.get("ss", 0),
                "medicare": emp.get("medicare", 0),
                "other": emp.get("otherDeductions", 0)
            },
            "total_deductions": emp.get("deductions", 0),
            "net_pay": emp.get("netPay", 0),
            "payment_method": emp.get("payment_method", "check"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_contractor": emp.get("is_contractor", False)
        }
        await db.pay_stubs.insert_one(pay_stub)
    
    await log_audit(user.user_id, user.name, "create", "payroll", payroll_id, f"Nómina {period_start} - {period_end}", {"total": data.get("totals", {}).get("net", 0), "employees": len(employees)})
    return {"message": "Nómina procesada y talonarios generados", "id": payroll_id, "stubs_generated": len(employees)}

@api_router.get("/pay-stubs/my")
async def get_my_pay_stubs(request: Request, session_token: Optional[str] = Cookie(None)):
    """Obtener mis talonarios de pago"""
    user = await get_current_user(request, session_token)
    stubs = await db.pay_stubs.find({"employee_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return stubs

@api_router.get("/pay-stubs/{employee_id}")
async def get_employee_pay_stubs(employee_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Obtener talonarios de un empleado (admin/rrhh)"""
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value] and user.user_id != employee_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    stubs = await db.pay_stubs.find({"employee_id": employee_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return stubs

@api_router.get("/pay-stubs/all")
async def get_all_pay_stubs(request: Request, session_token: Optional[str] = Cookie(None)):
    """Obtener todos los talonarios (solo super_admin y rrhh)"""
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="No autorizado")
    stubs = await db.pay_stubs.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return stubs

@api_router.delete("/pay-stubs/{stub_id}")
async def delete_pay_stub(stub_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    """Eliminar un talonario (solo super_admin y rrhh)"""
    user = await get_current_user(request, session_token)
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Verificar que existe el talonario
    stub = await db.pay_stubs.find_one({"id": stub_id})
    if not stub:
        raise HTTPException(status_code=404, detail="Talonario no encontrado")
    
    await db.pay_stubs.delete_one({"id": stub_id})
    await log_audit(user.user_id, user.name, "delete", "pay_stub", stub_id, 
                    f"Talonario de {stub.get('employee_name', 'N/A')}", 
                    {"period": f"{stub.get('period_start')} - {stub.get('period_end')}"})
    return {"message": "Talonario eliminado exitosamente"}

@api_router.get("/payroll/history")
async def get_payroll_history(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    runs = await db.payroll_runs.find({}, {"_id": 0}).sort("processed_at", -1).to_list(100)
    return runs

@api_router.post("/payroll/nacha")
async def generate_nacha(data: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    """Genera archivo NACHA para depósitos directos"""
    from fastapi.responses import Response
    user = await get_current_user(request, session_token)
    
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    employees_data = data.get("employees", [])
    
    now = datetime.now(timezone.utc)
    file_date = now.strftime("%y%m%d")
    file_time = now.strftime("%H%M")
    
    # Configuración bancaria (debería venir de settings)
    dest_routing = data.get("dest_routing", "121000248")  # Routing del banco destino
    company_id = data.get("company_id", "2313801042")     # ID de la empresa (10 dígitos)
    origin_name = (company.get("company_name", "COMPANY") + " " * 23)[:23]
    dest_name = (data.get("bank_name", "BANK OF PR") + " " * 23)[:23]
    
    lines = []
    
    # Record 1: File Header
    # 101 + dest_routing(9) + origin(10) + date(6) + time(4) + A094101 + dest_name(23) + origin_name(23) + ref(8)
    file_header = f"101 {dest_routing} {company_id}{file_date}{file_time}A094101{origin_name}{dest_name}        "
    lines.append(file_header)
    
    # Record 5: Batch Header  
    # 5220 + company_name(16) + discretionary(20) + company_id(10) + PPD + desc(10) + date(6) + spaces(3) + 1 + routing(8) + batch(7)
    batch_header = f"5220{origin_name[:16]:<16}{' '*20}{company_id:<10}PPD PAYROLL   {file_date}   1{dest_routing[:8]}0000001"
    lines.append(batch_header)
    
    # Record 6: Entry Detail
    entry_count = 0
    total_amount = 0
    entry_hash = 0
    
    for emp in employees_data:
        routing = emp.get("routing_number", dest_routing)
        account = emp.get("bank_account", "")
        amount = int(emp.get("net_pay", 0) * 100)  # Centavos
        name = (emp.get("employee_name", "")[:22] + " " * 22)[:22]
        
        if amount > 0:
            entry_count += 1
            total_amount += amount
            routing_8 = routing[:8] if routing else dest_routing[:8]
            entry_hash += int(routing_8) if routing_8.isdigit() else 0
            
            # 6 + trans_code(2) + routing(8) + check(1) + account(17) + amount(10) + name(22) + trace(7)
            acc_type = "22" if emp.get("account_type", "checking") == "checking" else "32"
            check_digit = routing[8] if len(routing) > 8 else "0"
            entry = f"6{acc_type}{routing_8}{check_digit}{account:<17}{amount:010d}{name}{entry_count:07d}"
            lines.append(entry)
    
    # Record 8: Batch Control
    # 8220 + count(6) + hash(10) + debit(12) + credit(12) + company_id(10) + batch(7)
    entry_hash_10 = entry_hash % 10000000000
    batch_control = f"8220{entry_count:06d}{entry_hash_10:010d}{0:012d}{total_amount:012d}{company_id:<10}{' '*25}{dest_routing[:8]}0000001"
    lines.append(batch_control)
    
    # Record 9: File Control
    # 9 + batch_count(6) + block_count(6) + entry_count(8) + hash(10) + debit(12) + credit(12) + reserved(39)
    block_count = (len(lines) + 2) // 10 + 1
    file_control = f"9{1:06d}{block_count:06d}{entry_count:08d}{entry_hash_10:010d}{0:012d}{total_amount:012d}{' '*39}"
    lines.append(file_control)
    
    # Padding con 9s hasta múltiplo de 10
    while len(lines) % 10 != 0:
        lines.append("9" * 94)
    
    nacha_content = "\n".join(lines)
    
    return Response(
        content=nacha_content,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename=NACHA_{file_date}.txt"}
    )

# ==================== HUMAN RESOURCES / EMPLOYEE DOCUMENTS ====================
EMPLOYEE_DOCS_DIR = Path("/app/uploads/employee_docs")
EMPLOYEE_DOCS_DIR.mkdir(parents=True, exist_ok=True)

@api_router.get("/employees")
async def get_employees(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    # Filter out hidden users and sort alphabetically
    employees = await db.users.find({"hidden": {"$ne": True}}, {"_id": 0, "password": 0}).sort("name", 1).to_list(1000)
    
    # Attach profiles to employees
    for emp in employees:
        profile = await db.employee_profiles.find_one({"user_id": emp["user_id"]}, {"_id": 0})
        emp["profile"] = profile or {}
    
    return employees

@api_router.get("/employees/{employee_id}/profile")
async def get_employee_profile(employee_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    profile = await db.employee_profiles.find_one({"user_id": employee_id}, {"_id": 0})
    return profile or {}

@api_router.put("/employees/{employee_id}/profile")
async def update_employee_profile(
    employee_id: str,
    profile_data: EmployeeProfile,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    import logging
    logger = logging.getLogger("employee_profile")
    
    try:
        logger.info(f"=== UPDATE EMPLOYEE PROFILE START ===")
        logger.info(f"Employee ID: {employee_id}")
        logger.info(f"Profile data received: {profile_data}")
        
        user = await get_current_user(request, session_token)
        logger.info(f"User requesting update: {user.name} (role: {user.role})")
        
        if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
            logger.warning(f"Access denied for user {user.name} with role {user.role}")
            raise HTTPException(status_code=403, detail="Solo RRHH o administradores")
        
        profile_dict = profile_data.model_dump()
        logger.info(f"Profile dict after model_dump: {profile_dict}")
        
        profile_dict["user_id"] = employee_id
        profile_dict["updated_at"] = datetime.now(PUERTO_RICO_TZ).isoformat()
        
        existing = await db.employee_profiles.find_one({"user_id": employee_id})
        logger.info(f"Existing profile found: {existing is not None}")
        
        if existing:
            logger.info(f"Updating existing profile for user_id: {employee_id}")
            result = await db.employee_profiles.update_one({"user_id": employee_id}, {"$set": profile_dict})
            logger.info(f"Update result - matched: {result.matched_count}, modified: {result.modified_count}")
        else:
            profile_dict["employee_id"] = f"emp_{uuid4().hex[:16]}"
            profile_dict["created_at"] = datetime.now(PUERTO_RICO_TZ).isoformat()
            logger.info(f"Creating new profile with employee_id: {profile_dict['employee_id']}")
            result = await db.employee_profiles.insert_one(profile_dict)
            logger.info(f"Insert result - inserted_id: {result.inserted_id}")
        
        logger.info(f"=== UPDATE EMPLOYEE PROFILE SUCCESS ===")
        return {"message": "Perfil actualizado"}
        
    except HTTPException as he:
        logger.error(f"HTTP Exception: {he.detail}")
        raise he
    except Exception as e:
        logger.error(f"=== UPDATE EMPLOYEE PROFILE ERROR ===")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error message: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al guardar perfil: {str(e)}")

# Endpoint para que empleados actualicen su propio perfil (sin campos de salario)
class EmployeeProfileSelfUpdate(BaseModel):
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    zipcode: Optional[str] = None
    country: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    marital_status: Optional[str] = None
    nationality: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    emergency_contact_relationship: Optional[str] = None

@api_router.put("/employees/{employee_id}/profile/self")
async def update_own_profile(
    employee_id: str,
    profile_data: EmployeeProfileSelfUpdate,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Empleados pueden actualizar su propia información personal (excepto salarios)"""
    user = await get_current_user(request, session_token)
    
    # Solo puede editar su propio perfil
    if user.user_id != employee_id:
        raise HTTPException(status_code=403, detail="Solo puedes editar tu propio perfil")
    
    # Obtener perfil existente
    existing = await db.employee_profiles.find_one({"user_id": employee_id})
    
    # Campos permitidos para auto-actualización
    update_data = {k: v for k, v in profile_data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(PUERTO_RICO_TZ).isoformat()
    
    if existing:
        await db.employee_profiles.update_one({"user_id": employee_id}, {"$set": update_data})
    else:
        update_data["user_id"] = employee_id
        update_data["employee_id"] = f"emp_{uuid4().hex[:16]}"
        update_data["created_at"] = datetime.now(PUERTO_RICO_TZ).isoformat()
        await db.employee_profiles.insert_one(update_data)
    
    return {"message": "Perfil actualizado correctamente"}

# Endpoint para obtener tareas asignadas al usuario actual
@api_router.get("/my-tasks")
async def get_my_tasks(request: Request, session_token: Optional[str] = Cookie(None)):
    """Obtener tareas asignadas al usuario actual"""
    user = await get_current_user(request, session_token)
    
    # Obtener tareas asignadas al usuario
    tasks = await db.tasks.find({"assigned_to": user.user_id}, {"_id": 0}).to_list(1000)
    
    # Añadir nombre del proyecto a cada tarea
    for task in tasks:
        project = await db.projects.find_one({"project_id": task.get("project_id")}, {"_id": 0, "name": 1})
        task["project_name"] = project.get("name", "Sin proyecto") if project else "Sin proyecto"
    
    # Ordenar por fecha de vencimiento y prioridad
    priority_order = {"high": 0, "medium": 1, "low": 2}
    tasks.sort(key=lambda x: (
        x.get("status") == "done",  # Completadas al final
        x.get("due_date") or "9999-99-99",  # Por fecha de vencimiento
        priority_order.get(x.get("priority", "medium"), 1)  # Por prioridad
    ))
    
    return tasks

# Endpoint para actualizar solo el estado de una tarea (para empleados)
class TaskStatusUpdate(BaseModel):
    status: str

@api_router.put("/tasks/{task_id}/status")
async def update_task_status(
    task_id: str, 
    status_data: TaskStatusUpdate, 
    request: Request, 
    session_token: Optional[str] = Cookie(None)
):
    """Empleados pueden actualizar el estado de sus tareas asignadas"""
    user = await get_current_user(request, session_token)
    
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    # Verificar que la tarea está asignada al usuario (o es admin/PM)
    if task.get("assigned_to") != user.user_id and user.role not in [UserRole.SUPER_ADMIN.value, UserRole.PROJECT_MANAGER.value]:
        raise HTTPException(status_code=403, detail="No tienes permiso para actualizar esta tarea")
    
    old_status = task.get("status")
    new_status = status_data.status
    now = datetime.now(timezone.utc).isoformat()
    
    await db.tasks.update_one(
        {"task_id": task_id}, 
        {"$set": {"status": new_status, "updated_at": now}}
    )
    
    # Enviar notificación si la tarea se completó
    if new_status == "done" and old_status != "done":
        project = await db.projects.find_one({"project_id": task["project_id"]}, {"_id": 0})
        if project and project.get("created_by"):
            # Notificación en app
            notification_doc = {
                "notification_id": f"notif_{uuid4().hex[:12]}",
                "user_id": project["created_by"],
                "type": "task_completed",
                "message": f"{user.name} completó la tarea: {task['title']}",
                "read": False,
                "timestamp": now,
                "related_id": task_id
            }
            await db.notifications.insert_one(notification_doc)
            
            # Email al creador del proyecto
            owner = await db.users.find_one({"user_id": project["created_by"]}, {"_id": 0})
            if owner:
                html, text = get_task_completed_email(
                    owner["name"],
                    task["title"],
                    project["name"],
                    user.name
                )
                await send_email(
                    owner["email"],
                    f"Tarea completada: {task['title']}",
                    html,
                    text
                )
    
    # Notificar cambio de estado (si no es completada)
    elif new_status != old_status:
        project = await db.projects.find_one({"project_id": task["project_id"]}, {"_id": 0})
        if project and project.get("created_by") and project["created_by"] != user.user_id:
            status_labels = {
                "todo": "Pendiente",
                "in_progress": "En Progreso",
                "review": "En Revisión",
                "done": "Completada"
            }
            notification_doc = {
                "notification_id": f"notif_{uuid4().hex[:12]}",
                "user_id": project["created_by"],
                "type": "task_status_changed",
                "message": f"{user.name} cambió '{task['title']}' a {status_labels.get(new_status, new_status)}",
                "read": False,
                "timestamp": now,
                "related_id": task_id
            }
            await db.notifications.insert_one(notification_doc)
    
    return {"message": "Estado actualizado", "status": new_status}

@api_router.get("/employees/{employee_id}/documents")
async def get_employee_documents(employee_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user(request, session_token)
    docs = await db.employee_documents.find({"employee_id": employee_id}, {"_id": 0}).to_list(1000)
    return docs

@api_router.post("/employees/{employee_id}/documents")
async def upload_employee_document(
    employee_id: str,
    file: UploadFile = File(...),
    document_type: str = Query(...),
    request: Request = None,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="Solo RRHH o administradores")
    
    # Create employee folder
    employee_folder = EMPLOYEE_DOCS_DIR / employee_id
    employee_folder.mkdir(parents=True, exist_ok=True)
    
    # Save file
    doc_id = f"edoc_{uuid4().hex[:16]}"
    file_ext = Path(file.filename).suffix
    filename = f"{doc_id}{file_ext}"
    file_path = employee_folder / filename
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    doc = {
        "doc_id": doc_id,
        "employee_id": employee_id,
        "document_type": document_type,
        "original_filename": file.filename,
        "stored_filename": filename,
        "file_path": str(file_path),
        "file_url": f"/uploads/employee_docs/{employee_id}/{filename}",
        "uploaded_by": user.user_id,
        "uploaded_at": datetime.now(PUERTO_RICO_TZ).isoformat()
    }
    
    await db.employee_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/employees/{employee_id}/documents/{doc_id}")
async def delete_employee_document(
    employee_id: str,
    doc_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    if user.role not in [UserRole.SUPER_ADMIN.value, UserRole.RRHH.value]:
        raise HTTPException(status_code=403, detail="Solo RRHH o administradores")
    
    doc = await db.employee_documents.find_one({"doc_id": doc_id})
    if doc and Path(doc["file_path"]).exists():
        Path(doc["file_path"]).unlink()
    
    await db.employee_documents.delete_one({"doc_id": doc_id})
    return {"message": "Documento eliminado"}

@api_router.get("/employees/{employee_id}/documents/{doc_id}/download")
async def download_employee_document(
    employee_id: str,
    doc_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    user = await get_current_user(request, session_token)
    
    doc = await db.employee_documents.find_one({"doc_id": doc_id, "employee_id": employee_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    file_path = Path(doc["file_path"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Determine media type
    suffix = file_path.suffix.lower()
    media_types = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xls': 'application/vnd.ms-excel',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    media_type = media_types.get(suffix, 'application/octet-stream')
    
    return FileResponse(
        file_path,
        media_type=media_type,
        filename=doc.get("original_filename", file_path.name)
    )

@api_router.get("/employees/{employee_id}/documents/{doc_id}/preview")
async def preview_employee_document(
    employee_id: str,
    doc_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Preview document inline without forcing download"""
    user = await get_current_user(request, session_token)
    
    doc = await db.employee_documents.find_one({"doc_id": doc_id, "employee_id": employee_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    file_path = Path(doc["file_path"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Determine media type
    suffix = file_path.suffix.lower()
    media_types = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp',
        '.svg': 'image/svg+xml'
    }
    media_type = media_types.get(suffix, 'application/octet-stream')
    
    # Return file without Content-Disposition attachment header (inline display)
    return FileResponse(
        file_path,
        media_type=media_type
        # No filename parameter = inline display instead of download
    )

app.include_router(api_router)
app.include_router(accounting_router)

# Middleware para prevenir caché en respuestas de API
@app.middleware("http")
async def add_no_cache_headers(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()