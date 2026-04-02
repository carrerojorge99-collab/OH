# Backend: Cost Estimate Excel Export
# Extracted from backend/server.py lines 12580-12814

@api_router.get("/cost-estimates/{estimate_id}/export/excel")
async def export_cost_estimate_excel(
    estimate_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    user = await get_current_user(request, session_token)
    
    estimate = await db.cost_estimates.find_one({"estimate_id": estimate_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimación no encontrada")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    ws_summary['A1'] = "Estimación de Costos"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Nombre: {estimate.get('estimate_name', '')}"
    ws_summary['A3'] = f"Proyecto: {estimate.get('project_name', '')}"
    ws_summary['A4'] = f"Fecha: {datetime.now(PUERTO_RICO_TZ).strftime('%d/%m/%Y')}"
    
    # Summary table starting at row 6
    summary_headers = ['Categoría', 'Total']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    summary_data = [
        ('Mano de Obra', estimate.get('total_labor', 0)),
        ('Subcontratistas', estimate.get('total_subcontractors', 0)),
        ('Materiales', estimate.get('total_materials', 0)),
        ('Equipos', estimate.get('total_equipment', 0)),
        ('Transporte', estimate.get('total_transportation', 0)),
        ('Condiciones Generales', estimate.get('total_general_conditions', 0)),
        ('Subtotal', estimate.get('subtotal', 0)),
        (f"Gastos Generales ({estimate.get('overhead_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('overhead_percentage', 0) / 100),
        (f"Utilidad ({estimate.get('profit_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('profit_percentage', 0) / 100),
        (f"Contingencia ({estimate.get('contingency_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('contingency_percentage', 0) / 100),
        (f"Impuestos ({estimate.get('tax_percentage', 0)}%)", estimate.get('subtotal', 0) * estimate.get('tax_percentage', 0) / 100),
        ('GRAN TOTAL', estimate.get('grand_total', 0)),
    ]
    
    for row_idx, (category, total) in enumerate(summary_data, 7):
        ws_summary.cell(row=row_idx, column=1, value=category).border = border
        cell = ws_summary.cell(row=row_idx, column=2, value=total)
        cell.border = border
        cell.number_format = currency_format
    
    # Make last row bold
    ws_summary.cell(row=18, column=1).font = Font(bold=True)
    ws_summary.cell(row=18, column=2).font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    
    # Labor Costs Sheet
    labor_costs = estimate.get('labor_costs', [])
    if labor_costs:
        ws_labor = wb.create_sheet("Mano de Obra")
        headers = ['Rol', 'Horas', 'Tarifa/Hora', 'Subtotal']
        for col, header in enumerate(headers, 1):
            cell = ws_labor.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(labor_costs, 2):
            ws_labor.cell(row=row_idx, column=1, value=item.get('role_name', '')).border = border
            ws_labor.cell(row=row_idx, column=2, value=item.get('hours', 0)).border = border
            cell = ws_labor.cell(row=row_idx, column=3, value=item.get('hourly_rate', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_labor.cell(row=row_idx, column=4, value=item.get('subtotal', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_labor.column_dimensions[get_column_letter(col)].width = 18
    
    # Subcontractors Sheet
    subcontractors = estimate.get('subcontractors', [])
    if subcontractors:
        ws_sub = wb.create_sheet("Subcontratistas")
        headers = ['Oficio', 'Descripción', 'Costo']
        for col, header in enumerate(headers, 1):
            cell = ws_sub.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(subcontractors, 2):
            ws_sub.cell(row=row_idx, column=1, value=item.get('trade', '')).border = border
            ws_sub.cell(row=row_idx, column=2, value=item.get('description', '')).border = border
            cell = ws_sub.cell(row=row_idx, column=3, value=item.get('cost', 0))
            cell.border = border
            cell.number_format = currency_format
        
        ws_sub.column_dimensions['A'].width = 20
        ws_sub.column_dimensions['B'].width = 40
        ws_sub.column_dimensions['C'].width = 18
    
    # Materials Sheet
    materials = estimate.get('materials', [])
    if materials:
        ws_mat = wb.create_sheet("Materiales")
        headers = ['Descripción', 'Cantidad', 'Precio Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_mat.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(materials, 2):
            ws_mat.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_mat.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_mat.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_mat.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_mat.column_dimensions[get_column_letter(col)].width = 18
    
    # Equipment Sheet
    equipment = estimate.get('equipment', [])
    if equipment:
        ws_eq = wb.create_sheet("Equipos")
        headers = ['Descripción', 'Cantidad', 'Días', 'Tarifa/Día', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_eq.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(equipment, 2):
            ws_eq.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_eq.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            ws_eq.cell(row=row_idx, column=3, value=item.get('days', 0)).border = border
            cell = ws_eq.cell(row=row_idx, column=4, value=item.get('rate_per_day', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_eq.cell(row=row_idx, column=5, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 6):
            ws_eq.column_dimensions[get_column_letter(col)].width = 18
    
    # Transportation Sheet
    transportation = estimate.get('transportation', [])
    if transportation:
        ws_trans = wb.create_sheet("Transporte")
        headers = ['Descripción', 'Cantidad', 'Costo Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(transportation, 2):
            ws_trans.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_trans.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_trans.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_trans.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_trans.column_dimensions[get_column_letter(col)].width = 18
    
    # General Conditions Sheet
    general_conditions = estimate.get('general_conditions', [])
    if general_conditions:
        ws_gc = wb.create_sheet("Condiciones Generales")
        headers = ['Descripción', 'Cantidad', 'Costo Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_gc.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(general_conditions, 2):
            ws_gc.cell(row=row_idx, column=1, value=item.get('description', '')).border = border
            ws_gc.cell(row=row_idx, column=2, value=item.get('quantity', 0)).border = border
            cell = ws_gc.cell(row=row_idx, column=3, value=item.get('unit_cost', 0))
            cell.border = border
            cell.number_format = currency_format
            cell = ws_gc.cell(row=row_idx, column=4, value=item.get('total', 0))
            cell.border = border
            cell.number_format = currency_format
        
        for col in range(1, 5):
            ws_gc.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"estimacion_{estimate_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== REQUIRED DOCUMENTS ENDPOINTS ====================
